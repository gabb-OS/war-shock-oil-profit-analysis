"""
01_data_pipeline.py  (fix: normalizzazione unità EUR/1000L → EUR/L)
====================
Raccolta e preparazione dati.

Produciamo tre serie settimanali (W-MON, dal 2019):
  - Brent crude EUR/barile       (yfinance BZ=F + EURUSD=X)
  - Prezzi pompa Italia senza tasse (EU Weekly Oil Bulletin, foglio "wo taxes")
  - Futures wholesale europei    (Eurobob ARA benzina, Gas Oil ICE diesel)

NOTA UNITÀ: l'EU Oil Bulletin pubblica i prezzi in EUR/1000L.
I futures Eurobob/GasOil sono convertiti in EUR/L.
Per coerenza, normalizziamo i prezzi pompa in EUR/L (/1000) se mediana > 10.

Output:
  data/brent_weekly_eur.csv
  data/prezzi_pompa_italia.csv
  data/dataset_merged.csv
  data/dataset_merged_with_futures.csv
  data/missing_weeks.csv / .json
  plots/01a_brent.png  |  01b_benzina.png  |  01c_diesel.png
"""

import os, json, warnings
from datetime import date, timedelta

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import requests
import openpyxl as _opxl
import yfinance as yf

warnings.filterwarnings("ignore")
os.makedirs("data",  exist_ok=True)
os.makedirs("plots", exist_ok=True)

TODAY_END  = (date.today() + timedelta(days=7)).strftime("%Y-%m-%d")
DATA_START = "2019-01-01"
DPI        = 180

WAR_EVENTS = {
    "Invasione Ucraina":   ("2022-02-24", "#e74c3c"),
    "Guerra Iran-Israele": ("2025-06-13", "#e67e22"),
    "Chiusura Hormuz":     ("2026-02-28", "#8e44ad"),
}

# Conversioni futures (USD/tonnellata → EUR/litro)
DENSITY_BENZ = 0.74   # kg/L
DENSITY_DIES = 0.84   # kg/L
L_PER_T_BENZ = 1000.0 / DENSITY_BENZ   # ≈ 1351 L/t
L_PER_T_DIES = 1000.0 / DENSITY_DIES   # ≈ 1190 L/t

missing_log: dict = {}


# ─────────────────────────────────────────────────────────────────────────────
# 1. EUR/USD settimanale
# ─────────────────────────────────────────────────────────────────────────────
print("Scarico EUR/USD (yfinance)...")
try:
    raw = yf.download("EURUSD=X", start=DATA_START, end=TODAY_END, progress=False)
    if raw.empty:
        raise ValueError("risposta vuota")
    eurusd_w = raw["Close"].squeeze().rename("eurusd").resample("W-MON").mean()
    eurusd_w.index = pd.to_datetime(eurusd_w.index)
    miss = pd.date_range(eurusd_w.index[0], eurusd_w.index[-1], freq="W-MON")
    miss = miss[~miss.isin(eurusd_w.dropna().index)]
    if len(miss):
        missing_log["eurusd"] = [str(d.date()) for d in miss]
    print(f"  {eurusd_w.dropna().shape[0]} settimane")
except Exception as exc:
    print(f"  ERRORE: {exc}")
    eurusd_w = None
    missing_log["eurusd"] = ["download_fallito"]


# ─────────────────────────────────────────────────────────────────────────────
# 2. Brent settimanale in EUR/barile
# ─────────────────────────────────────────────────────────────────────────────
print("\nScarico Brent (yfinance BZ=F)...")
try:
    raw = yf.download("BZ=F", start=DATA_START, end=TODAY_END, progress=False)
    if raw.empty:
        raise ValueError("risposta vuota")
    brent_usd = raw["Close"].squeeze().rename("brent_usd")
    brent_usd.index = pd.to_datetime(brent_usd.index)
    brent_w = brent_usd.resample("W-MON").mean().to_frame()
    brent_w["brent_7d_usd"] = brent_w["brent_usd"]

    if eurusd_w is not None:
        brent_w = brent_w.join(eurusd_w, how="left")
        brent_w["eurusd"] = brent_w["eurusd"].ffill()
        brent_w["brent_eur"]    = brent_w["brent_usd"]    / brent_w["eurusd"]
        brent_w["brent_7d_eur"] = brent_w["brent_7d_usd"] / brent_w["eurusd"]
    else:
        brent_w["brent_eur"] = brent_w["brent_7d_eur"] = brent_w["brent_usd"]
        print("  Brent rimane in USD (EUR/USD non disponibile)")

    for col in ["brent_eur", "brent_7d_eur"]:
        n = brent_w[col].isna().sum()
        if n:
            brent_w[col] = brent_w[col].interpolate(method="time", limit=2)
            missing_log.setdefault("brent", []).append(f"{n} settimane mancanti interpolate")

    brent_w["log_brent"] = np.log(brent_w["brent_7d_eur"])
    brent_w.to_csv("data/brent_weekly_eur.csv")
    print(f"  {brent_w.dropna().shape[0]} settimane | "
          f"{brent_w.index[0].date()} – {brent_w.index[-1].date()}")

except Exception as exc:
    raise SystemExit(f"ERRORE CRITICO Brent: {exc}")


# ─────────────────────────────────────────────────────────────────────────────
# 3. Prezzi pompa Italia senza tasse (EU Weekly Oil Bulletin)
# ─────────────────────────────────────────────────────────────────────────────
print("\nScarico prezzi pompa (EU Weekly Oil Bulletin)...")

EU_SOURCES = [
    ("storico 2005-presente",
     "https://energy.ec.europa.eu/document/download/"
     "906e60ca-8b6a-44e7-8589-652854d2fd3f_en?"
     "filename=Weekly_Oil_Bulletin_Prices_History_maticni_4web.xlsx",
     "data/eu_oil_bulletin_history.xlsx", True),
    ("settimanale senza tasse",
     "https://energy.ec.europa.eu/document/download/"
     "78311f92-68f8-4b82-b5cf-1293beeaae77_en?"
     "filename=Weekly+Oil+Bulletin+Weekly+prices+without+taxes+-+2024-02-19.xlsx",
     "data/eu_oil_bulletin_notax.xlsx", True),
    ("settimanale con tasse (fallback)",
     "https://energy.ec.europa.eu/document/download/"
     "264c2d0f-f161-4ea3-a777-78faae59bea0_en?"
     "filename=Weekly+Oil+Bulletin+Weekly+prices+with+Taxes+-+2024-02-19.xlsx",
     "data/eu_oil_bulletin_tax.xlsx", False),
]


def _sheet_names(path):
    wb = _opxl.load_workbook(path, read_only=True, data_only=True)
    ns = wb.sheetnames; wb.close(); return ns


def _read_sheet(path, sheet):
    wb = _opxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if not rows:
        return pd.DataFrame()
    hi = next((i for i, r in enumerate(rows) if any(v is not None for v in r)), 0)
    hdrs = [str(h).strip() if h is not None else f"_c{i}"
            for i, h in enumerate(rows[hi])]
    df = pd.DataFrame(rows[hi+1:], columns=hdrs)
    ic = hdrs[0]
    df[ic] = pd.to_datetime(df[ic], errors="coerce")
    return df.set_index(ic).loc[lambda x: x.index.notna()].sort_index()


def _notax_sheet(names):
    for s in names:
        if any(k in s.upper() for k in ["WO TAX","WITHOUT","NO TAX","NOTAX"]):
            return s
    return names[1] if len(names) > 1 else names[0]


def _it_cols(df):
    it_all = [c for c in df.columns if
              str(c).upper().startswith("IT") or "ITAL" in str(c).upper()]
    benz = [c for c in it_all if any(k in str(c).lower()
            for k in ["95","benz","petrol","gasol","euro","unleaded"])]
    dies = [c for c in it_all if any(k in str(c).lower()
            for k in ["diesel","gas_oil","gasoil"])]
    if not benz and it_all:    benz = [it_all[0]]
    if not dies and len(it_all) >= 2: dies = [it_all[1]]
    elif not dies and it_all:  dies = [it_all[0]]
    return benz, dies


pompa = None
used_pretax = False

for label, url, fname, is_pretax in EU_SOURCES:
    try:
        print(f"  Tentativo: {label}...")
        resp = requests.get(url, timeout=60, headers={"User-Agent":"Mozilla/5.0"})
        resp.raise_for_status()
        cnt = resp.content
        if not (cnt[:2] == b"PK" or cnt[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"):
            raise ValueError("risposta non-Excel")
        with open(fname, "wb") as fh:
            fh.write(cnt)

        names  = _sheet_names(fname)
        sheet  = _notax_sheet(names)
        df_raw = _read_sheet(fname, sheet).apply(pd.to_numeric, errors="coerce")
        bc, dc = _it_cols(df_raw)
        if not bc or not dc:
            raise ValueError(f"colonne IT non trovate nel foglio '{sheet}'")

        pompa = pd.concat([df_raw[bc[0]].rename("benzina_eur_l"),
                           df_raw[dc[0]].rename("diesel_eur_l")], axis=1)
        pompa = (pompa[pompa.index >= DATA_START]
                 .dropna(how="all")
                 .sort_index())
        pompa.index = pd.to_datetime(pompa.index)
        pompa = pompa.resample("W-MON").mean()
        used_pretax = is_pretax
        print(f"  OK | foglio '{sheet}' | {pompa.dropna().shape[0]} settimane")
        break
    except Exception as exc:
        print(f"  Fallito: {exc}")

if pompa is None:
    missing_log["prezzi_pompa"] = ["download_fallito"]
    raise SystemExit("Prezzi pompa non disponibili.")

# Traccia e interpola gap brevi
full_idx = pd.date_range(pompa.index[0], pompa.index[-1], freq="W-MON")
miss_p   = full_idx[pompa.reindex(full_idx)["benzina_eur_l"].isna()]
if len(miss_p):
    missing_log["prezzi_pompa"] = [str(d.date()) for d in miss_p]
    pompa = pompa.reindex(full_idx)
    for col in ["benzina_eur_l","diesel_eur_l"]:
        pompa[col] = pompa[col].interpolate(method="time", limit=2)
    print(f"  {len(miss_p)} settimane interpolate (limit=2w)")

if not used_pretax:
    missing_log["nota_tasse"] = "File senza tasse non scaricabile — usati prezzi lordi."
    print("  NOTA: prezzi includono tasse (fallback).")

# ── NORMALIZZAZIONE UNITÀ EUR/1000L → EUR/L ──────────────────────────────
# L'EU Bulletin pubblica in EUR/1000L (valori tipici 400-700).
# I futures Eurobob/GasOil sono in EUR/L (valori tipici 0.35-0.75).
# Per coerenza nel calcolo del margine crack spread, normalizziamo.
for col in ["benzina_eur_l", "diesel_eur_l"]:
    med_val = pompa[col].dropna().median()
    if med_val > 10:
        pompa[col] = pompa[col] / 1000.0
        print(f"  Normalizzato {col}: /1000 (mediana era {med_val:.1f} EUR/1000L -> "
              f"{med_val/1000:.4f} EUR/L)")

pompa["benzina_4w"]  = pompa["benzina_eur_l"]
pompa["diesel_4w"]   = pompa["diesel_eur_l"]
pompa["log_benzina"] = np.log(pompa["benzina_eur_l"])
pompa["log_diesel"]  = np.log(pompa["diesel_eur_l"])
pompa.to_csv("data/prezzi_pompa_italia.csv")
print(f"  Pompa: {pompa.index[0].date()} – {pompa.index[-1].date()}")
print(f"  Check benzina 2019 media: {pompa.loc['2019','benzina_eur_l'].mean():.4f} EUR/L")


# ─────────────────────────────────────────────────────────────────────────────
# 4. Merge brent + pompa
# ─────────────────────────────────────────────────────────────────────────────
merged = brent_w.reindex(pompa.index).join(pompa)
for col in ["brent_eur","brent_7d_eur","log_brent","eurusd"]:
    if col in merged.columns:
        merged[col] = merged[col].ffill(limit=4)
merged = merged.dropna(subset=["benzina_eur_l","diesel_eur_l"])
merged.to_csv("data/dataset_merged.csv")
print(f"\nDataset unificato: {len(merged)} settimane | "
      f"{merged.index[0].date()} – {merged.index[-1].date()}")


# ─────────────────────────────────────────────────────────────────────────────
# 5. Futures wholesale europei (CSV Investing.com pre-scaricati)
# ─────────────────────────────────────────────────────────────────────────────
print("\nCarico futures wholesale (CSV)...")


def _load_investing_csv(path, col="Price"):
    df = pd.read_csv(path, thousands=",")
    df.columns = [c.lstrip("\ufeff") for c in df.columns]
    df["Date"] = pd.to_datetime(df["Date"], format="%m/%d/%Y", errors="coerce")
    df = df.dropna(subset=["Date"]).set_index("Date").sort_index()
    if col in df.columns:
        df[col] = pd.to_numeric(df[col].astype(str).str.replace(",",""), errors="coerce")
    return df


merged_f   = merged.copy()
eurusd_al  = merged_f["eurusd"].ffill().bfill()
futures_ok = {}

for path, raw_col, eur_l_col, l_per_t in [
    ("data/Eurobob Futures Historical Data.csv",
     "eurobob_usd_tonne", "eurobob_eur_l", L_PER_T_BENZ),
    ("data/London Gas Oil Futures Historical Data.csv",
     "gasoil_usd_tonne",  "gasoil_eur_l",  L_PER_T_DIES),
]:
    try:
        df_f = _load_investing_csv(path)
        df_f.rename(columns={"Price": raw_col}, inplace=True)
        wk = df_f[raw_col].resample("W-MON").mean().ffill()
        merged_f = merged_f.join(wk.rename(raw_col), how="left")
        merged_f[eur_l_col] = (merged_f[raw_col] / eurusd_al) / l_per_t
        futures_ok[raw_col] = True
        print(f"  {raw_col}: {len(df_f)} righe | "
              f"mediana {merged_f[eur_l_col].median():.4f} EUR/L")
    except Exception as exc:
        futures_ok[raw_col] = False
        print(f"  {raw_col}: {exc}")

if futures_ok.get("eurobob_usd_tonne"):
    merged_f["margine_benz_crack"] = (merged_f["benzina_eur_l"]
                                      - merged_f["eurobob_eur_l"])
if futures_ok.get("gasoil_usd_tonne"):
    merged_f["margine_dies_crack"] = (merged_f["diesel_eur_l"]
                                      - merged_f["gasoil_eur_l"])



merged_f.to_csv("data/dataset_merged_with_futures.csv")
print(f"Dataset con futures: {len(merged_f)} settimane")


# ─────────────────────────────────────────────────────────────────────────────
# 6. Log settimane mancanti
# ─────────────────────────────────────────────────────────────────────────────
rows_m = []
for fonte, val in missing_log.items():
    if isinstance(val, list):
        for s in val:
            rows_m.append({"fonte": fonte, "settimana_mancante": s})
    else:
        rows_m.append({"fonte": fonte, "settimana_mancante": str(val)})
if rows_m:
    pd.DataFrame(rows_m).to_csv("data/missing_weeks.csv", index=False)
    with open("data/missing_weeks.json","w") as fh:
        json.dump(missing_log, fh, indent=2, default=str)
    print(f"\nLog mancanti: {len(rows_m)} voci → data/missing_weeks.csv")


# ─────────────────────────────────────────────────────────────────────────────
# 7. Plot overview
# ─────────────────────────────────────────────────────────────────────────────
tax_note = "" if used_pretax else " (prezzi con tasse — fallback)"

def _war_lines(ax, y_top):
    for label, (dt, color) in WAR_EVENTS.items():
        ts = pd.Timestamp(dt)
        if merged.index[0] <= ts <= merged.index[-1]:
            ax.axvline(ts, color=color, lw=1.8, ls="--", alpha=0.85)
            ax.text(ts + pd.Timedelta(days=5), y_top * 0.96,
                    label, rotation=90, fontsize=8, color=color, va="top")

for fname_out, col, color, title, unit in [
    ("plots/01a_brent.png",   "brent_eur",     "#2166ac",
     "Brent Crude Oil 2019–2026 (EUR/barile)", "EUR/barile"),
    ("plots/01b_benzina.png", "benzina_eur_l", "#d6604d",
     f"Benzina Italia senza tasse{tax_note} (EUR/litro)", "EUR/litro"),
    ("plots/01c_diesel.png",  "diesel_eur_l",  "#31a354",
     f"Diesel Italia senza tasse{tax_note} (EUR/litro)", "EUR/litro"),
]:
    if col not in merged.columns:
        continue
    fig, ax = plt.subplots(figsize=(14, 4.5))
    ax.plot(merged.index, merged[col], color=color, lw=2.0)
    _war_lines(ax, merged[col].max())
    ax.set_ylabel(unit, fontsize=11)
    ax.set_title(title, fontsize=12, fontweight="bold")
    ax.grid(alpha=0.3)
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %Y"))
    ax.xaxis.set_major_locator(mdates.MonthLocator(interval=4))
    plt.xticks(rotation=40, fontsize=9)
    plt.tight_layout()
    fig.savefig(fname_out, dpi=DPI, bbox_inches="tight")
    plt.close(fig)

print("Plot: plots/01a_brent.png | 01b_benzina.png | 01c_diesel.png")
print("Script 01 completato.")
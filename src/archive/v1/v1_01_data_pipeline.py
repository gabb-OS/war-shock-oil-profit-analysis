"""
01_data_pipeline.py
====================
Scarica e prepara i dati per l'analisi.
  - Brent crude oil (giornaliero) via yfinance → convertito in EUR/barile
  - Prezzi carburanti Italia SENZA TASSE (settimanale) via EU Weekly Oil Bulletin
  - Tre eventi: Ucraina 2022 | Iran-Israele giu 2025 | Hormuz feb 2026
  - Granularità: settimanale (W-MON)
  - In caso di dati mancanti: salva le settimane fallite, NON simula dati
"""

import os
import requests
from datetime import date, timedelta
_TODAY_STR = (date.today() + timedelta(days=7)).strftime('%Y-%m-%d')  # end dinamico: sempre aggiornato
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import matplotlib.patches as mpatches
import warnings
warnings.filterwarnings("ignore")
import yfinance as yf

os.makedirs("data",  exist_ok=True)
os.makedirs("plots", exist_ok=True)

WAR_EVENTS = {
    "Invasione Ucraina":   ("2022-02-24", "#e74c3c"),
    "Guerra Iran-Israele": ("2025-06-13", "#e67e22"),
    "Chiusura Hormuz":     ("2026-02-28", "#8e44ad"),
}

missing_log = {}   # dizionario {fonte: [date mancanti]}

# ─────────────────────────────────────────
# 1. TASSO DI CAMBIO EUR/USD (settimanale)
#    Serve per convertire Brent da USD a EUR
# ─────────────────────────────────────────
print("Scarico tasso EUR/USD (yfinance)...")
try:
    eurusd_raw = yf.download("EURUSD=X", start="2019-01-01", end=_TODAY_STR, progress=False)
    if eurusd_raw.empty:
        raise ValueError("Download vuoto")
    eurusd = eurusd_raw[["Close"]].copy()
    eurusd.columns = ["eurusd"]
    eurusd.index = pd.to_datetime(eurusd.index)
    eurusd.dropna(inplace=True)
    # Resample settimanale (media della settimana)
    eurusd_weekly = eurusd["eurusd"].resample("W-MON").mean()

    # Traccia settimane mancanti
    full_weeks = pd.date_range(eurusd_weekly.index[0], eurusd_weekly.index[-1], freq="W-MON")
    missing_eurusd = full_weeks[~full_weeks.isin(eurusd_weekly.dropna().index)].tolist()
    if missing_eurusd:
        missing_log["eurusd"] = [str(d.date()) for d in missing_eurusd]
        print(f"  Attenzione: {len(missing_eurusd)} settimane mancanti in EUR/USD")
    print(f"  EUR/USD: {len(eurusd_weekly.dropna())} settimane disponibili")
except Exception as e:
    print(f"  ERRORE download EUR/USD: {e}")
    eurusd_weekly = None
    missing_log["eurusd"] = ["download_fallito"]


# ─────────────────────────────────────────
# 2. BRENT (giornaliero → settimanale, in EUR/barile)
# ─────────────────────────────────────────
print("\nScarico Brent crude (yfinance)...")
try:
    brent_raw = yf.download("BZ=F", start="2019-01-01", end=_TODAY_STR, progress=False)
    if brent_raw.empty:
        raise ValueError("Download vuoto")
    brent = brent_raw[["Close"]].copy()
    brent.columns = ["brent_usd"]
    brent.index = pd.to_datetime(brent.index)
    brent.dropna(inplace=True)
        
    brent["brent_7d_usd"] = brent["brent_usd"]
    
    # Resample settimanale
    brent_weekly = brent[["brent_usd", "brent_7d_usd"]].resample("W-MON").mean()

    # Converti in EUR usando il tasso di cambio
    if eurusd_weekly is not None:
        # EUR/USD → per convertire: prezzo_eur = prezzo_usd / tasso_eurusd
        brent_weekly = brent_weekly.join(eurusd_weekly.rename("eurusd"), how="left")
        brent_weekly["eurusd"] = brent_weekly["eurusd"].ffill()   # forward fill settimane vuote
        brent_weekly["brent_eur"] = brent_weekly["brent_usd"] / brent_weekly["eurusd"]
        brent_weekly["brent_7d_eur"] = brent_weekly["brent_7d_usd"] / brent_weekly["eurusd"]
    else:
        # Se EUR/USD non disponibile usa USD con nota
        brent_weekly["brent_eur"]    = brent_weekly["brent_usd"]
        brent_weekly["brent_7d_eur"] = brent_weekly["brent_7d_usd"]
        print("  Attenzione: Brent in USD (conversione EUR non disponibile)")

    brent_weekly["log_brent"] = np.log(brent_weekly["brent_7d_eur"])

    # Traccia settimane mancanti
    full_range = pd.date_range(brent_weekly.index[0], brent_weekly.index[-1], freq="W-MON")
    missing_brent = full_range[brent_weekly["brent_eur"].isna()].tolist()
    if missing_brent:
        missing_log["brent"] = [str(d.date()) for d in missing_brent]
        print(f"  Attenzione: {len(missing_brent)} settimane mancanti in Brent")
        brent_weekly["brent_eur"]    = brent_weekly["brent_eur"].interpolate(method="time")
        brent_weekly["brent_7d_eur"] = brent_weekly["brent_7d_eur"].interpolate(method="time")
        brent_weekly["log_brent"]    = np.log(brent_weekly["brent_7d_eur"])
        print(f"  Settimane mancanti interpolate linearmente.")

    brent_weekly.to_csv("data/brent_weekly_eur.csv")
    print(f"  Brent: {len(brent_weekly.dropna())} settimane | "
          f"{brent_weekly.index[0].date()} - {brent_weekly.index[-1].date()}")

except Exception as e:
    print(f"  ERRORE CRITICO download Brent: {e}")
    missing_log["brent"] = ["download_fallito_completamente"]
    brent_weekly = None

# ─────────────────────────────────────────
# 3. PREZZI POMPA ITALIA — SENZA TASSE (EUR/litro)
#    Fonte: EU Weekly Oil Bulletin — foglio "Prices without taxes"
#    Granularità: settimanale (ogni lunedì)
# ─────────────────────────────────────────
print("\nScarico prezzi pompa Italia SENZA TASSE (EU Oil Bulletin)...")

# EU pubblica i dati tramite UUID nel path — gli URL statici /system/files/ sono deprecati.
# Tre URL in ordine di preferenza:
#   1. File storico completo dal 2005 (contiene entrambe le serie, un foglio per tipo)
#   2. File settimanale senza tasse (aggiornato ogni lunedi)
#   3. File settimanale con tasse (fallback documentato)
# Aggiornati a marzo 2026 dai bulletin ufficiali EC Newsroom.
EU_URLS = [
    # 1. Storico completo 2005-presente (priorita -- copre tutta la finestra di analisi)
    ("https://energy.ec.europa.eu/document/download/"
     "906e60ca-8b6a-44e7-8589-652854d2fd3f_en?"
     "filename=Weekly_Oil_Bulletin_Prices_History_maticni_4web.xlsx"),
    # 2. Settimanale senza tasse
    ("https://energy.ec.europa.eu/document/download/"
     "78311f92-68f8-4b82-b5cf-1293beeaae77_en?"
     "filename=Weekly+Oil+Bulletin+Weekly+prices+without+taxes+-+2024-02-19.xlsx"),
    # 3. Settimanale con tasse (fallback)
    ("https://energy.ec.europa.eu/document/download/"
     "264c2d0f-f161-4ea3-a777-78faae59bea0_en?"
     "filename=Weekly+Oil+Bulletin+Weekly+prices+with+Taxes+-+2024-02-19.xlsx"),
]

pompa = None
used_pretax = False

# Label e nome file per ogni URL
EU_URL_META = [
    ("storico 2005-presente (senza tasse)",  "data/eu_oil_bulletin_history.xlsx",  True),
    ("settimanale senza tasse",              "data/eu_oil_bulletin_notax.xlsx",    True),
    ("settimanale con tasse (fallback)",     "data/eu_oil_bulletin_tax.xlsx",      False),
]

import openpyxl as _opxl

def _safe_sheet_names(path):
    """Legge nomi fogli con openpyxl read_only=True.
    Evita il bug 'argument of type float is not iterable' causato da
    data-validation con formula1 numerica nei file EU Bulletin."""
    wb = _opxl.load_workbook(path, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names

def _safe_read_sheet(path, sheet_name):
    """Legge un foglio xlsx via openpyxl read_only=True → DataFrame.
    Prima riga non-vuota = header, prima colonna = indice datetime."""
    wb = _opxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if not rows:
        return pd.DataFrame()
    header_idx = next((i for i, r in enumerate(rows) if any(v is not None for v in r)), 0)
    headers = [str(h).strip() if h is not None else f"_col{i}"
               for i, h in enumerate(rows[header_idx])]
    df = pd.DataFrame(rows[header_idx + 1:], columns=headers)
    idx_col = headers[0]
    df[idx_col] = pd.to_datetime(df[idx_col], errors="coerce")
    df = df.set_index(idx_col)
    df = df[df.index.notna()].sort_index()
    return df

def _find_notax_sheet(sheet_names):
    for s in sheet_names:
        su = s.upper()
        if any(k in su for k in ["WO TAX", "WITHOUT", "NO TAX", "NOTAX", "WO TAXES"]):
            return s
    return sheet_names[1] if len(sheet_names) > 1 else sheet_names[0]

for url_idx, (eu_url, (label, fname, is_pretax)) in enumerate(zip(EU_URLS, EU_URL_META)):
    try:
        print(f"  Tentativo: {label}...")
        resp = requests.get(
            eu_url, timeout=60,
            headers={"User-Agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            )},
        )
        resp.raise_for_status()

        # Valida che sia un vero file Excel (non HTML di errore con status 200)
        content = resp.content
        if not (content[:2] == b'PK' or content[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'):
            raise ValueError(
                f"Risposta non-Excel (possibile pagina HTML di errore). "
                f"Scarica manualmente: {eu_url}"
            )
        with open(fname, "wb") as fh:
            fh.write(content)

        # ── Usa openpyxl read_only=True (salta data-validation) ───────────────
        # pd.ExcelFile/pd.read_excel carica ANCHE le data-validation del file.
        # Il file EU Bulletin contiene validation con formula1 float nel XML,
        # che causa TypeError "argument of type float is not iterable".
        # Con read_only=True openpyxl salta completamente data-validation.
        sheet_names = _safe_sheet_names(fname)
        sheet_notax = _find_notax_sheet(sheet_names)
        print(f"  Foglio senza tasse: '{sheet_notax}'")

        df_notax = _safe_read_sheet(fname, sheet_notax)
        df_notax = df_notax.apply(pd.to_numeric, errors="coerce")

        # Trova colonne IT senza tasse: benzina (euro95/gasoline) e diesel
        it_all = [c for c in df_notax.columns if str(c).upper().startswith("IT")
                  or "ITALY" in str(c).upper() or "ITAL" in str(c).upper()]

        # Separa benzina e diesel dalle colonne IT disponibili
        it_b = [c for c in it_all if any(k in str(c).lower()
                for k in ["95", "benz", "petrol", "gasol", "euro"])]
        it_d = [c for c in it_all if any(k in str(c).lower()
                for k in ["diesel", "gas_oil", "gasoil"])]

        # Se la separazione non ha funzionato, usa le prime due colonne IT disponibili
        if not it_b and len(it_all) >= 1:
            it_b = [it_all[0]]
        if not it_d and len(it_all) >= 2:
            it_d = [it_all[1]]
        elif not it_d and len(it_all) == 1:
            it_d = [it_all[0]]

        if not it_b or not it_d:
            print(f"  Colonna Italia non trovata ({label}). Foglio: {sheet_notax}, cols IT: {it_all}")
            continue

        benzina_it = df_notax[it_b[0]].copy()
        diesel_it  = df_notax[it_d[0]].copy()

        # Indice temporale: gia DatetimeIndex da _safe_read_sheet, ma rinforzato
        benzina_it.index = pd.to_datetime(benzina_it.index, errors="coerce")
        diesel_it.index  = pd.to_datetime(diesel_it.index,  errors="coerce")

        benzina_it = benzina_it[benzina_it.index.notna()]
        diesel_it  = diesel_it[diesel_it.index.notna()]

        # Assicura dtype float (coerce gia fatto sopra, ma per sicurezza)
        benzina_it = pd.to_numeric(benzina_it, errors="coerce")
        diesel_it  = pd.to_numeric(diesel_it,  errors="coerce")

        pompa = pd.concat([
            benzina_it.rename("benzina_eur_l"),
            diesel_it.rename("diesel_eur_l"),
        ], axis=1)
        print(pompa.index)
        pompa = pompa[pompa.index >= "2019-01-01"]
        pompa.dropna(how="all", inplace=True)
        pompa.sort_index(inplace=True)
        pompa.index = pd.to_datetime(pompa.index)

        print(pompa.info())
        # Resample a W-MON per allineamento col Brent
        pompa = pompa.resample("W-MON").mean()
        print(pompa.info())


        used_pretax = is_pretax
        print(f"  OK: {label} — {len(pompa.dropna())} settimane")
        print(f"  Foglio senza tasse: '{sheet_notax}'")
        print(f"  Colonne IT: benzina={it_b[0]}, diesel={it_d[0]}")
        break

    except Exception as e:
        print(f"  Fallito ({label}): {e}")
        continue

# Traccia settimane mancanti nei prezzi pompa
if pompa is not None and len(pompa) > 0:
    full_range_p = pd.date_range(pompa.index[0], pompa.index[-1], freq="W-MON")
    miss_benz = full_range_p[pompa["benzina_eur_l"].isna()].tolist()
    miss_dies = full_range_p[pompa["diesel_eur_l"].isna()].tolist()

    all_missing_p = sorted(set(miss_benz + miss_dies))
    if all_missing_p:
        missing_log["prezzi_pompa"] = [str(d.date()) for d in all_missing_p]
        print(f"  Attenzione: {len(all_missing_p)} settimane con dati parziali/mancanti")
        # Interpolazione lineare per settimane isolate mancanti
        pompa["benzina_eur_l"] = pompa["benzina_eur_l"].interpolate(method="time")
        pompa["diesel_eur_l"]  = pompa["diesel_eur_l"].interpolate(method="time")
        print(f"  Settimane mancanti interpolate.")

    if not used_pretax:
        missing_log["prezzi_pretax_note"] = (
            "File senza tasse non disponibile. Utilizzati prezzi con tasse come fallback. "
            "I prezzi includono IVA e accise italiane."
        )
        print("  NOTA: prezzi includono tasse (file senza tasse non scaricabile).")

else:
    # Nessun dato pompa disponibile — registra il problema e interrompi
    missing_log["prezzi_pompa"] = ["download_fallito_completamente"]
    print("  ERRORE CRITICO: impossibile scaricare i prezzi pompa.")
    print("  Salvo il log dei fallimenti e termino.")

    pd.DataFrame([{"fonte": k, "settimane_mancanti": v}
                  for k, v in missing_log.items()]).to_csv("data/missing_weeks.csv", index=False)
    import json
    with open("data/missing_weeks.json", "w") as f:
        json.dump(missing_log, f, indent=2, default=str)

    raise SystemExit(
        "\nScript interrotto: dati prezzi pompa non disponibili.\n"
        "Controlla data/missing_weeks.csv per il dettaglio delle settimane mancanti."
    )

# ─────────────────────────────────────────
# 4. LOG TRANSFORM (no rolling average)
# ─────────────────────────────────────────

pompa["benzina_4w"]  = pompa["benzina_eur_l"]   # alias per retrocompatibilità
pompa["diesel_4w"]   = pompa["diesel_eur_l"]    # alias per retrocompatibilità
pompa["log_benzina"] = np.log(pompa["benzina_eur_l"])
pompa["log_diesel"]  = np.log(pompa["diesel_eur_l"])
pompa.to_csv("data/prezzi_pompa_italia.csv")
print(f"  Prezzi pompa: {pompa.index[0].date()} - {pompa.index[-1].date()}")

# ─────────────────────────────────────────
# 5. MERGE (allineamento settimanale)
# ─────────────────────────────────────────
if brent_weekly is None:
    pd.DataFrame([{"fonte": k, "settimane_mancanti": v}
                  for k, v in missing_log.items()]).to_csv("data/missing_weeks.csv", index=False)
    raise SystemExit(
        "\nScript interrotto: dati Brent non disponibili.\n"
        "Controlla data/missing_weeks.csv per il dettaglio."
    )

# Left join su pompa (EU Oil Bulletin, più recente): mantiene tutte le settimane
# del Bulletin. Le ultime 2-3 settimane di Brent mancanti vengono portate avanti
# con ffill (il prezzo futures cambia poco in pochi giorni: errore trascurabile).
# NOTA: pd.concat con join="left" è rimosso in pandas ≥ 2.0 → uso .join() + reindex.
merged = brent_weekly.reindex(pompa.index).join(pompa)
for col in ["brent_eur", "brent_7d_eur", "log_brent", "eurusd"]:
    if col in merged.columns:
        merged[col] = merged[col].ffill(limit=4)  # max 4 settimane ffill
merged = merged.dropna(subset=["benzina_eur_l", "diesel_eur_l"])
merged.to_csv("data/dataset_merged.csv")

print(f"\n  Dataset unificato: {len(merged)} settimane | "
      f"{merged.index[0].date()} - {merged.index[-1].date()}")
print(f"  Brent in EUR: {'si' if eurusd_weekly is not None else 'no (USD)'}")
print(f"  Prezzi pompa senza tasse: {'si' if used_pretax else 'no (con tasse)'}")

# ─────────────────────────────────────────
# 6. SALVA LOG SETTIMANE MANCANTI
# ─────────────────────────────────────────
import json

missing_df_rows = []
for fonte, valore in missing_log.items():
    if isinstance(valore, list):
        for settimana in valore:
            missing_df_rows.append({"fonte": fonte, "settimana_mancante": settimana})
    else:
        missing_df_rows.append({"fonte": fonte, "settimana_mancante": valore})

if missing_df_rows:
    pd.DataFrame(missing_df_rows).to_csv("data/missing_weeks.csv", index=False)
    with open("data/missing_weeks.json", "w") as f:
        json.dump(missing_log, f, indent=2, default=str)
    print(f"\n  Log settimane mancanti salvato: data/missing_weeks.csv")
    for fonte, val in missing_log.items():
        if isinstance(val, list) and val and val[0] not in ("download_fallito", "download_fallito_completamente"):
            print(f"  Problema rilevato in '{fonte}': {len(val)} settimane mancanti")
        elif isinstance(val, str):
            print(f"  Nota '{fonte}': {val}")
else:
    print("\n  Nessuna settimana mancante rilevata.")


# ─────────────────────────────────────────
# 7. PLOT OVERVIEW — paper quality
# ─────────────────────────────────────────
FIGSIZE   = (14, 5)
DPI       = 180
FONT_AXIS = 12
FONT_TICK = 10
FONT_LEG  = 10

brent_col    = "brent_eur"
brent_ma_col = "brent_7d_eur"
brent_label  = "EUR / barile"
brent_ma_lbl = "Brent 7d avg (EUR)"

tax_note = "" if used_pretax else " (prezzi con tasse — file senza tasse non disponibile)"

def add_war_lines(ax, ylim_top):
    for label, (date, color) in WAR_EVENTS.items():
        ts = pd.Timestamp(date)
        if merged.index[0] <= ts <= merged.index[-1]:
            ax.axvline(ts, color=color, lw=1.8, linestyle="--", alpha=0.9)
            ax.text(ts + pd.Timedelta(days=5), ylim_top * 0.97,
                    label, rotation=90, fontsize=9, color=color,
                    va="top", ha="left")

# --- Plot A: Brent in EUR ---
fig, ax = plt.subplots(figsize=FIGSIZE)
ax.plot(merged.index, merged[brent_col],    color="#aec6e8", lw=1.0, alpha=0.7)
ax.plot(merged.index, merged[brent_ma_col], color="#2166ac", lw=2.2)
add_war_lines(ax, merged[brent_col].max())
ax.set_ylabel(brent_label, fontsize=FONT_AXIS)
ax.set_title("Prezzo Brent Crude Oil — 2019–2026 (EUR/barile)", fontsize=14, fontweight="bold")
patches_war = [mpatches.Patch(color=c, label=l) for l, (_, c) in WAR_EVENTS.items()]
ax.legend(handles=[
    plt.Line2D([0],[0], color="#aec6e8", lw=1.5, label="Brent settimanale"),
    plt.Line2D([0],[0], color="#2166ac", lw=2.2, label=brent_ma_lbl),
] + patches_war, fontsize=FONT_LEG, loc="upper left")
ax.grid(alpha=0.3)
ax.tick_params(labelsize=FONT_TICK)
ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %Y"))
ax.xaxis.set_major_locator(mdates.MonthLocator(interval=3))
plt.xticks(rotation=45)
plt.tight_layout()
plt.savefig("plots/01a_brent.png", dpi=DPI, bbox_inches="tight")
plt.close()

# --- Plot B: Benzina senza tasse ---
fig, ax = plt.subplots(figsize=FIGSIZE)
ax.plot(merged.index, merged["benzina_eur_l"], color="#d6604d", lw=2.0)
add_war_lines(ax, merged["benzina_eur_l"].max())
ax.set_ylabel("EUR / litro", fontsize=FONT_AXIS)
ax.set_title(f"Prezzo Benzina Italia — senza tasse{tax_note}\n2019–2026",
             fontsize=13, fontweight="bold")
ax.legend(handles=[
    plt.Line2D([0],[0], color="#d6604d", lw=2.0, label="Benzina settimanale (EUR/litro)"),
] + patches_war, fontsize=FONT_LEG, loc="upper left")
ax.grid(alpha=0.3)
ax.tick_params(labelsize=FONT_TICK)
ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %Y"))
ax.xaxis.set_major_locator(mdates.MonthLocator(interval=3))
plt.xticks(rotation=45)
plt.tight_layout()
plt.savefig("plots/01b_benzina.png", dpi=DPI, bbox_inches="tight")
plt.close()

# --- Plot C: Diesel senza tasse ---
fig, ax = plt.subplots(figsize=FIGSIZE)
ax.plot(merged.index, merged["diesel_eur_l"], color="#31a354", lw=2.0)
add_war_lines(ax, merged["diesel_eur_l"].max())
ax.set_ylabel("EUR / litro", fontsize=FONT_AXIS)
ax.set_title(f"Prezzo Diesel Italia — senza tasse{tax_note}\n2019–2026",
             fontsize=13, fontweight="bold")
ax.legend(handles=[
    plt.Line2D([0],[0], color="#31a354", lw=2.0, label="Diesel settimanale (EUR/litro)"),
] + patches_war, fontsize=FONT_LEG, loc="upper left")
ax.grid(alpha=0.3)
ax.tick_params(labelsize=FONT_TICK)
ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %Y"))
ax.xaxis.set_major_locator(mdates.MonthLocator(interval=3))
plt.xticks(rotation=45)
plt.tight_layout()
plt.savefig("plots/01c_diesel.png", dpi=DPI, bbox_inches="tight")
plt.close()

print("\nPlot salvati: plots/01a_brent.png | 01b_benzina.png | 01c_diesel.png")
print("Script 01 completato.")
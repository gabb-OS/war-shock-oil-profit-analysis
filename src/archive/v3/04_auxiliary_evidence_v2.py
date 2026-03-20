"""
04_auxiliary_evidence_v2.py
============================
Evidenza ausiliare e test di specificità italiana.

ARCHITETTURA EPISTEMICA (v2)
─────────────────────────────
Questo script opera al LIVELLO 3 della pipeline v2:

  LIVELLO 1 (03_margin_hypothesis_v2.py)  → H₀: μ_post = μ_2019  → BH famiglia A
  LIVELLO 2 (03_margin_hypothesis_v2.py)  → esplorativi, τ_margin  → no BH
  LIVELLO 3 (questo script)               → H₀: δ_DiD = 0          → BH famiglia B
                                           → windfall, Granger, R&F  → descriptivo/exploratorio

SEZIONI
───────
  §1. Difference-in-Differences (DiD) — IT vs DE, IT vs SE
      H₀: δ_DiD = 0  (margine IT non aumenta più di quello del paese controllo)
      • Famiglia BH separata (8 test: 2 eventi × 2 carburanti × 2 paesi)
      • Parallel Trends Assumption (PTA) testata pre-shock
      • Interpretazione corretta DiD negativo (Italy < Germania → contro opportunismo)
      Output: data/did_results_v2.csv

  §2. Granger causality — Brent → prezzi pompa  [esplorativo]
      Output: data/granger_results_v2.csv

  §3. Rockets & Feathers — asimmetria rialzo/ribasso  [esplorativo]
      Output: data/rockets_feathers_v2.csv

  §4. Windfall profits — stima con correzione trend consumi
      Miglioria v2: volumi corretti per trend lineare (-1.5%/anno da 2022)
      invece di volumi fissi 2022.
      Output: data/windfall_v2.csv

  §5. Hormuz (Feb 2026) — analisi preliminare  [fuori da BH, flaggato]
      Solo 7 settimane post-shock → stime instabili. Nessun p-value in BH.
      Output: data/hormuz_preliminary.csv

NOTA SU DiD NEGATIVO (Ucraina, benzina e diesel)
─────────────────────────────────────────────────
I δ_DiD IT vs DE per Ucraina risultano negativi (benzina ≈ −0.024, diesel ≈ −0.033
EUR/l, CI 95% che include zero). Questo NON è un errore: significa che
l'Italia ha avuto margini inferiori (o non superiori) alla Germania post-Ucraina.
È un finding contro l'ipotesi di opportunismo specifico italiano per quell'evento.
Interpretazione nel paper: cfr. interpretation_note nel CSV output.

NOTA SU SVEZIA come controllo
──────────────────────────────
Svezia sostituisce Francia v1: la Francia ha introdotto ristorno fiscale
~0.15–0.18 €/l (set–dic 2022, Décret n°2022-1153) che comprime artificialmente
il margine osservato FR e confonde il DiD per Ucraina 2022.
La Svezia non ha avuto misure fiscali comparabili nel periodo analizzato.

Input:
  data/dataset_merged_with_futures.csv  (o dataset_merged.csv)
  data/eu_oil_bulletin_history.xlsx
  data/confirmatory_pvalues_v2.csv      (per τ_price da MCMC)
  data/table1_changepoints.csv          (τ_price per evento/serie)

Output:
  data/did_results_v2.csv
  data/granger_results_v2.csv
  data/rockets_feathers_v2.csv
  data/windfall_v2.csv
  data/hormuz_preliminary.csv
  plots/04_did_v2.png
  plots/04_rf_v2.png
  plots/04_windfall_v2.png

Rif: Angrist & Pischke (2009) cap. 5; Card & Krueger (1994);
     Andrews (1991) Econometrica; Newey & West (1987).
"""

from __future__ import annotations

import os
import warnings
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from scipy import stats
from scipy.stats import mannwhitneyu
import statsmodels.api as sm
from statsmodels.tsa.stattools import grangercausalitytests
from statsmodels.stats.stattools import durbin_watson

warnings.filterwarnings("ignore")

# ─── Configurazione ───────────────────────────────────────────────────────────
ALPHA      = 0.05
DPI        = 150
PTA_WEEKS  = 8   # finestra pre-shock per il test PTA (settimane)
               # La guida dichiara 8 settimane. Usare l'intero pre-shock
               # (14-17 mesi) causa rigetti spurii per divergenze strutturali
               # di lungo periodo non correlate all'evento analizzato.
               # Rif: DeepSeek review 2026-04-27 / §2.7 report metodologico.

# Fattore conversione Brent: $/bbl → EUR/litro
# 1 bbl = 158.987 L; usiamo BRENT_EUR_L = brent_eur / (159 * yield_factor)
BRENT_BBL_TO_L = 158.987
YIELD_BENZ     = 0.45    # rendimento benzina da barile grezzo
YIELD_DIES     = 0.52    # rendimento diesel

# Trend consumi v2: riduzione lineare stimata (-1.5%/anno dal 2022)
# Fonte: MISE dati vendite carburanti 2019-2023 + proiezione IEA 2024-2026
CONSUMPTION_TREND_PCT_PER_YEAR = -0.015   # -1.5% per anno

# ── Volume proxy 2022 (milioni litri/settimana)
# Fonte: MISE - Dipartimento Energia, vendite mensili 2022 / 52
VOLUME_PROXY_2022 = {
    "Benzina": 1_750_000_000 / 52,   # ~33.65 ML/settimana
    "Diesel":  2_250_000_000 / 52,   # ~43.27 ML/settimana
}

os.makedirs("data",  exist_ok=True)
os.makedirs("plots", exist_ok=True)

# ─── Carica dataset principale ────────────────────────────────────────────────
def _load_merged() -> pd.DataFrame:
    for fname in ["data/dataset_merged_with_futures.csv", "data/dataset_merged.csv"]:
        if os.path.exists(fname):
            df = pd.read_csv(fname, index_col=0, parse_dates=True)
            return df
    raise FileNotFoundError(
        "dataset_merged.csv non trovato — eseguire 01_data_pipeline.py"
    )

try:
    merged = _load_merged()
    print(f"  Dataset caricato: {len(merged)} righe, {merged.columns.tolist()}")
except FileNotFoundError as e:
    print(f"  ERRORE: {e}")
    merged = None

# ─── Risolvi nomi colonne margine (robusto a varianti) ────────────────────────
def _margine_col(df: pd.DataFrame, fuel: str) -> str | None:
    """Trova la colonna margine per carburante (benzina/diesel), case-insensitive."""
    fuel_l = fuel.lower()
    candidates = [c for c in df.columns if "margine" in c.lower() or "margin" in c.lower()]
    for c in candidates:
        cl = c.lower()
        if fuel_l == "benzina" and any(k in cl for k in ["benz", "petrol", "gasol_95", "crack_b"]):
            return c
        if fuel_l == "diesel" and any(k in cl for k in ["dies", "gasoil", "gas_oil", "crack_d"]):
            return c
    # fallback: primo/secondo candidato
    if fuel_l == "benzina" and len(candidates) > 0:
        return candidates[0]
    if fuel_l == "diesel" and len(candidates) > 1:
        return candidates[1]
    return None

# ─── Configurazione eventi ────────────────────────────────────────────────────
EVENTS = {
    "Ucraina (Feb 2022)": {
        "shock":      pd.Timestamp("2022-02-24"),
        "pre_start":  pd.Timestamp("2021-01-11"),
        "post_end":   pd.Timestamp("2022-12-31"),
        "preliminary": False,
    },
    "Iran-Israele (Giu 2025)": {
        "shock":      pd.Timestamp("2025-06-13"),
        "pre_start":  pd.Timestamp("2024-01-01"),
        "post_end":   pd.Timestamp("2025-12-31"),
        "preliminary": False,
    },
    "Hormuz (Feb 2026)": {
        "shock":      pd.Timestamp("2026-02-01"),
        "pre_start":  pd.Timestamp("2025-06-01"),
        "post_end":   pd.Timestamp("2026-04-30"),
        "preliminary": True,
        "preliminary_note": (
            "Solo 7 settimane post-shock al momento dell'analisi. "
            "Nessun p-value in BH. Risultati altamente instabili."
        ),
    },
}

FUELS = ["Benzina", "Diesel"]

# ─── Helper: stelle significatività ──────────────────────────────────────────
def _stars(p: float) -> str:
    if p < 0.001: return "***"
    if p < 0.01:  return "**"
    if p < 0.05:  return "*"
    if p < 0.10:  return "."
    return ""

# ─── Helper: BH correction ───────────────────────────────────────────────────
def bh_correction(p_values: np.ndarray, alpha: float = 0.05):
    """Benjamini-Hochberg FDR correction."""
    p = np.asarray(p_values, dtype=float)
    n = len(p)
    if n == 0:
        return np.array([], dtype=bool), np.array([])
    order      = np.argsort(p)
    ranked     = np.empty(n)
    ranked[order] = np.arange(1, n + 1)
    p_adj      = np.minimum(1.0, p * n / ranked)
    # Monotonicity (step-up)
    p_mono     = np.minimum.accumulate(p_adj[order][::-1])[::-1]
    p_out      = np.empty(n)
    p_out[order] = p_mono
    return p_out <= alpha, p_out

# ─────────────────────────────────────────────────────────────────────────────
# §1. DIFFERENCE-IN-DIFFERENCES
# ─────────────────────────────────────────────────────────────────────────────

print("\n" + "=" * 70)
print("§1. DIFFERENCE-IN-DIFFERENCES — IT vs DE, IT vs SE")
print("    H₀: δ_DiD = 0  (famiglia BH separata da confermativi margine)")
print("=" * 70)

# Carica paesi controllo da EU Oil Bulletin
DID_CONTROLS = {
    "Germania": {
        "prefissi_col": ["DE ", "GERMANY", "DEUTSCH", "DE_"],
        "nota": None,
    },
    "Svezia": {
        "prefissi_col": ["SE ", "SWEDEN", "SVERIGE", "SE_"],
        "nota": None,
    },
}

CONTROL_PUMPS: dict[str, pd.DataFrame] = {}

def _safe_wb_sheets(path):
    try:
        import openpyxl as _opxl
        wb = _opxl.load_workbook(path, read_only=True, data_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    except Exception:
        return []

def _safe_wb_read(path, sheet):
    import openpyxl as _opxl
    wb = _opxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if not rows:
        return pd.DataFrame()
    # Trova header
    hdr_i = next(
        (i for i, r in enumerate(rows) if sum(v is not None for v in r) >= 3), 0
    )
    headers = [
        str(h).strip() if h is not None else f"_col{j}"
        for j, h in enumerate(rows[hdr_i])
    ]
    df = pd.DataFrame(rows[hdr_i + 1:], columns=headers)
    date_col = headers[0]
    df[date_col] = pd.to_datetime(df[date_col], errors="coerce")
    df = df.set_index(date_col).sort_index()
    return df[df.index.notna()]

EU_FILE = "data/eu_oil_bulletin_history.xlsx"

def _notax_sheet(names: list[str]) -> str:
    for s in names:
        su = s.upper()
        if any(k in su for k in ["WO TAX", "WITHOUT", "NO TAX", "NOTAX", "WO TAXES"]):
            return s
    return names[1] if len(names) > 1 else names[0]

if merged is not None and os.path.exists(EU_FILE):
    try:
        sheets  = _safe_wb_sheets(EU_FILE)
        sheet   = _notax_sheet(sheets)
        df_eu   = _safe_wb_read(EU_FILE, sheet)
        df_eu   = df_eu.apply(pd.to_numeric, errors="coerce")

        for paese, cfg in DID_CONTROLS.items():
            prefs = cfg["prefissi_col"]
            all_c = [
                c for c in df_eu.columns
                if any(str(c).upper().startswith(p.strip()) or p.strip() in str(c).upper()
                       for p in prefs)
            ]
            b_cols = [c for c in all_c if any(
                k in str(c).lower() for k in ["95", "benz", "petrol", "unleaded", "super", "euro_"]
            )]
            d_cols = [c for c in all_c if any(
                k in str(c).lower() for k in ["diesel", "gasoil", "gas_oil"]
            )]
            if not b_cols and len(all_c) >= 1:
                b_cols = [all_c[0]]
            if not d_cols and len(all_c) >= 2:
                d_cols = [all_c[1]]

            if b_cols and d_cols:
                pump = pd.concat([
                    df_eu[b_cols[0]].rename("benzina_eur_l"),
                    df_eu[d_cols[0]].rename("diesel_eur_l"),
                ], axis=1)
                pump = pump[pump.index >= "2019-01-01"].dropna(how="all")
                pump = pump.resample("W-MON").mean()
                # Normalizza se valori in EUR/1000L anziché EUR/L
                for col in ["benzina_eur_l", "diesel_eur_l"]:
                    med = pump[col].dropna().median()
                    if med > 10:
                        pump[col] = pump[col] / 1000.0

                # Crack spread usando stesso Brent del dataset IT
                brent_aligned = merged["brent_eur"].reindex(pump.index).ffill(limit=4)
                pump["margine_benzina"] = pump["benzina_eur_l"] - brent_aligned / (BRENT_BBL_TO_L * YIELD_BENZ)
                pump["margine_diesel"]  = pump["diesel_eur_l"]  - brent_aligned / (BRENT_BBL_TO_L * YIELD_DIES)
                CONTROL_PUMPS[paese] = pump
                print(f"  {paese}: {len(pump.dropna())} settimane  "
                      f"({b_cols[0]}, {d_cols[0]})")
            else:
                print(f"  {paese}: colonne non trovate nel foglio '{sheet}' — skip")

    except Exception as e:
        print(f"  Errore caricamento EU Bulletin: {e}")
else:
    if merged is None:
        print("  Skip DiD: dataset principale non disponibile")
    else:
        print(f"  Skip DiD: {EU_FILE} non trovato")

# ── Esegui DiD ────────────────────────────────────────────────────────────────
did_rows: list[dict] = []

FUEL_TO_COL = {
    "Benzina": "margine_benzina",
    "Diesel":  "margine_diesel",
}

if merged is not None and CONTROL_PUMPS:
    for paese, ctrl in CONTROL_PUMPS.items():
        for event_name, ecfg in EVENTS.items():
            if ecfg["preliminary"]:
                continue   # Hormuz: non in questa famiglia BH

            shock     = ecfg["shock"]
            pre_start = ecfg["pre_start"]
            post_end  = ecfg["post_end"]

            for fuel in FUELS:
                mcol_it = _margine_col(merged, fuel) or FUEL_TO_COL[fuel]
                mcol_ct = FUEL_TO_COL[fuel]

                if mcol_it not in merged.columns or mcol_ct not in ctrl.columns:
                    print(f"  SKIP {event_name} {fuel} ({paese}): colonna margine non trovata")
                    continue

                it_pre  = merged.loc[pre_start:shock,  mcol_it].dropna()
                it_post = merged.loc[shock:post_end,   mcol_it].dropna()
                ct_pre  = ctrl.loc[pre_start:shock,    mcol_ct].dropna()
                ct_post = ctrl.loc[shock:post_end,     mcol_ct].dropna()

                if any(len(s) < 3 for s in [it_pre, it_post, ct_pre, ct_post]):
                    print(f"  SKIP {event_name} {fuel} ({paese}): dati insufficienti "
                          f"({len(it_pre)}/{len(it_post)}/{len(ct_pre)}/{len(ct_post)})")
                    continue

                # ── Parallel Trends Assumption ────────────────────────────────
                # [v2.1 FIX] PTA testata sulle ultime PTA_WEEKS settimane
                # pre-shock (8 settimane, come dichiarato nella guida).
                # In v2.0 usava l'intero pre-shock (14-17 mesi), causando
                # rigetti spurii per divergenze strutturali pre-esistenti
                # non legate all'evento. Rif: §2.7 report metodologico.
                pta_p     = np.nan
                pta_pass  = None
                try:
                    pta_window  = pd.Timedelta(weeks=PTA_WEEKS)
                    it_pta = it_pre[it_pre.index >= (shock - pta_window)]
                    ct_pta = ct_pre[ct_pre.index >= (shock - pta_window)]
                    if len(it_pta) < 3 or len(ct_pta) < 3:
                        # finestra troppo corta: usa l'intera pre con nota
                        it_pta, ct_pta = it_pre, ct_pre
                    all_pta = it_pta.index.union(ct_pta.index).sort_values()
                    t_it = np.array([(d - all_pta[0]).days for d in it_pta.index], dtype=float)
                    t_ct = np.array([(d - all_pta[0]).days for d in ct_pta.index], dtype=float)
                    rows_pt = (
                        [(1, t, v) for t, v in zip(t_it, it_pta.values)] +
                        [(0, t, v) for t, v in zip(t_ct, ct_pta.values)]
                    )
                    dpt = pd.DataFrame(rows_pt, columns=["IT", "t", "M"])
                    dpt["IT_x_t"] = dpt["IT"] * dpt["t"]
                    X_pt = sm.add_constant(dpt[["IT", "t", "IT_x_t"]].values)
                    ols_pt = sm.OLS(dpt["M"].values, X_pt).fit(cov_type="HC3")
                    pta_p    = float(ols_pt.pvalues[3])   # H₀: coeff trend × IT = 0
                    pta_pass = pta_p >= ALPHA
                except Exception as ep:
                    print(f"    PTA errore: {ep}")

                # ── Modello DiD ──────────────────────────────────────────────
                rows_panel = (
                    [(1, 0, v) for v in it_pre]  +
                    [(1, 1, v) for v in it_post] +
                    [(0, 0, v) for v in ct_pre]  +
                    [(0, 1, v) for v in ct_post]
                )
                dpanel = pd.DataFrame(rows_panel, columns=["IT", "Post", "M"])
                dpanel["IT_x_Post"] = dpanel["IT"] * dpanel["Post"]

                X   = sm.add_constant(dpanel[["IT", "Post", "IT_x_Post"]].values)
                ols = sm.OLS(dpanel["M"].values, X).fit(cov_type="HC3")

                delta  = float(ols.params[3])
                se     = float(ols.bse[3])
                t_stat = float(ols.tvalues[3])
                p_val  = float(ols.pvalues[3])
                ci_lo  = delta - 1.96 * se
                ci_hi  = delta + 1.96 * se

                # Interpretazione economica
                if delta < 0 and p_val >= ALPHA:
                    note = (
                        "δ negativo non significativo: IT non ha margini superiori a "
                        f"{paese} post-shock → evidenza CONTRO specificità italiana"
                    )
                elif delta < 0 and p_val < ALPHA:
                    note = (
                        f"δ significativamente negativo: IT ha margini INFERIORI a "
                        f"{paese} post-shock → forte evidenza CONTRO opportunismo IT"
                    )
                elif delta > 0 and p_val < ALPHA:
                    note = (
                        f"δ positivo significativo: IT ha margini superiori a "
                        f"{paese} post-shock → evidenza A FAVORE di specificità italiana"
                    )
                else:
                    note = "δ positivo non significativo: differenza IT vs controllo non distinguibile da zero"

                pta_warn = " ⚠ PTA violata" if pta_pass is False else ""
                print(
                    f"  [{paese}] {event_name.split('(')[0].strip():<22} | {fuel:<8}: "
                    f"δ={delta:+.4f} EUR/l  SE={se:.4f}  p={p_val:.4f}{_stars(p_val)}"
                    f"  CI=[{ci_lo:+.4f}, {ci_hi:+.4f}]  → {'RIFIUTATA' if p_val<ALPHA else 'n.s.'}"
                    f"{pta_warn}"
                )

                did_rows.append({
                    "Evento":              event_name,
                    "Paese_controllo":     paese,
                    "Carburante":          fuel,
                    "n_IT_pre":            len(it_pre),
                    "n_IT_post":           len(it_post),
                    "n_CT_pre":            len(ct_pre),
                    "n_CT_post":           len(ct_post),
                    "PTA_pvalue":          round(pta_p, 4) if not np.isnan(pta_p) else "N/A",
                    "PTA_non_rigettata":   str(pta_pass) if pta_pass is not None else "N/A",
                    "delta_DiD_EUR_L":     round(delta, 5),
                    "SE_HC3":              round(se,    5),
                    "CI_95_low":           round(ci_lo, 5),
                    "CI_95_high":          round(ci_hi, 5),
                    "t_stat":              round(t_stat, 3),
                    "p_value":             round(p_val, 6),
                    "R2_OLS":              round(float(ols.rsquared), 3),
                    "H0":                  "RIFIUTATA" if p_val < ALPHA else "non rifiutata",
                    "interpretation_note": note,
                    "famiglia_BH":         "auxiliary_DiD",
                })

# ── BH correction sulla famiglia DiD ─────────────────────────────────────────
if did_rows:
    df_did = pd.DataFrame(did_rows)
    p_arr = df_did["p_value"].values.astype(float)
    reject, p_adj = bh_correction(p_arr, alpha=ALPHA)
    df_did["BH_DiD_reject"]    = reject
    df_did["p_DiD_BH_adjusted"] = np.round(p_adj, 6)
    df_did.to_csv("data/did_results_v2.csv", index=False)
    n_rej = int(reject.sum())
    print(f"\n  Salvato: data/did_results_v2.csv  ({len(df_did)} test DiD)")
    print(f"  BH famiglia DiD: {n_rej}/{len(df_did)} rigettati a α={ALPHA}")
else:
    print("\n  Nessun risultato DiD prodotto.")
    df_did = pd.DataFrame()

# ── Plot DiD — forest plot compatto (v2.1 fix: eliminato spazio vuoto centrale)
if not df_did.empty:
    events_noprel = [e for e, cfg in EVENTS.items() if not cfg["preliminary"]]
    countries     = list(CONTROL_PUMPS.keys())
    n_events      = len(events_noprel)
    n_countries   = len(countries)

    if n_events > 0 and n_countries > 0:
        # Un panel per evento, dentro ogni panel le righe sono paese×carburante
        n_cols   = n_events
        fig, axes = plt.subplots(1, n_cols, figsize=(6 * n_cols, max(4, len(countries) * 2 + 2)),
                                  squeeze=False)
        colors_fuel  = {"Benzina": "#d6604d", "Diesel": "#4393c3"}
        markers_fuel = {"Benzina": "o", "Diesel": "s"}

        for ei, event in enumerate(events_noprel):
            ax        = axes[0][ei]
            ev_label  = event.split("(")[0].strip()
            # Righe del forest plot: per ogni paese × carburante
            # Ordine: Germania/Benzina, Germania/Diesel, Svezia/Benzina, Svezia/Diesel
            rows = []
            for paese in countries:
                sub = df_did[
                    (df_did["Paese_controllo"] == paese) &
                    (df_did["Evento"] == event)
                ].sort_values("Carburante")
                for _, r in sub.iterrows():
                    rows.append({
                        "fuel":    r["Carburante"],
                        "paese":   paese,
                        "delta":   r["delta_DiD_EUR_L"],
                        "lo":      r["CI_95_low"],
                        "hi":      r["CI_95_high"],
                        "pval":    r["p_value"],
                        "pta_ok":  r.get("PTA_non_rigettata", True),
                        "label":   f"{paese} — {r['Carburante']}",
                    })

            n_rows = len(rows)
            y_pos  = np.arange(n_rows - 1, -1, -1, dtype=float)  # alto → basso

            for j, row in enumerate(rows):
                y = y_pos[j]
                d  = row["delta"]
                lo = row["lo"]
                hi = row["hi"]
                pv = row["pval"]
                col = colors_fuel.get(row["fuel"], "grey")
                mk  = markers_fuel.get(row["fuel"], "o")
                # Tratto dell'IC
                ax.plot([lo, hi], [y, y], color=col, lw=2.0, solid_capstyle="round")
                # Punto centrale
                ax.scatter([d], [y], color=col, s=70, zorder=5, marker=mk,
                           edgecolors="white", linewidths=0.6)
                # Stella significatività
                stars_s = _stars(pv)
                ax.text(hi + 0.003, y, f" {stars_s}", va="center", fontsize=8,
                        color=col, fontweight="bold")
                # PTA: cerchio aperto se violata
                if row["pta_ok"] is False or str(row["pta_ok"]).lower() == "false":
                    ax.scatter([d], [y], s=160, facecolors="none",
                               edgecolors="darkorange", lw=1.5, zorder=6,
                               label="_nolegend_")

            ax.axvline(0, color="black", lw=1.0, ls="--", alpha=0.7)
            ax.set_yticks(y_pos)
            ax.set_yticklabels([r["label"] for r in rows], fontsize=9)
            ax.set_xlabel("δ_DiD (EUR/L)", fontsize=9)
            ax.set_title(f"{ev_label}\n(IT vs paese controllo)", fontsize=10, fontweight="bold")
            ax.set_ylim([-0.6, n_rows - 0.4])
            ax.grid(True, axis="x", alpha=0.25, linestyle=":")

            # Ombra per δ > 0 (IT margini superiori → evidenza specificità)
            ax.axvspan(0, ax.get_xlim()[1] if ax.get_xlim()[1] > 0 else 0.3,
                       alpha=0.04, color="#c0392b")

        # Legenda comune
        from matplotlib.lines import Line2D
        legend_els = [
            Line2D([0],[0], marker="o", color="#d6604d", lw=2, ms=8, label="Benzina"),
            Line2D([0],[0], marker="s", color="#4393c3", lw=2, ms=8, label="Diesel"),
            Line2D([0],[0], marker="o", color="none", markeredgecolor="darkorange",
                   markeredgewidth=1.5, ms=10, lw=0, label="PTA violata (∘)"),
        ]
        fig.legend(handles=legend_els, loc="lower center", ncol=3, fontsize=9,
                   bbox_to_anchor=(0.5, -0.04))

        fig.suptitle(
            "Difference-in-Differences — δ margine IT relativo al paese controllo\n"
            "Barre = IC 95%.  Linea tratteggiata = 0.  "
            "δ > 0 → IT margini superiori (area rossa).  ∘ = PTA violata.",
            fontsize=10, y=1.02
        )
        fig.tight_layout()
        fig.savefig("plots/04_did_v2.png", dpi=DPI, bbox_inches="tight")
        plt.close(fig)
        print("  Plot: plots/04_did_v2.png")


# ─────────────────────────────────────────────────────────────────────────────
# §2. GRANGER CAUSALITY  [esplorativo — non in BH]
# ─────────────────────────────────────────────────────────────────────────────

print("\n" + "=" * 70)
print("§2. GRANGER — Brent → prezzi pompa  [ESPLORATIVO — no BH]")
print("=" * 70)

granger_rows: list[dict] = []

GRANGER_LAGS = [1, 2, 3, 4]   # settimane

if merged is not None:
    for fuel in FUELS:
        price_col = "log_benzina" if fuel == "Benzina" else "log_diesel"
        if "log_brent" not in merged.columns or price_col not in merged.columns:
            print(f"  {fuel}: colonne log non trovate — skip")
            continue

        df_g = merged[["log_brent", price_col]].dropna().copy()
        if len(df_g) < 20:
            print(f"  {fuel}: dati insufficienti per Granger")
            continue

        try:
            res_g = grangercausalitytests(
                df_g[[price_col, "log_brent"]].values,
                maxlag=max(GRANGER_LAGS),
                verbose=False
            )
            for lag in GRANGER_LAGS:
                if lag not in res_g:
                    continue
                # Usa test F (più robusto per piccoli campioni)
                f_stat = res_g[lag][0]["ssr_ftest"][0]
                f_p    = res_g[lag][0]["ssr_ftest"][1]
                granger_rows.append({
                    "Carburante":  fuel,
                    "lag_settimane": lag,
                    "lag_giorni":  lag * 7,
                    "F_stat":      round(f_stat, 4),
                    "p_value":     round(f_p, 6),
                    "significativo": f_p < ALPHA,
                    "tipo":        "esplorativo",
                    "nota":        "Non in famiglia BH — evidenza ausiliaria direzione causalità",
                })
                print(f"  {fuel:<8} lag={lag}w  F={f_stat:.3f}  p={f_p:.4f}{_stars(f_p)}")
        except Exception as e:
            print(f"  {fuel}: errore Granger: {e}")

if granger_rows:
    pd.DataFrame(granger_rows).to_csv("data/granger_results_v2.csv", index=False)
    print(f"\n  Salvato: data/granger_results_v2.csv ({len(granger_rows)} test)")
else:
    print("  Nessun risultato Granger prodotto.")


# ─────────────────────────────────────────────────────────────────────────────
# §3. ROCKETS & FEATHERS  [esplorativo — non in BH]
# ─────────────────────────────────────────────────────────────────────────────

print("\n" + "=" * 70)
print("§3. ROCKETS & FEATHERS — asimmetria rialzo/ribasso  [ESPLORATIVO]")
print("=" * 70)

rf_rows: list[dict] = []

if merged is not None:
    brent_col = "brent_eur"
    if brent_col not in merged.columns:
        brent_col = "brent_7d_eur"

    for fuel in FUELS:
        price_col = "benzina_eur_l" if fuel == "Benzina" else "diesel_eur_l"
        if price_col not in merged.columns or brent_col not in merged.columns:
            print(f"  {fuel}: colonne prezzo non trovate — skip")
            continue

        df_rf = merged[[brent_col, price_col]].dropna().copy()
        df_rf["delta_brent"] = df_rf[brent_col].diff()
        df_rf["delta_price"] = df_rf[price_col].diff()
        df_rf = df_rf.dropna()

        df_rf["up"]   = df_rf["delta_brent"].clip(lower=0)
        df_rf["down"] = df_rf["delta_brent"].clip(upper=0)

        if len(df_rf) < 15:
            print(f"  {fuel}: dati insufficienti per R&F")
            continue

        try:
            X = sm.add_constant(df_rf[["up", "down"]].values)
            ols_rf = sm.OLS(df_rf["delta_price"].values, X).fit(
                cov_type="HAC", cov_kwds={"maxlags": 4}
            )
            beta_up   = float(ols_rf.params[1])
            beta_down = float(ols_rf.params[2])
            se_up     = float(ols_rf.bse[1])
            se_down   = float(ols_rf.bse[2])
            rf_index  = beta_up / abs(beta_down) if abs(beta_down) > 1e-9 else np.nan

            # Test asimmetria: H₀: β_up = β_down
            # R_matrix: [0, 1, -1]
            try:
                t_asym = float(
                    (beta_up - abs(beta_down)) /
                    np.sqrt(se_up**2 + se_down**2)
                )
                p_asym = 2 * (1 - stats.t.cdf(abs(t_asym), df=len(df_rf) - 3))
            except Exception:
                t_asym, p_asym = np.nan, np.nan

            rho_dw = durbin_watson(ols_rf.resid)
            rho_hat = 1 - rho_dw / 2   # approssimazione da DW

            print(
                f"  {fuel:<8}: β_up={beta_up:+.4f}  β_down={beta_down:+.4f}  "
                f"R&F={rf_index:.3f}  t_asym={t_asym:.3f}  p={p_asym:.4f}{_stars(p_asym)}"
            )

            rf_rows.append({
                "Carburante":    fuel,
                "beta_up":       round(beta_up, 5),
                "SE_up_HAC":     round(se_up, 5),
                "beta_down":     round(beta_down, 5),
                "SE_down_HAC":   round(se_down, 5),
                "RF_index":      round(rf_index, 4) if not np.isnan(rf_index) else "N/A",
                "t_asimmetria":  round(t_asym, 3) if not np.isnan(t_asym) else "N/A",
                "p_asimmetria":  round(p_asym, 6) if not np.isnan(p_asym) else "N/A",
                "rho_hat_DW":    round(rho_hat, 3),
                "n_obs":         len(df_rf),
                "tipo":          "esplorativo",
                "nota":          "HAC maxlags=4 (fixed). Non in BH.",
            })
        except Exception as e:
            print(f"  {fuel}: errore R&F: {e}")

if rf_rows:
    pd.DataFrame(rf_rows).to_csv("data/rockets_feathers_v2.csv", index=False)
    print(f"\n  Salvato: data/rockets_feathers_v2.csv ({len(rf_rows)} stime)")

    # Plot R&F
    if merged is not None and len(rf_rows) == 2:
        fig, axes = plt.subplots(1, 2, figsize=(12, 5))
        colors_rf = {"Benzina": "#d6604d", "Diesel": "#4393c3"}
        for i, (row, fuel) in enumerate(zip(rf_rows, FUELS)):
            ax     = axes[i]
            price_col = "benzina_eur_l" if fuel == "Benzina" else "diesel_eur_l"
            df_s   = merged[[brent_col, price_col]].dropna().copy()
            delta_b = df_s[brent_col].diff().dropna()
            delta_p = df_s[price_col].diff().dropna()
            ax.scatter(
                delta_b, delta_p,
                alpha=0.4, s=18,
                color=colors_rf[fuel], label=fuel
            )
            x_line = np.linspace(delta_b.min(), delta_b.max(), 100)
            y_up   = row["beta_up"]   * np.maximum(x_line, 0)
            y_down = row["beta_down"] * np.minimum(x_line, 0)
            ax.plot(x_line, y_up + y_down, color="black", lw=1.5,
                    label=f"R&F={row['RF_index']}  p={row['p_asimmetria']}")
            ax.axhline(0, color="grey", lw=0.5, ls="--")
            ax.axvline(0, color="grey", lw=0.5, ls="--")
            ax.set_xlabel("ΔBrent (EUR/L)")
            ax.set_ylabel(f"Δ{fuel} (EUR/L)")
            ax.set_title(f"Rockets & Feathers — {fuel}")
            ax.legend(fontsize=8)
        fig.tight_layout()
        fig.savefig("plots/04_rf_v2.png", dpi=DPI, bbox_inches="tight")
        plt.close(fig)
        print("  Plot: plots/04_rf_v2.png")
else:
    print("  Nessun risultato R&F prodotto.")


# ─────────────────────────────────────────────────────────────────────────────
# §4. WINDFALL PROFITS  [descrittivo — no test, no BH]
# Miglioria v2: volumi corretti per trend lineare consumi
# ─────────────────────────────────────────────────────────────────────────────

print("\n" + "=" * 70)
print("§4. WINDFALL PROFITS — stima con correzione trend consumi")
print("    Miglioria v2: volumi × (1 + trend)^(anno-2022)  invece di fissi 2022")
print("=" * 70)

windfall_rows: list[dict] = []

if merged is not None:
    # Carica δ_margine da tabella confirmatory v2 (o da calcolo diretto)
    conf_path = "data/confirmatory_pvalues_v2.csv"
    table2_path = "data/table2_margin_anomaly.csv"

    # Prova a caricare delta_margine da confirmatory v2
    delta_lookup: dict[tuple, float] = {}
    if os.path.exists(conf_path):
        df_conf = pd.read_csv(conf_path)
        for _, row in df_conf.iterrows():
            if str(row.get("split_type", "")).lower() == "shock_hard":
                key = (str(row["evento"]), str(row["carburante"]))
                if "delta_margine" in df_conf.columns:
                    delta_lookup[key] = float(row["delta_margine"])

    # Fallback: usa tabella v1 se disponibile
    if not delta_lookup and os.path.exists(table2_path):
        df_t2 = pd.read_csv(table2_path)
        for _, row in df_t2.iterrows():
            fuel = str(row.get("Carburante", row.get("Serie", "?")))
            event = str(row.get("Evento", "?"))
            for dcol in ["delta_margine_eur", "Delta_margine", "delta_margine"]:
                if dcol in df_t2.columns:
                    try:
                        delta_lookup[(event, fuel)] = float(row[dcol])
                    except (ValueError, TypeError):
                        pass
                    break

    for event_name, ecfg in EVENTS.items():
        if ecfg["preliminary"]:
            continue

        shock     = ecfg["shock"]
        post_end  = ecfg["post_end"]
        post_data = merged.loc[shock:post_end]
        n_weeks   = max(len(post_data), 1)

        # Anno shock per calcolo trend
        shock_year  = shock.year

        for fuel in FUELS:
            mcol = _margine_col(merged, fuel) or FUEL_TO_COL[fuel]
            if mcol not in merged.columns:
                continue

            # δ margine: media post − media 2019
            post_margin  = merged.loc[shock:post_end, mcol].dropna()
            base_2019    = merged.loc["2019-01-01":"2019-12-31", mcol].dropna()

            if len(post_margin) < 3 or len(base_2019) < 3:
                print(f"  {event_name} {fuel}: dati insufficienti per windfall")
                continue

            delta = float(post_margin.mean() - base_2019.mean())

            # Correzione volumi: trend lineare da 2022
            years_since_2022  = shock_year - 2022
            volume_correction = (1 + CONSUMPTION_TREND_PCT_PER_YEAR) ** years_since_2022
            vol_2022          = VOLUME_PROXY_2022[fuel]
            vol_corrected     = vol_2022 * volume_correction

            windfall_EUR      = delta * vol_corrected * n_weeks
            windfall_MLD      = windfall_EUR / 1e9

            # CI bootstrap sul δ_margine → CI windfall
            np.random.seed(42)
            boot_deltas = [
                (np.random.choice(post_margin.values, len(post_margin), replace=True).mean()
                 - np.random.choice(base_2019.values, len(base_2019), replace=True).mean())
                for _ in range(5000)
            ]
            ci_lo_delta = np.percentile(boot_deltas, 2.5)
            ci_hi_delta = np.percentile(boot_deltas, 97.5)

            windfall_lo = ci_lo_delta * vol_corrected * n_weeks / 1e9
            windfall_hi = ci_hi_delta * vol_corrected * n_weeks / 1e9

            print(
                f"  {event_name.split('(')[0].strip():<22} | {fuel:<8}: "
                f"δ={delta:+.4f} EUR/L  n={n_weeks}w  "
                f"vol_corr={volume_correction:.3f}  "
                f"windfall={windfall_MLD:+.2f} Mld EUR  "
                f"[{windfall_lo:+.2f}, {windfall_hi:+.2f}]"
            )

            windfall_rows.append({
                "Evento":                 event_name,
                "Carburante":             fuel,
                "n_settimane_post":       n_weeks,
                "shock_year":             shock_year,
                "delta_margine_EUR_L":    round(delta, 5),
                "CI_boot_95_lo":          round(ci_lo_delta, 5),
                "CI_boot_95_hi":          round(ci_hi_delta, 5),
                "volume_proxy_2022_L_w":  round(vol_2022, 0),
                "trend_correzione_v2":    round(volume_correction, 4),
                "volume_corretto_L_w":    round(vol_corrected, 0),
                "windfall_MLD_EUR":       round(windfall_MLD, 3),
                "windfall_CI_lo_MLD":     round(windfall_lo, 3),
                "windfall_CI_hi_MLD":     round(windfall_hi, 3),
                "trend_%_annuo":          f"{CONSUMPTION_TREND_PCT_PER_YEAR*100:.1f}%",
                "nota": (
                    "DESCRITTIVO — nessun p-value. "
                    "Volume corretto per trend EV + efficienza motori dal 2022."
                ),
            })

if windfall_rows:
    pd.DataFrame(windfall_rows).to_csv("data/windfall_v2.csv", index=False)
    print(f"\n  Salvato: data/windfall_v2.csv ({len(windfall_rows)} stime)")

    # ── [FIX critic.4] ANALISI SENSITIVITÀ VOLUMI ±30% ───────────────────────
    # I volumi sono proxy 2022 corretti per trend. L'incertezza sulla stima
    # dei volumi reali è alta (dati MISE/Eni non in licenza libera).
    # Si fornisce un range sensitività ±30% per valutare la robustezza
    # delle stime windfall alle ipotesi sui volumi.
    # Rif: critique metodologica 27-apr-2026, punto 4.
    print("\n  Sensitività windfall a ipotesi sui volumi [FIX critic.4]:")
    print(f"  {'Evento':<26} {'Fuel':<8} {'Vol -30%':>10} {'Vol base':>10} {'Vol +30%':>10}  (Mld EUR)")
    print("  " + "─" * 72)
    sens_rows: list[dict] = []
    for row_w in windfall_rows:
        vol_base  = row_w["volume_corretto_L_w"]
        delta_val = row_w["delta_margine_EUR_L"]
        n_wks_val = row_w["n_settimane_post"]
        ev_s      = row_w["Evento"].split("(")[0].strip()
        fuel_s    = row_w["Carburante"]
        wf_base   = row_w["windfall_MLD_EUR"]
        for mult, label in [(0.70, "-30%"), (1.00, "base"), (1.30, "+30%")]:
            wf_s = delta_val * vol_base * mult * n_wks_val / 1e9
            sens_rows.append({
                "Evento":         row_w["Evento"],
                "Carburante":     fuel_s,
                "vol_scenario":   label,
                "vol_mult":       mult,
                "windfall_MLD":   round(wf_s, 3),
            })
        wf_lo = delta_val * vol_base * 0.70 * n_wks_val / 1e9
        wf_hi = delta_val * vol_base * 1.30 * n_wks_val / 1e9
        print(f"  {ev_s:<26} {fuel_s:<8} {wf_lo:>+10.2f} {wf_base:>+10.2f} {wf_hi:>+10.2f}")
    if sens_rows:
        pd.DataFrame(sens_rows).to_csv("data/windfall_sensitivity_v2.csv", index=False)
        print(f"\n  Salvato: data/windfall_sensitivity_v2.csv ({len(sens_rows)} righe)")

    # Plot windfall
    df_wf = pd.DataFrame(windfall_rows)
    if not df_wf.empty:
        events_u = df_wf["Evento"].unique()
        n_ev     = len(events_u)
        fig, axes = plt.subplots(1, n_ev, figsize=(5 * n_ev, 4), squeeze=False)
        colors_wf = {"Benzina": "#d6604d", "Diesel": "#4393c3"}
        for i, ev in enumerate(events_u):
            ax   = axes[0][i]
            sub  = df_wf[df_wf["Evento"] == ev]
            xs   = np.arange(len(sub))
            bars = ax.bar(
                xs,
                sub["windfall_MLD_EUR"].values,
                color=[colors_wf.get(f, "grey") for f in sub["Carburante"]],
                alpha=0.8
            )
            # IC bootstrap
            ax.errorbar(
                xs,
                sub["windfall_MLD_EUR"].values,
                yerr=[
                    sub["windfall_MLD_EUR"].values - sub["windfall_CI_lo_MLD"].values,
                    sub["windfall_CI_hi_MLD"].values - sub["windfall_MLD_EUR"].values,
                ],
                fmt="none", color="black", capsize=4
            )
            ax.set_xticks(xs)
            ax.set_xticklabels(sub["Carburante"].values)
            ax.axhline(0, color="black", lw=0.6, ls="--")
            ax.set_ylabel("Windfall (Mld EUR)")
            ax.set_title(f"Windfall — {ev.split('(')[0].strip()}\n(volumi v2 corretti per trend)")
        fig.tight_layout()
        fig.savefig("plots/04_windfall_v2.png", dpi=DPI, bbox_inches="tight")
        plt.close(fig)
        print("  Plot: plots/04_windfall_v2.png")
else:
    print("  Nessun risultato windfall prodotto.")


# ─────────────────────────────────────────────────────────────────────────────
# §5. HORMUZ (Feb 2026) — PRELIMINARE  [fuori da BH]
# ─────────────────────────────────────────────────────────────────────────────

print("\n" + "=" * 70)
print("§5. HORMUZ — ANALISI PRELIMINARE  [FLAG: instabile, NO p-value in BH]")
print("=" * 70)

hormuz_cfg    = EVENTS.get("Hormuz (Feb 2026)", {})
hormuz_shock  = hormuz_cfg.get("shock", pd.Timestamp("2026-02-01"))
hormuz_rows: list[dict] = []

if merged is not None:
    post_data_hormuz = merged.loc[hormuz_shock:]
    n_weeks_hormuz   = len(post_data_hormuz)
    print(f"  Settimane post-Hormuz disponibili: {n_weeks_hormuz}")
    print(f"  Nota: {hormuz_cfg.get('preliminary_note', 'Analisi preliminare')}")

    for fuel in FUELS:
        mcol = _margine_col(merged, fuel) or FUEL_TO_COL[fuel]
        if mcol not in merged.columns:
            continue

        post_m  = merged.loc[hormuz_shock:, mcol].dropna()
        base_19 = merged.loc["2019-01-01":"2019-12-31", mcol].dropna()
        pre_m   = merged.loc[hormuz_cfg.get("pre_start", "2025-06-01"):hormuz_shock, mcol].dropna()

        if len(post_m) < 2 or len(base_19) < 3:
            print(f"  {fuel}: dati insufficienti")
            continue

        delta_vs_2019 = float(post_m.mean() - base_19.mean())
        delta_pre_post = float(post_m.mean() - pre_m.mean()) if len(pre_m) >= 3 else np.nan
        std_2019       = float(base_19.std())
        z_score        = delta_vs_2019 / std_2019 if std_2019 > 0 else np.nan

        # τ_margin descrittivo (SOLO informativo — nessun test)
        # Trova il punto di massima variazione nella serie post-shock
        if len(post_m) >= 4:
            best_k = int(np.argmax(np.abs(
                [post_m.iloc[k:].mean() - post_m.iloc[:k].mean()
                 for k in range(2, len(post_m) - 1)]
            )) + 2)
            tau_margin_desc = str(post_m.index[best_k].date()) if best_k < len(post_m) else "N/A"
        else:
            tau_margin_desc = "N/A (troppo pochi dati)"

        print(
            f"  {fuel:<8}: δ_vs2019={delta_vs_2019:+.4f}  z={z_score:.2f}  "
            f"n_post={len(post_m)}  "
            f"[INSTABILE — {n_weeks_hormuz} sett. post-shock]"
        )

        hormuz_rows.append({
            "Evento":              "Hormuz (Feb 2026)",
            "Carburante":          fuel,
            "n_settimane_post":    len(post_m),
            "delta_vs_2019":       round(delta_vs_2019, 5),
            "delta_pre_post":      round(delta_pre_post, 5) if not np.isnan(delta_pre_post) else "N/A",
            "std_2019":            round(std_2019, 5),
            "z_score_vs_2019":     round(z_score, 3) if not np.isnan(z_score) else "N/A",
            "tau_margin_descrittivo": tau_margin_desc,
            "flag_preliminare":    True,
            "flag_in_BH":          False,
            "nota": (
                f"PRELIMINARE: solo {n_weeks_hormuz} settimane post-shock. "
                "Nessun p-value in BH globale. "
                "Stime altamente instabili — da aggiornare quando n ≥ 20 settimane."
            ),
        })

if hormuz_rows:
    pd.DataFrame(hormuz_rows).to_csv("data/hormuz_preliminary.csv", index=False)
    print(f"\n  Salvato: data/hormuz_preliminary.csv ({len(hormuz_rows)} righe)")
    print("  ⚠ Ricordare: questi dati NON entrano nella BH globale (05_global_corrections_v2.py)")
else:
    print("  Nessun dato Hormuz disponibile.")


# ─────────────────────────────────────────────────────────────────────────────
# SOMMARIO
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 70)
print("  SOMMARIO 04_auxiliary_evidence_v2.py")
print("=" * 70)
print(f"  §1 DiD:       {len(did_rows)} test → data/did_results_v2.csv"
      + (f"  ({int(df_did['BH_DiD_reject'].sum())} BH-rigettati)" if not df_did.empty else ""))
print(f"  §2 Granger:   {len(granger_rows)} test → data/granger_results_v2.csv  [esplorativo]")
print(f"  §3 R&F:       {len(rf_rows)} stime  → data/rockets_feathers_v2.csv  [esplorativo]")
print(f"  §4 Windfall:  {len(windfall_rows)} stime → data/windfall_v2.csv  [descrittivo]")
print(f"  §5 Hormuz:    {len(hormuz_rows)} righe → data/hormuz_preliminary.csv  [preliminare, no BH]")
print()
print("  Il file data/did_results_v2.csv viene letto da 05_global_corrections_v2.py")
print("  come FAMIGLIA BH AUSILIARIA separata dai test confirmativi sul margine.")
print("=" * 70)
print("\nScript 04 completato.")

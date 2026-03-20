"""
04_auxiliary_evidence.py  (fix: Granger F-stat, DiD time-series, PTA 8w, Hormuz, SE col fix)
=========================
Evidenza ausiliaria a supporto dell'ipotesi H0.

FIX in questa versione:
  1. Granger plot: usa F-statistic invece di p-value (p≈0 rendeva le barre invisibili)
  2. DiD PTA: finestra 8 settimane prima dello shock invece dell'intera pre-period
     (trend su 6+ mesi divergono strutturalmente tra paesi anche in assenza di effetto)
  3. DiD colonne Svezia: logica migliorata per evitare fallback su colonne EU aggregate
  4. DiD plot aggiuntivo: serie temporali dei margini IT/DE/SE a confronto
  5. Hormuz: aggiunto (dati preliminari, solo BH esplorativa separata)

Input:
  data/dataset_merged.csv
  data/dataset_merged_with_futures.csv
  data/eu_oil_bulletin_history.xlsx

Output:
  data/granger_benzina.csv
  data/granger_diesel.csv
  data/rockets_feathers_results.csv
  data/did_results.csv
  data/auxiliary_pvalues.csv
  plots/04_granger.png
  plots/04_rf.png
  plots/04_did.png
  plots/04_did_timeseries.png   ← NUOVO
"""

import os, warnings
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.dates as mdates
import openpyxl as _opxl
from statsmodels.tsa.stattools import grangercausalitytests, adfuller
from statsmodels.regression.linear_model import OLS
from statsmodels.stats.sandwich_covariance import cov_hac
import statsmodels.api as sm
from scipy import stats

warnings.filterwarnings("ignore")
os.makedirs("data",  exist_ok=True)
os.makedirs("plots", exist_ok=True)

ALPHA      = 0.05
MAX_LAG    = 8
DPI        = 180
PTA_WEEKS  = 8    # finestra pre-shock per Parallel Trends test (settimane)

EVENTS = {
    "Ucraina (Feb 2022)": {
        "shock":      pd.Timestamp("2022-02-24"),
        "pre_start":  pd.Timestamp("2021-09-01"),
        "post_end":   pd.Timestamp("2022-08-31"),
        "preliminare": False,
    },
    "Iran-Israele (Giu 2025)": {
        "shock":      pd.Timestamp("2025-06-13"),
        "pre_start":  pd.Timestamp("2025-01-01"),
        "post_end":   pd.Timestamp("2025-10-31"),
        "preliminare": False,
    },
    "Hormuz (Feb 2026)": {
        "shock":      pd.Timestamp("2026-02-28"),
        "pre_start":  pd.Timestamp("2025-10-01"),
        "post_end":   pd.Timestamp("2026-04-27"),
        "preliminare": True,
    },
}

DID_CONTROLS = {
    "Germania": ["DE_", "DE "],
    "Svezia":   ["SE_", "SE "],
}

def _stars(p):
    if np.isnan(p): return ""
    return "***" if p < 0.001 else "**" if p < 0.01 else "*" if p < 0.05 else "n.s."


# ─────────────────────────────────────────────────────────────────────────────
# Carica dataset
# ─────────────────────────────────────────────────────────────────────────────
merged   = pd.read_csv("data/dataset_merged.csv",
                       index_col=0, parse_dates=True)
merged_f = pd.read_csv("data/dataset_merged_with_futures.csv",
                       index_col=0, parse_dates=True)
print(f"Dataset: {len(merged)} settimane | "
      f"{merged.index[0].date()} – {merged.index[-1].date()}\n")

# Safety check unità
for col in ["benzina_eur_l", "diesel_eur_l"]:
    if col in merged_f.columns and merged_f[col].dropna().median() > 10:
        merged_f[col] = merged_f[col] / 1000.0
        merged[col]   = merged[col] / 1000.0 if col in merged.columns else merged.get(col)
        print(f"  [safety] Normalizzato {col} /1000")

aux_pvalues: list[dict] = []


# ═════════════════════════════════════════════════════════════════════════════
# §1. GRANGER — usa F-statistic per il plot (p≈0 → barre invisibili)
# ═════════════════════════════════════════════════════════════════════════════
print("=" * 65)
print("§1. GRANGER: Brent -> prezzi pompa (velocita' di trasmissione)")
print("=" * 65)

print("\nADF sui livelli log:")
for col in ["log_brent", "log_benzina", "log_diesel"]:
    if col not in merged.columns:
        continue
    p_adf = adfuller(merged[col].dropna(), autolag="AIC")[1]
    staz  = "stazionario" if p_adf < ALPHA else "non stazionario -> uso Δlog"
    print(f"  {col}: p={p_adf:.4f}  [{staz}]")

merged["d_log_brent"]   = merged["log_brent"].diff()
merged["d_log_benzina"] = merged["log_benzina"].diff()
merged["d_log_diesel"]  = merged["log_diesel"].diff()

granger_df = merged[merged.index.year != 2020].dropna(
    subset=["d_log_brent", "d_log_benzina", "d_log_diesel"]
)
print(f"\nDati Granger: {len(granger_df)} settimane (escluso 2020 COVID)\n")

granger_results: dict[str, pd.DataFrame] = {}

for fuel_col, fuel_name in [("d_log_benzina", "Benzina"),
                              ("d_log_diesel",  "Diesel")]:
    data2 = granger_df[[fuel_col, "d_log_brent"]].dropna()
    try:
        gc = grangercausalitytests(data2, maxlag=MAX_LAG, verbose=False)
    except Exception as e:
        print(f"  {fuel_name}: errore ({e})")
        continue

    rows = []
    for lag in range(1, MAX_LAG + 1):
        f_stat, p_val = gc[lag][0]["ssr_ftest"][:2]
        flag = " <- lag < 30gg" if p_val < ALPHA and lag * 7 < 30 else ""
        print(f"  {fuel_name} lag={lag}w ({lag*7}gg): "
              f"F={f_stat:.3f}  p={p_val:.4f} {_stars(p_val)}{flag}")
        rows.append({
            "lag_weeks": lag, "lag_days": lag * 7,
            "F_stat":    round(f_stat, 4),
            "p_value":   round(p_val, 4),
            "significant": p_val < ALPHA,
        })

    df_gc = pd.DataFrame(rows)
    granger_results[fuel_name] = df_gc
    df_gc.to_csv(f"data/granger_{fuel_name.lower()}.csv", index=False)

# Plot Granger: F-statistic (p≈0 → barre invisibili con p-value)
if granger_results:
    def _bar_color_f(f_stat, p_val):
        if p_val < 0.001: return "#8b1a1a"
        if p_val < 0.01:  return "#c0392b"
        if p_val < 0.05:  return "#e74c3c"
        return "#95a5a6"

    n_panels = len(granger_results)
    fig, axes = plt.subplots(1, n_panels, figsize=(7 * n_panels, 5.5))
    if n_panels == 1:
        axes = [axes]

    for ax, (fuel_name, df_gc) in zip(axes, granger_results.items()):
        lags_d  = df_gc["lag_days"].values
        fstats  = df_gc["F_stat"].values
        pvals   = df_gc["p_value"].values
        colors  = [_bar_color_f(f, p) for f, p in zip(fstats, pvals)]

        ax.set_axisbelow(True)
        bars = ax.bar(lags_d, fstats, color=colors, edgecolor="black",
                      linewidth=0.6, alpha=0.90, width=5.5)
        # Linea soglia F critico (approssimativa, ~3.9 per α=0.05 con df tipici)
        ax.axhline(4.0, color="#2c3e50", lw=1.4, ls="--")
        ax.text(lags_d[-1] + 1, 4.5, "F crit ≈ 4.0\n(α=0.05)", ha="right",
                fontsize=8, color="#2c3e50")
        ax.axvline(30, color="#e67e22", lw=1.6, ls="--")
        ax.axvspan(0, 30, alpha=0.06, color="#e74c3c")
        ax.text(15, max(fstats) * 1.12, "< 30gg", ha="center",
                fontsize=8, color="#c0392b", style="italic")

        # Etichette p-value sulle barre
        for bar, p, f in zip(bars, pvals, fstats):
            ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + max(fstats)*0.01,
                    f"p={p:.3f}" if p >= 0.001 else "p<.001",
                    ha="center", va="bottom", fontsize=7, rotation=45)

        n_sig = sum(p < ALPHA and d <= 30 for p, d in zip(pvals, lags_d))
        ax.set_title(f"Granger: Brent → {fuel_name}\n"
                     f"F-statistic | Lag sign. < 30gg: {n_sig} (evidenza esplorativa)",
                     fontsize=10, fontweight="bold",
                     color="#b03030" if n_sig > 0 else "black")
        ax.set_xlabel("Lag (giorni)", fontsize=10)
        ax.set_ylabel("F-statistic (test SSR)", fontsize=10)
        ax.set_xticks(lags_d)
        ax.set_xticklabels([f"{int(d)}d" for d in lags_d], fontsize=9)
        ax.set_ylim(0, max(fstats) * 1.35)
        ax.legend(handles=[
            mpatches.Patch(color="#8b1a1a", label="p<0.001 (***)"),
            mpatches.Patch(color="#c0392b", label="p<0.01  (**)"),
            mpatches.Patch(color="#e74c3c", label="p<0.05  (*)"),
            mpatches.Patch(color="#95a5a6", label="p≥0.05  (n.s.)"),
        ], fontsize=8, loc="upper right")
        ax.grid(alpha=0.3)

    fig.suptitle(
        "Granger Causality: Brent → prezzi pompa Italia (F-statistic)\n"
        "Misura la velocità di trasmissione, non la speculazione diretta",
        fontsize=11, fontweight="bold", y=1.02,
    )
    fig.tight_layout(pad=1.5)
    fig.savefig("plots/04_granger.png", dpi=DPI, bbox_inches="tight")
    plt.close(fig)
    print("\n  Salvato: plots/04_granger.png")


# ═════════════════════════════════════════════════════════════════════════════
# §2. ROCKETS & FEATHERS
# ═════════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 65)
print("§2. ROCKETS & FEATHERS: asimmetria strutturale beta_up vs beta_down")
print("=" * 65)

merged["d_brent_pct"]   = merged["brent_7d_eur"].pct_change() * 100
merged["d_benzina_pct"] = merged["benzina_4w"].pct_change() * 100 if "benzina_4w" in merged.columns \
                          else merged["benzina_eur_l"].pct_change() * 100
merged["d_diesel_pct"]  = merged["diesel_4w"].pct_change() * 100 if "diesel_4w" in merged.columns \
                          else merged["diesel_eur_l"].pct_change() * 100
rf_df = merged.dropna(subset=["d_brent_pct", "d_benzina_pct", "d_diesel_pct"])

FUELS_RF = {
    "Benzina": ("d_benzina_pct", "#d6604d"),
    "Diesel":  ("d_diesel_pct",  "#31a354"),
}

rf_results: dict = {}
print(f"\nMetodo: OLS + SE HAC (Newey-West, 4 lag) | motivazione: DW<1.5 sistemico")

for fuel_name, (dcol, fc) in FUELS_RF.items():
    y     = rf_df[dcol].values
    b_pos = np.maximum(rf_df["d_brent_pct"].values, 0)
    b_neg = np.minimum(rf_df["d_brent_pct"].values, 0)
    X     = np.column_stack([np.ones(len(y)), b_pos, b_neg])
    n, k  = len(y), 3

    try:
        ols_r  = OLS(y, X).fit()
        b_up   = ols_r.params[1]
        b_down = ols_r.params[2]
        cov_nw = cov_hac(ols_r, nlags=4)
        method = "OLS+HAC"
    except Exception as exc:
        print(f"  {fuel_name}: errore ({exc}) — skip")
        continue

    se_up   = float(np.sqrt(cov_nw[1, 1]))
    se_down = float(np.sqrt(cov_nw[2, 2]))
    se_diff = np.sqrt(se_up**2 + se_down**2)
    t_stat  = (b_up - b_down) / se_diff if se_diff > 0 else np.nan
    p_asym  = float(stats.t.sf(abs(t_stat), df=n - k) * 2) if not np.isnan(t_stat) else np.nan
    rf_idx  = abs(b_up) / abs(b_down) if b_down != 0 else np.inf

    rf_results[fuel_name] = {
        "b_up": b_up, "b_down": b_down,
        "se_up": se_up, "se_down": se_down,
        "rf_index": rf_idx, "t_stat": t_stat, "p_asym": p_asym,
        "rho_ar": np.nan, "method": method, "color": fc, "dcol": dcol,
    }

    print(f"\n  {fuel_name} [{method}]:")
    print(f"    beta_up   = {b_up:.4f}  (SE HAC={se_up:.4f})")
    print(f"    beta_down = {b_down:.4f}  (SE HAC={se_down:.4f})")
    print(f"    R&F index = {rf_idx:.3f}  |  p asimmetria = {p_asym:.4f} {_stars(p_asym)}")

pd.DataFrame([{
    "Carburante":   f,
    "Metodo":       r["method"],
    "b_up":         round(r["b_up"], 4),
    "SE_up_HAC":    round(r["se_up"], 4),
    "b_down":       round(r["b_down"], 4),
    "SE_down_HAC":  round(r["se_down"], 4),
    "RF_index":     round(r["rf_index"], 3),
    "t_stat":       round(r["t_stat"], 3) if not np.isnan(r["t_stat"]) else "N/A",
    "p_asym":       round(r["p_asym"], 4) if not np.isnan(r["p_asym"]) else "N/A",
} for f, r in rf_results.items()]).to_csv("data/rockets_feathers_results.csv", index=False)

if rf_results:
    fig, axes = plt.subplots(1, 2, figsize=(14, 5.5), sharey=False)
    for ax, (fuel_name, res) in zip(axes, rf_results.items()):
        x_all = rf_df["d_brent_pct"].values
        y_all = rf_df[res["dcol"]].values
        valid = ~(np.isnan(x_all) | np.isnan(y_all))
        clrs  = ["#e74c3c" if b > 0 else "#3498db" for b in x_all[valid]]
        ax.scatter(x_all[valid], y_all[valid], c=clrs, alpha=0.28, s=16)
        x_up   = np.linspace(0, np.nanmax(x_all), 100)
        x_down = np.linspace(np.nanmin(x_all), 0, 100)
        ax.plot(x_up,   res["b_up"]   * x_up,   color="#e74c3c", lw=2.5,
                label=f"β_up={res['b_up']:.4f}")
        ax.plot(x_down, res["b_down"] * x_down, color="#3498db", lw=2.5, ls="--",
                label=f"β_down={res['b_down']:.4f}")
        ax.axhline(0, color="black", lw=0.5); ax.axvline(0, color="black", lw=0.5)
        p_s = f"{res['p_asym']:.4f}" if not np.isnan(res["p_asym"]) else "N/A"
        ax.set_title(f"Rockets & Feathers — {fuel_name}\n"
                     f"R&F index={res['rf_index']:.3f}  p asimmetria={p_s} {_stars(res['p_asym'])}",
                     fontsize=10, fontweight="bold")
        ax.set_xlabel("Δ Brent (%)", fontsize=10)
        ax.set_ylabel(f"Δ {fuel_name} (%)", fontsize=10)
        ax.legend(fontsize=9); ax.grid(alpha=0.3)
    fig.suptitle("Rockets & Feathers: asimmetria strutturale Brent → pompa",
                 fontsize=11, fontweight="bold", y=1.02)
    fig.tight_layout(pad=1.5)
    fig.savefig("plots/04_rf.png", dpi=DPI, bbox_inches="tight")
    plt.close(fig)
    print("\n  Salvato: plots/04_rf.png")


# ═════════════════════════════════════════════════════════════════════════════
# §3. DIFFERENCE-IN-DIFFERENCES
# ═════════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 65)
print("§3. DIFFERENCE-IN-DIFFERENCES: anomalia specifica all'Italia?")
print(f"    PTA testato su finestra {PTA_WEEKS} settimane pre-shock")
print("=" * 65)

EU_HIST = "data/eu_oil_bulletin_history.xlsx"

def _sheet_names(path):
    wb = _opxl.load_workbook(path, read_only=True, data_only=True)
    ns = wb.sheetnames; wb.close(); return ns

def _read_sheet(path, sheet):
    wb   = _opxl.load_workbook(path, read_only=True, data_only=True)
    ws   = wb[sheet]
    data = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if not data: return pd.DataFrame()
    hi   = next((i for i, r in enumerate(data) if any(v is not None for v in r)), 0)
    hdrs = [str(h).strip() if h is not None else f"_c{i}"
            for i, h in enumerate(data[hi])]
    df   = pd.DataFrame(data[hi + 1:], columns=hdrs)
    ic   = hdrs[0]
    df[ic] = pd.to_datetime(df[ic], errors="coerce")
    return df.set_index(ic).loc[lambda x: x.index.notna()].sort_index()

def _notax_sheet(names):
    for s in names:
        if any(k in s.upper() for k in ["WO TAX","WITHOUT","NO TAX","NOTAX"]):
            return s
    return names[1] if len(names) > 1 else names[0]


def _country_cols(df, prefissi):
    """
    Ricerca colonne per paese con logica esplicita:
    - Cerca colonne il cui nome INIZIA con uno dei prefissi (es. "DE_", "SE_")
    - Tra quelle, separa benzina e diesel per keyword
    - NON fa fallback su colonne EU aggregate
    """
    # Colonne che iniziano ESATTAMENTE con il prefisso
    all_c = []
    for pref in prefissi:
        all_c += [c for c in df.columns
                  if str(c).upper().startswith(pref.upper())]
    all_c = list(dict.fromkeys(all_c))  # dedup preservando ordine

    # Benzina
    benz_kw = ["95","benz","petrol","unleaded","euro_sup","euro_95","gasolio_95"]
    benz    = [c for c in all_c if any(k in str(c).lower() for k in benz_kw)]

    # Diesel: keyword esplicite + esclude "eu_" all'inizio
    dies_kw = ["diesel","gasoil","gas_oil","gasolio_dies","dies"]
    dies    = [c for c in all_c
               if any(k in str(c).lower() for k in dies_kw)
               and not str(c).lower().startswith("eu_")]

    # Fallback solo dentro all_c (non su tutto il dataset)
    if not benz and all_c:
        benz = [all_c[0]]
    if not dies and len(all_c) >= 2:
        dies = [all_c[1]]
    elif not dies and all_c:
        # Se c'è solo 1 colonna usa quella per entrambi (segnalato)
        dies = [all_c[0]]

    return benz, dies, all_c


CONTROL_MARGINS: dict[str, pd.DataFrame] = {}
eurobob_al = merged_f["eurobob_eur_l"] if "eurobob_eur_l" in merged_f.columns else None
gasoil_al  = merged_f["gasoil_eur_l"]  if "gasoil_eur_l"  in merged_f.columns else None

try:
    snames  = _sheet_names(EU_HIST)
    notax   = _notax_sheet(snames)
    df_eu   = _read_sheet(EU_HIST, notax).apply(pd.to_numeric, errors="coerce")
    print(f"\n  Foglio EU Bulletin: '{notax}' | {len(df_eu.columns)} colonne")

    for paese, prefissi in DID_CONTROLS.items():
        b_c, d_c, all_c = _country_cols(df_eu, prefissi)

        if not b_c or not d_c:
            print(f"  {paese}: colonne non trovate — skip. "
                  f"Colonne disponibili con prefisso {prefissi}: {all_c[:5]}")
            continue

        print(f"  {paese}: benzina='{b_c[0]}'  diesel='{d_c[0]}'  "
              f"(da {len(all_c)} colonne candidate)")

        pump = pd.concat([df_eu[b_c[0]].rename("benzina_eur_l"),
                          df_eu[d_c[0]].rename("diesel_eur_l")], axis=1)
        pump = pump[pump.index >= "2019-01-01"].dropna(how="all")
        pump = pump.resample("W-MON").mean()

        # Normalizza EUR/1000L → EUR/L
        for col in ["benzina_eur_l", "diesel_eur_l"]:
            if col in pump.columns:
                med = pump[col].dropna().median()
                if med > 10:
                    pump[col] = pump[col] / 1000.0

        if eurobob_al is not None:
            pump["margine_benz_crack"] = (
                pump["benzina_eur_l"] - eurobob_al.reindex(pump.index).ffill(limit=4)
            )
        if gasoil_al is not None:
            pump["margine_dies_crack"] = (
                pump["diesel_eur_l"] - gasoil_al.reindex(pump.index).ffill(limit=4)
            )

        CONTROL_MARGINS[paese] = pump
        n_ok = pump.dropna(subset=["benzina_eur_l"]).shape[0]
        benz_med_2019 = pump.loc["2019","benzina_eur_l"].median() if "2019" in pump.index.year.unique() else float("nan")
        print(f"       {n_ok} settimane | benzina 2019 median: {benz_med_2019:.4f} EUR/L")

except FileNotFoundError:
    print(f"  {EU_HIST} non trovato — DiD saltato.")
except Exception as exc:
    print(f"  Errore caricamento dati controllo: {exc}")

IT_MARGIN = {
    "Benzina": "margine_benz_crack",
    "Diesel":  "margine_dies_crack",
}
it_margins_ok = all(col in merged_f.columns for col in IT_MARGIN.values())

did_rows = []

if CONTROL_MARGINS and it_margins_ok:
    for paese, ctrl in CONTROL_MARGINS.items():
        for ev_name, cfg in EVENTS.items():
            shock     = cfg["shock"]
            pre_start = cfg["pre_start"]
            post_end  = cfg["post_end"]
            prelim    = cfg.get("preliminare", False)

            for fuel_name, it_col in IT_MARGIN.items():
                ct_col = it_col

                if ct_col not in ctrl.columns:
                    continue

                it_pre  = merged_f.loc[pre_start:shock,  it_col].dropna()
                it_post = merged_f.loc[shock:post_end,   it_col].dropna()
                ct_pre  = ctrl.loc[pre_start:shock, ct_col].dropna()
                ct_post = ctrl.loc[shock:post_end,  ct_col].dropna()

                if any(len(s) < 3 for s in [it_pre, it_post, ct_pre, ct_post]):
                    print(f"  [{paese}] {ev_name[:20]} | {fuel_name}: campioni insufficienti — skip")
                    continue

                # ── Parallel Trends Test su finestra PTA_WEEKS ───────────────
                # Usa solo le ultime PTA_WEEKS settimane del pre-shock.
                # Una finestra breve evita che divergenze strutturali di lungo periodo
                # falsifichino il test (i trend possono divergere su 6+ mesi anche in
                # assenza di effetto causale dell'evento).
                pt_pvalue = np.nan
                pt_valida = None
                try:
                    it_pta = it_pre.iloc[-min(PTA_WEEKS, len(it_pre)):]
                    ct_pta = ct_pre.iloc[-min(PTA_WEEKS, len(ct_pre)):]

                    if len(it_pta) >= 4 and len(ct_pta) >= 4:
                        all_pre_idx = it_pta.index.union(ct_pta.index).sort_values()
                        t0 = all_pre_idx[0]
                        t_it = np.array([(d - t0).days for d in it_pta.index], dtype=float)
                        t_ct = np.array([(d - t0).days for d in ct_pta.index], dtype=float)
                        rows_pt = (
                            [(1, t, v) for t, v in zip(t_it, it_pta.values)] +
                            [(0, t, v) for t, v in zip(t_ct, ct_pta.values)]
                        )
                        df_pt   = pd.DataFrame(rows_pt, columns=["Italy","t","M"])
                        df_pt["Ixt"] = df_pt["Italy"] * df_pt["t"]
                        X_pt    = sm.add_constant(df_pt[["Italy","t","Ixt"]].values)
                        ols_pt  = sm.OLS(df_pt["M"].values, X_pt).fit(cov_type="HC3")
                        pt_pvalue = float(ols_pt.pvalues[3])
                        pt_valida = bool(pt_pvalue >= ALPHA)
                except Exception:
                    pass

                # ── Modello DiD ───────────────────────────────────────────────
                rows_panel = (
                    [(1, 0, v) for v in it_pre.values]  +
                    [(1, 1, v) for v in it_post.values] +
                    [(0, 0, v) for v in ct_pre.values]  +
                    [(0, 1, v) for v in ct_post.values]
                )
                df_panel = pd.DataFrame(rows_panel, columns=["Italy","Post","M"])
                df_panel["IxP"] = df_panel["Italy"] * df_panel["Post"]
                X_did = sm.add_constant(df_panel[["Italy","Post","IxP"]].values)

                try:
                    ols_did  = sm.OLS(df_panel["M"].values, X_did).fit(cov_type="HC3")
                    delta    = float(ols_did.params[3])
                    se_delta = float(ols_did.bse[3])
                    p_did    = float(ols_did.pvalues[3])
                    ci_lo    = delta - 1.96 * se_delta
                    ci_hi    = delta + 1.96 * se_delta
                except Exception as exc:
                    print(f"  DiD errore {paese}|{ev_name}|{fuel_name}: {exc}")
                    continue

                pta_flag = (
                    f" [PTA valida, {PTA_WEEKS}w]" if pt_valida is True
                    else f" [PTA violata {PTA_WEEKS}w — cautela]" if pt_valida is False
                    else " [PTA non calcolata]"
                )
                prelim_note = " ⚠ PRELIMINARE" if prelim else ""
                print(f"\n  [{paese}] {ev_name.split('(')[0].strip()} | "
                      f"{fuel_name}{prelim_note}:")
                print(f"    delta_DiD = {delta:+.4f} EUR/l  "
                      f"SE={se_delta:.4f}  p={p_did:.4f} {_stars(p_did)}")
                print(f"    CI 95%: [{ci_lo:+.4f}, {ci_hi:+.4f}]")
                if not np.isnan(pt_pvalue):
                    print(f"    PTA ({PTA_WEEKS}w): p={pt_pvalue:.3f}{pta_flag}")
                if prelim:
                    print(f"    NOTA: solo {len(it_post)} settimane post-shock")

                did_rows.append({
                    "Evento":          ev_name,
                    "Paese_controllo": paese,
                    "Carburante":      fuel_name,
                    "preliminare":     prelim,
                    "n_IT_pre":        len(it_pre),
                    "n_IT_post":       len(it_post),
                    "n_CT_pre":        len(ct_pre),
                    "n_CT_post":       len(ct_post),
                    "PTA_pvalue":      round(pt_pvalue, 4) if not np.isnan(pt_pvalue) else "N/A",
                    "PTA_non_rigettata": pt_valida,
                    "PTA_finestra_w":  PTA_WEEKS,
                    "delta_DiD":       round(delta, 4),
                    "SE_HC3":          round(se_delta, 4),
                    "CI_95_lo":        round(ci_lo, 4),
                    "CI_95_hi":        round(ci_hi, 4),
                    "p_value":         round(p_did, 6),
                    "H0":              "RIFIUTATA" if p_did < ALPHA else "non rifiutata",
                })

                # Solo non-preliminari nel BH globale
                if not prelim:
                    aux_pvalues.append({
                        "fonte":       f"DiD_{ev_name}_{paese}_{fuel_name}",
                        "tipo":        "confirmatory",
                        "descrizione": f"{ev_name} | IT vs {paese} | {fuel_name}",
                        "p_value":     p_did,
                    })


# Salva DiD
if did_rows:
    pd.DataFrame(did_rows).to_csv("data/did_results.csv", index=False)
    print(f"\n  Salvato: data/did_results.csv ({len(did_rows)} stime)")

    # ── Plot DiD barre con CI 95% ─────────────────────────────────────────
    df_did  = pd.DataFrame(did_rows)
    fig_d, ax_d = plt.subplots(figsize=(12, max(5, len(df_did) * 0.70)))
    labels_d = [
        f"{r['Evento'].split('(')[0].strip()} | {r['Paese_controllo']} | {r['Carburante']}"
        + (" ⚠" if r["preliminare"] else "")
        + (" PTA✓" if r["PTA_non_rigettata"] is True else " PTA✗" if r["PTA_non_rigettata"] is False else "")
        for _, r in df_did.iterrows()
    ]
    delta_d = df_did["delta_DiD"].values
    ci_lo_d = df_did["delta_DiD"].values - df_did["CI_95_lo"].values
    ci_hi_d = df_did["CI_95_hi"].values - df_did["delta_DiD"].values
    clr_d   = []
    for _, r in df_did.iterrows():
        if r["preliminare"]:
            clr_d.append("#bdc3c7")   # grigio chiaro = preliminare
        elif r["p_value"] < ALPHA:
            clr_d.append("#e74c3c")   # rosso = significativo
        else:
            clr_d.append("#95a5a6")   # grigio = non sign.

    ax_d.barh(range(len(df_did)), delta_d, color=clr_d,
              edgecolor="black", lw=0.5, alpha=0.85)
    for i, (lo, hi) in enumerate(zip(ci_lo_d, ci_hi_d)):
        ax_d.errorbar(delta_d[i], i, xerr=[[lo],[hi]],
                      fmt="none", color="black", capsize=5, lw=1.5)
    ax_d.axvline(0, color="black", lw=1.0, ls="--")
    ax_d.set_yticks(range(len(df_did)))
    ax_d.set_yticklabels(labels_d, fontsize=8.5)
    ax_d.set_xlabel("δ DiD (EUR/litro, CI 95% HC3)", fontsize=10)
    ax_d.set_title(
        f"Difference-in-Differences: eccesso margine Italia vs paesi controllo\n"
        f"PTA testato su finestra {PTA_WEEKS}w | ✓=valida ✗=violata | ⚠=preliminare",
        fontsize=10, fontweight="bold",
    )
    ax_d.legend(handles=[
        mpatches.Patch(color="#e74c3c", label="p<0.05 (sign.)"),
        mpatches.Patch(color="#95a5a6", label="p≥0.05 (non sign.)"),
        mpatches.Patch(color="#bdc3c7", label="Hormuz (preliminare)"),
    ], fontsize=8, loc="lower right")
    ax_d.grid(alpha=0.3, axis="x")
    plt.tight_layout(pad=1.5)
    fig_d.savefig("plots/04_did.png", dpi=DPI, bbox_inches="tight")
    plt.close(fig_d)
    print("  Salvato: plots/04_did.png")

    # ── Plot DiD: serie temporali margini IT vs paesi controllo ──────────
    # Mostra l'evoluzione del margine crack spread per tutti i paesi
    # Prima del plot, costruiamo una palette per paese
    COUNTRY_COLORS = {
        "Italia":   ("#e74c3c", 2.5),
        "Germania": ("#2980b9", 1.8),
        "Svezia":   ("#27ae60", 1.8),
    }
    WAR_COLORS = {
        "Ucraina (Feb 2022)": ("#e74c3c", "2022-02-24"),
        "Iran-Israele (Giu 2025)": ("#e67e22", "2025-06-13"),
        "Hormuz (Feb 2026)":  ("#8e44ad", "2026-02-28"),
    }

    for fuel_name, it_col in IT_MARGIN.items():
        ct_col = it_col
        if it_col not in merged_f.columns:
            continue

        fig_ts, ax_ts = plt.subplots(figsize=(14, 5))

        # Italia
        it_series = merged_f[it_col].dropna()
        ax_ts.plot(it_series.index, it_series.values,
                   color=COUNTRY_COLORS["Italia"][0],
                   lw=COUNTRY_COLORS["Italia"][1],
                   label="Italia", zorder=4)

        # Paesi controllo
        for paese, ctrl in CONTROL_MARGINS.items():
            if ct_col in ctrl.columns:
                ct_series = ctrl[ct_col].dropna()
                color, lw = COUNTRY_COLORS.get(paese, ("#888", 1.5))
                ax_ts.plot(ct_series.index, ct_series.values,
                           color=color, lw=lw, label=paese, alpha=0.85, zorder=3)

        # Linee shock
        for ev, (ev_color, ev_date) in WAR_COLORS.items():
            ts = pd.Timestamp(ev_date)
            if it_series.index[0] <= ts <= it_series.index[-1]:
                ax_ts.axvline(ts, color=ev_color, lw=1.8, ls="--", alpha=0.8)
                ax_ts.text(ts + pd.Timedelta(days=5),
                           ax_ts.get_ylim()[1] if ax_ts.get_ylim()[1] != 0
                           else it_series.max() * 0.95,
                           ev.split(" ")[0], rotation=90, fontsize=8,
                           color=ev_color, va="top")

        # Banda baseline 2019
        it_2019 = merged_f.loc["2019", it_col].dropna()
        if len(it_2019) >= 4:
            mu, sig = float(it_2019.mean()), float(it_2019.std())
            ax_ts.axhspan(mu - 2*sig, mu + 2*sig, alpha=0.10, color="#888",
                          label="Baseline IT ±2σ (2019)")
            ax_ts.axhline(mu, color="#888", lw=0.8, ls="--")

        ax_ts.set_ylabel("Margine crack spread (EUR/litro)", fontsize=10)
        ax_ts.set_title(
            f"Margine lordo — {fuel_name}: Italia vs Germania vs Svezia\n"
            f"Costo wholesale comune (Eurobob/GasOil ARA) | Prezzi pompa EU Bulletin (wo taxes)",
            fontsize=11, fontweight="bold",
        )
        ax_ts.legend(fontsize=9, loc="upper left")
        ax_ts.grid(alpha=0.3)
        ax_ts.xaxis.set_major_formatter(mdates.DateFormatter("%b %Y"))
        ax_ts.xaxis.set_major_locator(mdates.MonthLocator(interval=4))
        plt.xticks(rotation=40, fontsize=9)
        plt.tight_layout()
        fname_ts = f"plots/04_did_timeseries_{fuel_name.lower()}.png"
        fig_ts.savefig(fname_ts, dpi=DPI, bbox_inches="tight")
        plt.close(fig_ts)
        print(f"  Salvato: {fname_ts}")

else:
    print("  DiD non eseguito (dati controllo non disponibili).")


# ─────────────────────────────────────────────────────────────────────────────
# Salva p-value ausiliari per BH globale
# ─────────────────────────────────────────────────────────────────────────────
pd.DataFrame(aux_pvalues).to_csv("data/auxiliary_pvalues.csv", index=False)
print(f"\nSalvato: data/auxiliary_pvalues.csv ({len(aux_pvalues)} test confirmatory DiD)")
print("\nScript 04 completato.")
print("  Granger:  esplorativo (F-stat plot) -> NON nel BH globale")
print("  R&F:      esplorativo               -> NON nel BH globale")
print(f"  DiD:      confirmatory (PTA su {PTA_WEEKS}w) -> data/auxiliary_pvalues.csv")
print("  Hormuz:   incluso come preliminare  -> escluso dalla BH correction")
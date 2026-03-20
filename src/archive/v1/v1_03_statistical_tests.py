"""
03_statistical_tests.py
========================
Test statistici di supporto all'ipotesi H0 sul margine.
Tutti i test qui sono EVIDENZA AUSILIARIA — il test primario di H0
è in 02_core_analysis.py (anomalia del margine).

CONTENUTO:
  §1. Granger causality: Brent → prezzi pompa
      → testa la VELOCITÀ di trasmissione (non speculazione diretta)
      → lag < 30gg: evidenza di mercato efficiente o speculativo (ambiguo)
  §2. Rockets & Feathers (GLSAR AR(1) + HAC)
      → testa ASIMMETRIA rialzo/ribasso
      → R&F index > 1 + p<0.05: prezzi salgono più veloce di quanto scendono
  §3. KS test, ANOVA, Chow test (prezzi pre vs post shock)
  §4. Cross-Correlation (CCF) e rolling correlation
  §5. Bootstrap 95% CI sul lag D (changepoint vs shock)
  §6. Selezione tipo di regressione (BP, Ljung-Box, DW, AIC/BIC)
  §7. Welch t-test — margine lordo pre vs post shock (test ausiliario)
      → two-sample t-test (varianze non uguali) sul crack spread IT
      → complementa lo z-score BH di script 02 con approccio frequentista classico
  §8. Difference-in-Differences (DiD) — Italia vs Germania
      → carica dati DE dallo stesso file EU Oil Bulletin già scaricato
      → modello OLS: M_{c,t} = α + β1·Italy + β2·Post + δ·(Italy×Post) + ε
      → δ = estimatore DiD: quota anomalia specifica al mercato IT

Output:
  data/granger_benzina.csv
  data/granger_diesel.csv
  data/rockets_feathers_results.csv
  data/ks_results.csv
  data/anova_results.csv
  data/chow_results.csv
  data/bootstrap_ci.csv
  data/regression_selection.csv
  data/ttest_margin.csv
  data/did_results.csv
  plots/03_granger_combined.png
  plots/04_rf_combined.png
  plots/06_statistical_tests.png
  plots/08_regression_selection.png
  plots/09_ttest_did.png
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import matplotlib.patches as mpatches
from statsmodels.tsa.stattools import grangercausalitytests, adfuller, ccf as tsm_ccf
from statsmodels.regression.linear_model import GLSAR, OLS
from statsmodels.stats.sandwich_covariance import cov_hac
from statsmodels.stats.diagnostic import acorr_ljungbox
from statsmodels.stats.stattools import durbin_watson
from scipy import stats
import openpyxl as _opxl
import warnings
warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────────────────────────────────────
# Configurazione
# ─────────────────────────────────────────────────────────────────────────────
MAX_LAG   = 8
ALPHA     = 0.05
DPI       = 180
ROLL_WIN  = 12   # settimane per rolling correlation
MAX_LAG_CCF = 12

EVENTS = {
    "Ucraina (Feb 2022)": {
        "shock":     pd.Timestamp("2022-02-24"),
        "pre_start": pd.Timestamp("2021-09-01"),
        "post_end":  pd.Timestamp("2022-08-31"),
    },
    "Iran-Israele (Giu 2025)": {
        "shock":     pd.Timestamp("2025-06-13"),
        "pre_start": pd.Timestamp("2025-01-01"),
        "post_end":  pd.Timestamp("2025-10-31"),
    },
    "Hormuz (Feb 2026)": {
        "shock":     pd.Timestamp("2026-02-28"),
        "pre_start": pd.Timestamp("2025-09-01"),
        "post_end":  pd.Timestamp("2026-04-01"),
    },
}

FUELS = {"Benzina": "benzina_4w", "Diesel": "diesel_4w"}

WAR_DATES = {ev: cfg["shock"] for ev, cfg in EVENTS.items()}

plt.rcParams.update({
    "font.family": "serif",
    "font.serif":  ["Times New Roman", "DejaVu Serif"],
    "axes.spines.top":   False,
    "axes.spines.right": False,
    "axes.grid":   True,
    "grid.color":  "#e0e0e0",
    "grid.linewidth": 0.5,
    "figure.dpi":  DPI,
})

# ─────────────────────────────────────────────────────────────────────────────
# Carica dataset
# ─────────────────────────────────────────────────────────────────────────────
merged = pd.read_csv("data/dataset_merged.csv", index_col=0, parse_dates=True)
merged.dropna(inplace=True)
print(f"Dataset: {len(merged)} settimane | "
      f"{merged.index[0].date()} → {merged.index[-1].date()}\n")


# ─────────────────────────────────────────────────────────────────────────────
# §1. GRANGER CAUSALITY: Brent → prezzi pompa
#     H0: Brent non Granger-causa il prezzo pompa al lag k
#     Rifiuto H0 a lag < 30gg → trasmissione Brent→pompa rapida
#     NOTA: trasmissione rapida è consistente sia con mercato efficiente
#     che con comportamento speculativo. Test AUSILIARIO.
# ─────────────────────────────────────────────────────────────────────────────
print("=" * 65)
print("§1. GRANGER CAUSALITY: Brent → prezzi pompa")
print("    (test ausiliario — vedi nota metodologica nel docstring)")
print("=" * 65)

print("\nADF — stazionarietà:")
for col in ["log_brent", "log_benzina", "log_diesel"]:
    if col in merged.columns:
        p_adf = adfuller(merged[col].dropna(), autolag="AIC")[1]
        print(f"  ADF {col}: p={p_adf:.4f} "
              f"{'[stazionario]' if p_adf < 0.05 else '[non stazionario → uso Δlog]'}")

merged["d_log_brent"]   = merged["log_brent"].diff()
merged["d_log_benzina"] = merged["log_benzina"].diff()
merged["d_log_diesel"]  = merged["log_diesel"].diff()
merged_d = merged.dropna()

# ── Filtro 2020 per Granger ───────────────────────────────────────────────────
# Il 2020 (COVID-19) ha distorto strutturalmente il mercato energetico:
# WTI negativo ad aprile, domanda collassata ~25%, prezzi pompa sganciati
# dalla dinamica normale Brent→distribuzione. Includerlo nel Granger
# introdurrebbe un regime non-stazionario estraneo alla domanda di ricerca
# (speculazione in condizioni normali di mercato).
# Tutte le altre analisi (KS, ANOVA, Chow, DiD) usano finestre per-evento
# che partono da settembre 2021 e non sono interessate da questo filtro.
# Rif: Baumeister & Kilian (2020) sull'eccezionalità del COVID per i mercati oil.
merged_granger = merged_d[merged_d.index.year != 2020].copy()
n_removed = len(merged_d) - len(merged_granger)
print(f"  [Granger] Escluso 2020 COVID: {n_removed} settimane rimosse "
      f"({len(merged_granger)} settimane usate per la stima)")

granger_results = {}


def _stars(p):
    return "***" if p < 0.001 else "**" if p < 0.01 else "*" if p < 0.05 else "(n.s.)"


def _bar_color(p):
    if p < 0.001: return "#8b1a1a"
    if p < 0.01:  return "#c0392b"
    if p < 0.05:  return "#e74c3c"
    return "#95a5a6"


for fuel_col, fuel_name in [("d_log_benzina", "Benzina"), ("d_log_diesel", "Diesel")]:
    data2 = merged_granger[[fuel_col, "d_log_brent"]].dropna()
    try:
        gc = grangercausalitytests(data2, maxlag=MAX_LAG, verbose=False)
    except Exception as e:
        print(f"  Errore {fuel_name}: {e}")
        continue
    rows = []
    for lag in range(1, MAX_LAG + 1):
        f_stat, p_val = gc[lag][0]["ssr_ftest"][:2]
        rows.append({
            "lag_weeks": lag, "lag_days": lag * 7,
            "F_stat": round(f_stat, 4), "p_value": round(p_val, 4),
            "significant": p_val < ALPHA,
        })
        print(f"  {fuel_name} lag={lag}w ({lag*7}gg): F={f_stat:.3f} "
              f"p={p_val:.4f} {_stars(p_val)}"
              f"{'  ← H₀ rifiutata' if p_val < ALPHA and lag * 7 < 30 else ''}")
    granger_results[fuel_name] = pd.DataFrame(rows)
    granger_results[fuel_name].to_csv(f"data/granger_{fuel_name.lower()}.csv", index=False)


def _granger_panel(ax, df_gc, fuel_name):
    lags_d = df_gc["lag_days"].values
    pvals  = df_gc["p_value"].values
    colors = [_bar_color(p) for p in pvals]
    ax.set_axisbelow(True)
    bars = ax.bar(lags_d, pvals, color=colors, edgecolor="black",
                  linewidth=0.6, alpha=0.88, width=5.5, zorder=3)
    ax.axhline(ALPHA, color="#2c3e50", lw=1.4, ls="--", zorder=4)
    ax.text(lags_d[-1] + 1, ALPHA + 0.005, f"α={ALPHA}",
            ha="right", va="bottom", fontsize=8, color="#2c3e50", style="italic")
    ax.axvline(30, color="#e67e22", lw=1.6, ls="--", zorder=4)
    ax.axvspan(0, 30, alpha=0.06, color="#e74c3c", zorder=1)
    ax.text(15, 0.90 * ax.get_ylim()[1],
            "< 30gg\n(trasmissione rapida)", ha="center", va="top",
            fontsize=7, color="#c0392b", style="italic")
    y_top = max(pvals) * 1.3
    for bar, p in zip(bars, pvals):
        bx = bar.get_x() + bar.get_width() / 2
        ax.text(bx, bar.get_height() + y_top * 0.02, f"{p:.3f}",
                ha="center", va="bottom", fontsize=6.8,
                color=_bar_color(p) if p < ALPHA else "#777777")
        if _stars(p) not in ("(n.s.)", ""):
            ax.text(bx, bar.get_height() + y_top * 0.09, _stars(p),
                    ha="center", va="bottom", fontsize=8, color=_bar_color(p), fontweight="bold")
    n_sig = sum(p < ALPHA and d <= 30 for p, d in zip(pvals, lags_d))
    ax.set_title(f"Granger: Brent → {fuel_name}\nLag sign. <30gg: {n_sig}",
                 fontsize=11, fontweight="bold",
                 color="#b03030" if n_sig > 0 else "black")
    ax.set_xlabel("Lag (settimane → giorni)", fontsize=10)
    ax.set_ylabel("p-value (F-test)", fontsize=10)
    ax.set_xticks(lags_d)
    ax.set_xticklabels([f"{int(d)}d\n(w{int(d)//7})" for d in lags_d], fontsize=8)
    ax.set_ylim(0, max(y_top, 0.18))
    ax.legend(handles=[
        mpatches.Patch(color="#8b1a1a", label="p<0.001 (***)"),
        mpatches.Patch(color="#c0392b", label="p<0.01  (**)"),
        mpatches.Patch(color="#e74c3c", label="p<0.05  (*)"),
        mpatches.Patch(color="#95a5a6", label="p≥0.05  (n.s.)"),
    ], fontsize=7.5, loc="upper right", title="Sign.")


if len(granger_results) >= 1:
    n_panels = len(granger_results)
    fig, axes = plt.subplots(1, n_panels, figsize=(7 * n_panels, 5), sharey=False)
    if n_panels == 1:
        axes = [axes]
    for ax, (fuel_name, df_gc) in zip(axes, granger_results.items()):
        _granger_panel(ax, df_gc, fuel_name)
    fig.suptitle(
        "Granger Causality: Brent → Prezzi Carburanti Italia\n"
        "Evidenza di trasmissione rapida (ausiliaria, non conclusiva su speculazione)",
        fontsize=11, fontweight="bold", y=1.02,
    )
    fig.tight_layout(pad=1.5)
    fig.savefig("plots/03_granger_combined.png", dpi=DPI, bbox_inches="tight")
    plt.close(fig)
    print("\n  Salvato: plots/03_granger_combined.png")


# ─────────────────────────────────────────────────────────────────────────────
# §2. ROCKETS & FEATHERS (GLSAR AR(1) + HAC Newey-West)
#     H0: β_up = β_down (simmetria rialzo/ribasso)
#     NOTA: R&F index > 1 suggerisce asimmetria, non speculazione diretta.
#     Metodo: GLSAR iterativo (Cochrane-Orcutt) + SE HAC (Newey-West, 4 lag)
#     Motivazione: DW ≈ 0.003-0.04 → OLS SE gravemente distorte
#     Rif: Greene (2012) cap.20; Newey & West (1987)
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 65)
print("§2. ROCKETS & FEATHERS (GLSAR AR(1) + HAC, test ausiliario)")
print("=" * 65)

merged_d["d_brent"]   = merged_d["brent_7d_eur"].pct_change() * 100
merged_d["d_benzina"] = merged_d["benzina_4w"].pct_change() * 100
merged_d["d_diesel"]  = merged_d["diesel_4w"].pct_change() * 100
merged_rf = merged_d.dropna(subset=["d_brent", "d_benzina", "d_diesel"])

FUELS_RF = {
    "Benzina": ("d_benzina", "benzina_4w", "#d6604d"),
    "Diesel":  ("d_diesel",  "diesel_4w",  "#31a354"),
}

rf_results = {}

for fuel_name, (dcol, col4w, fc) in FUELS_RF.items():
    y         = merged_rf[dcol].dropna().values
    brent_pos = np.maximum(merged_rf["d_brent"].values[:len(y)], 0)
    brent_neg = np.minimum(merged_rf["d_brent"].values[:len(y)], 0)
    X         = np.column_stack([np.ones(len(y)), brent_pos, brent_neg])
    n, k      = len(y), 3

    try:
        glsar_model = GLSAR(y, X, rho=1)
        glsar_res   = glsar_model.iterative_fit(maxiter=10)
        b_up, b_down = glsar_res.params[1], glsar_res.params[2]
        rho_ar       = float(glsar_model.rho)   # ← dal modello, non dai risultati
        cov_nw       = cov_hac(glsar_res, nlags=4)
        method       = "GLSAR-AR(1)+HAC"
    except Exception as e:
        print(f"  {fuel_name}: GLSAR fallito ({e}), uso OLS+HAC")
        ols_r        = OLS(y, X).fit()
        b_up, b_down = ols_r.params[1], ols_r.params[2]
        rho_ar       = np.nan
        cov_nw       = cov_hac(ols_r, nlags=4)
        method       = "OLS+HAC"

    se_up   = float(np.sqrt(cov_nw[1, 1]))
    se_down = float(np.sqrt(cov_nw[2, 2]))
    se_diff = np.sqrt(se_up**2 + se_down**2)
    t_stat  = (b_up - b_down) / se_diff
    p_asym  = float(stats.t.sf(abs(t_stat), df=n - k) * 2)
    rf_idx  = abs(b_up) / abs(b_down) if b_down != 0 else np.inf

    rf_results[fuel_name] = {
        "b_up": b_up, "b_down": b_down, "se_up": se_up, "se_down": se_down,
        "rf_index": rf_idx, "t_stat": t_stat, "p_asym": p_asym,
        "rho_ar": rho_ar, "method": method, "col": fc, "col4w": col4w,
    }
    print(f"\n  {fuel_name} [{method}]:")
    print(f"    β_up  = {b_up:.4f} (SE={se_up:.4f})")
    print(f"    β_down= {b_down:.4f} (SE={se_down:.4f})")
    print(f"    ρ AR(1)={rho_ar:.3f}  |  R&F index={rf_idx:.3f}  |  "
          f"p={p_asym:.4f} {_stars(p_asym)}")

# Plot R&F: scatter β_up vs β_down, affiancati
fig_rf, axes_rf = plt.subplots(1, 2, figsize=(14, 6), sharey=False)
for ax, (fuel_name, res) in zip(axes_rf, rf_results.items()):
    brent_x = merged_rf["d_brent"].values
    fuel_y  = merged_rf[FUELS_RF[fuel_name][0]].values
    valid   = ~(np.isnan(brent_x) | np.isnan(fuel_y))
    colors  = ["#e74c3c" if b > 0 else "#3498db" for b in brent_x[valid]]
    ax.scatter(brent_x[valid], fuel_y[valid], c=colors, alpha=0.35, s=18)
    x_up   = np.linspace(0, np.nanmax(brent_x), 100)
    x_down = np.linspace(np.nanmin(brent_x), 0, 100)
    ax.plot(x_up,   res["b_up"]   * x_up,   color="#e74c3c", lw=2.5,
            label=f"β_up={res['b_up']:.4f}")
    ax.plot(x_down, res["b_down"] * x_down, color="#3498db", lw=2.5, ls="--",
            label=f"β_down={res['b_down']:.4f}")
    ax.axhline(0, color="black", lw=0.5)
    ax.axvline(0, color="black", lw=0.5)
    p_s = f"{res['p_asym']:.4f}" if not np.isnan(res["p_asym"]) else "N/A"
    ax.set_title(f"R&F — {fuel_name} [{res['method']}]\n"
                 f"R&F index={res['rf_index']:.3f}  p={p_s} {_stars(res['p_asym'])}",
                 fontsize=11, fontweight="bold")
    ax.set_xlabel("ΔBrent (%)", fontsize=10)
    ax.set_ylabel(f"Δ{fuel_name} (%)", fontsize=10)
    ax.legend(fontsize=10)

fig_rf.suptitle("Rockets & Feathers: asimmetria trasmissione Brent→pompa\n"
                "(evidenza ausiliaria — non test diretto di H0 sul margine)",
                fontsize=11, fontweight="bold", y=1.02)
fig_rf.tight_layout(pad=1.5)
fig_rf.savefig("plots/04_rf_combined.png", dpi=DPI, bbox_inches="tight")
plt.close(fig_rf)
print("\n  Salvato: plots/04_rf_combined.png")

pd.DataFrame([{
    "Carburante": f, "Metodo": r["method"],
    "β_up": round(r["b_up"], 4), "SE_up_HAC": round(r["se_up"], 4),
    "β_down": round(r["b_down"], 4), "SE_down_HAC": round(r["se_down"], 4),
    "rho_AR1": round(r["rho_ar"], 4) if not np.isnan(r["rho_ar"]) else "N/A",
    "R&F_index": round(r["rf_index"], 3),
    "t_stat": round(r["t_stat"], 3),
    "p_asym": round(r["p_asym"], 4),
} for f, r in rf_results.items()]).to_csv("data/rockets_feathers_results.csv", index=False)


# ─────────────────────────────────────────────────────────────────────────────
# §3. KS TEST, ANOVA A UN FATTORE, CHOW TEST
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 65)
print("§3. KS test, ANOVA, Chow test")
print("=" * 65)


def chow_test(y: np.ndarray, breakpoint: int):
    """
    Chow test: F = [(RSS_r - RSS_u) / k] / [RSS_u / (n - 2k)]
    H0: nessuna rottura strutturale
    """
    n, k = len(y), 2
    x    = np.arange(n)
    X_f  = np.column_stack([np.ones(n), x])
    b_r, _, _, _ = np.linalg.lstsq(X_f, y, rcond=None)
    rss_r = np.sum((y - X_f @ b_r) ** 2)

    def _ols_rss(xv, yv):
        if len(xv) < k + 1:
            return 0.0
        Xv = np.column_stack([np.ones(len(xv)), xv])
        b, _, _, _ = np.linalg.lstsq(Xv, yv, rcond=None)
        return np.sum((yv - Xv @ b) ** 2)

    rss_u = _ols_rss(x[:breakpoint], y[:breakpoint]) + _ols_rss(x[breakpoint:], y[breakpoint:])
    if rss_u < 1e-12:
        return np.nan, np.nan
    f_stat = ((rss_r - rss_u) / k) / (rss_u / (n - 2 * k))
    return f_stat, 1 - stats.f.cdf(f_stat, dfn=k, dfd=n - 2 * k)


ks_rows, anova_rows, chow_rows = [], [], []

for event_name, cfg in EVENTS.items():
    shock = cfg["shock"]
    for fuel_name, fuel_col in FUELS.items():
        if fuel_col not in merged.columns:
            continue
        pre  = merged.loc[cfg["pre_start"]:shock, fuel_col].dropna()
        post = merged.loc[shock:cfg["post_end"],  fuel_col].dropna()
        if len(pre) < 4 or len(post) < 4:
            continue

        # KS
        ks_s, ks_p = stats.ks_2samp(pre.values, post.values)
        ks_rows.append({
            "Evento": event_name, "Carburante": fuel_name,
            "n_pre": len(pre), "n_post": len(post),
            "KS_stat": round(ks_s, 4), "p_value": round(ks_p, 6),
            "H0": "RIFIUTATA" if ks_p < ALPHA else "non rifiutata",
        })

        # ANOVA (3 periodi)
        pA = merged.loc[cfg["pre_start"]:shock, fuel_col].dropna()
        pB = merged.loc[shock:shock + pd.Timedelta(weeks=6), fuel_col].dropna()
        pC = merged.loc[shock + pd.Timedelta(weeks=6):cfg["post_end"], fuel_col].dropna()
        if len(pA) >= 3 and len(pB) >= 3 and len(pC) >= 3:
            f_st, an_p = stats.f_oneway(pA, pB, pC)
            anova_rows.append({
                "Evento": event_name, "Carburante": fuel_name,
                "F_stat": round(f_st, 4), "p_value": round(an_p, 6),
                "mean_pre": round(pA.mean(), 4), "mean_shock6w": round(pB.mean(), 4),
                "mean_post": round(pC.mean(), 4),
                "H0": "RIFIUTATA" if an_p < ALPHA else "non rifiutata",
            })

        # Chow (breakpoint al momento dello shock)
        series_full = merged.loc[cfg["pre_start"]:cfg["post_end"], fuel_col].dropna()
        bp_idx = int(series_full.index.searchsorted(shock))
        if 2 <= bp_idx <= len(series_full) - 3:
            ch_f, ch_p = chow_test(series_full.values, bp_idx)
            chow_rows.append({
                "Evento": event_name, "Carburante": fuel_name,
                "F_stat": round(float(ch_f), 4) if not np.isnan(ch_f) else "N/A",
                "p_value": round(float(ch_p), 6) if not np.isnan(ch_p) else "N/A",
                "H0": ("RIFIUTATA" if (not np.isnan(ch_p) and ch_p < ALPHA)
                       else "non rifiutata"),
            })

pd.DataFrame(ks_rows).to_csv("data/ks_results.csv", index=False)
pd.DataFrame(anova_rows).to_csv("data/anova_results.csv", index=False)
pd.DataFrame(chow_rows).to_csv("data/chow_results.csv", index=False)
print(f"  KS: {len(ks_rows)} test | ANOVA: {len(anova_rows)} | Chow: {len(chow_rows)}")
for r in ks_rows:
    print(f"  KS {r['Evento'].split('(')[0].strip()} | {r['Carburante']}: "
          f"p={r['p_value']:.4f} → H0 {r['H0']}")


# ─────────────────────────────────────────────────────────────────────────────
# §4. CCF e Rolling Correlation
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 65)
print("§4. Cross-Correlation (CCF) e Rolling Correlation")
print("=" * 65)

brent_clean = merged["brent_7d_eur"].dropna()
ccf_results = {}
rolling_corr = {}

for fuel_name, fuel_col in FUELS.items():
    if fuel_col not in merged.columns:
        continue
    fuel_clean = merged[fuel_col].dropna()
    common_idx = brent_clean.index.intersection(fuel_clean.index)
    if len(common_idx) < MAX_LAG_CCF + 2:
        continue

    b_vals = brent_clean[common_idx].values
    f_vals = fuel_clean[common_idx].values
    b_std  = (b_vals - b_vals.mean()) / (b_vals.std() or 1)
    f_std  = (f_vals - f_vals.mean()) / (f_vals.std() or 1)

    ccf_vals = []
    for lag in range(0, MAX_LAG_CCF + 1):
        corr = np.corrcoef(b_std[:len(b_std) - lag], f_std[lag:])[0, 1]
        ccf_vals.append(corr)
    ccf_results[fuel_name] = ccf_vals

    # Rolling correlation
    combined = pd.DataFrame({"brent": b_vals, "fuel": f_vals}, index=common_idx)
    rolling_corr[fuel_name] = combined["brent"].rolling(ROLL_WIN).corr(combined["fuel"])
    print(f"  CCF {fuel_name}: lag ottimale = "
          f"{np.argmax(ccf_vals)} settimane ({np.argmax(ccf_vals)*7}gg)")


# ─────────────────────────────────────────────────────────────────────────────
# §5. Bootstrap 95% CI sul lag D
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 65)
print("§5. Bootstrap 95% CI sul lag D (changepoint vs shock)")
print("=" * 65)

bootstrap_results = []
try:
    table1 = pd.read_csv("data/table1_changepoints.csv")
    for _, row in table1.iterrows():
        event_name = row["Evento"]
        fuel_name  = row["Serie"]
        if event_name not in EVENTS or fuel_name not in FUELS:
            continue
        fuel_col = FUELS.get(fuel_name)
        if fuel_col not in merged.columns:
            continue

        cfg   = EVENTS[event_name]
        shock = cfg["shock"]
        series = merged.loc[cfg["pre_start"]:cfg["post_end"], fuel_col].dropna()
        if len(series) < 10:
            continue
        shock_idx = int(series.index.searchsorted(shock))

        rng    = np.random.default_rng(42)
        lags_b = []
        for _ in range(2000):
            idx_boot = rng.integers(0, len(series), size=len(series))
            s_boot   = series.values[np.sort(idx_boot)]
            # stima changepoint come indice con massima differenza medie
            best_idx, best_score = shock_idx, 0.0
            for i in range(2, len(s_boot) - 2):
                score = abs(s_boot[:i].mean() - s_boot[i:].mean())
                if score > best_score:
                    best_score = score
                    best_idx   = i
            cp_date_b = series.index[min(best_idx, len(series) - 1)]
            lags_b.append((cp_date_b - shock).days)

        lags_arr = np.array(lags_b)
        bootstrap_results.append({
            "Evento":    event_name,
            "Carburante": fuel_name,
            "Lag_mean":  round(float(np.mean(lags_arr)), 1),
            "CI_95_low": round(float(np.percentile(lags_arr, 2.5)), 1),
            "CI_95_high": round(float(np.percentile(lags_arr, 97.5)), 1),
            "H0_30gg":   "RIFIUTATA" if np.percentile(lags_arr, 97.5) < 30 else "non rifiutata",
        })
        print(f"  {event_name[:20]} | {fuel_name}: "
              f"CI 95% lag = [{bootstrap_results[-1]['CI_95_low']:.0f}, "
              f"{bootstrap_results[-1]['CI_95_high']:.0f}] gg → "
              f"H0 {bootstrap_results[-1]['H0_30gg']}")
except FileNotFoundError:
    print("  data/table1_changepoints.csv non trovato — eseguire prima script 02")

if bootstrap_results:
    pd.DataFrame(bootstrap_results).to_csv("data/bootstrap_ci.csv", index=False)


# ─────────────────────────────────────────────────────────────────────────────
# §6. Selezione tipo di regressione (BP, Ljung-Box, DW, AIC/BIC, SE comparison)
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 65)
print("§6. Selezione tipo di regressione per ogni serie × evento")
print("=" * 65)

from statsmodels.stats.diagnostic import het_breuschpagan
import statsmodels.api as sm

selection_rows = []

for event_name, cfg in EVENTS.items():
    shock = cfg["shock"]
    for fuel_name, fuel_col in FUELS.items():
        if fuel_col not in merged.columns:
            continue
        series = merged.loc[cfg["pre_start"]:cfg["post_end"], fuel_col].dropna()
        if len(series) < 12:
            continue

        x = np.arange(len(series), dtype=float)
        y = series.values
        X = sm.add_constant(x)
        ols_r = sm.OLS(y, X).fit()
        resid = ols_r.resid

        # Breusch-Pagan
        try:
            _, bp_p, _, _ = het_breuschpagan(resid, X)
        except Exception:
            bp_p = np.nan

        # Ljung-Box (lag 4)
        try:
            lb_p = acorr_ljungbox(resid, lags=[4], return_df=True)["lb_pvalue"].values[0]
        except Exception:
            lb_p = np.nan

        # DW
        dw = durbin_watson(resid)

        # AR(1) stima ρ
        try:
            rho_hat = np.corrcoef(resid[:-1], resid[1:])[0, 1]
        except Exception:
            rho_hat = np.nan

        # SE OLS vs HAC
        se_ols = ols_r.bse[1]
        try:
            cov_nw = cov_hac(ols_r, nlags=4)
            se_hac = float(np.sqrt(cov_nw[1, 1]))
            se_inf = 100 * (se_hac - se_ols) / se_ols if se_ols > 0 else np.nan
        except Exception:
            se_hac, se_inf = np.nan, np.nan

        # Raccomandazione
        if dw < 0.5 or (not np.isnan(rho_hat) and abs(rho_hat) > 0.85):
            rec = "Bayesian StudentT + AR(1)"
        elif (not np.isnan(bp_p) and bp_p < ALPHA) or (not np.isnan(lb_p) and lb_p < ALPHA):
            rec = "GLSAR AR(1) + HAC"
        elif not np.isnan(bp_p) and bp_p < ALPHA:
            rec = "OLS + HAC Newey-West"
        else:
            rec = "OLS standard"

        selection_rows.append({
            "Evento":         event_name,
            "Serie":          fuel_name,
            "BP_p":           round(float(bp_p), 4) if not np.isnan(bp_p) else "N/A",
            "LjungBox_p_min": round(float(lb_p), 4) if not np.isnan(lb_p) else "N/A",
            "DW":             round(dw, 3),
            "rho_AR1":        round(float(rho_hat), 3) if not np.isnan(rho_hat) else "N/A",
            "SE_inflation_%": round(se_inf, 1) if not np.isnan(se_inf) else "N/A",
            "Raccomandazione": rec,
        })
        print(f"  {event_name[:22]} | {fuel_name}: BP_p={bp_p:.3f} "
              f"DW={dw:.2f} ρ={rho_hat:.2f} → {rec}")

pd.DataFrame(selection_rows).to_csv("data/regression_selection.csv", index=False)

# ─────────────────────────────────────────────────────────────────────────────
# §7. WELCH T-TEST — margine lordo pre vs post shock
#     H0: μ_margine_pre = μ_margine_post  (nessun aumento del crack spread)
#     Test di Welch (varianze non assunte uguali) sul margine lordo calcolato come:
#       margine = prezzo_pompa_senza_tasse - brent_eur_per_litro
#     Questo test è AUSILIARIO e complementare allo z-score BH di script 02.
#     Rif: Welch (1947); consigliato su Granger quando n è piccolo per evento.
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 65)
print("§7. WELCH T-TEST — margine lordo pre vs post shock")
print("    H0: media margine pre == media margine post")
print("=" * 65)

BRENT_L_FACTOR = 159.0  # litri per barile (barrel → litro)

# ── Normalizzazione unità pompa ───────────────────────────────────────────────
# benzina_4w / diesel_4w sono in milli-EUR/L (median > 10) oppure EUR/L.
# STESSA LOGICA di 02_core_analysis.py §C per coerenza.
# Senza questa correzione il margine sarebbe in unità miste → Δ ~300-400 invece di ~0.3-0.4.
for _raw_col, _eur_l_col in [("benzina_4w", "benzina_eur_l_s03"),
                               ("diesel_4w",  "diesel_eur_l_s03")]:
    if _raw_col in merged.columns:
        _med = merged[_raw_col].dropna().median()
        _uf  = 1000.0 if _med > 10 else 1.0
        merged[_eur_l_col] = merged[_raw_col] / _uf

# Calcola i margini nel dataset principale (EUR/L normalizzati)
for fuel_name, eur_l_col in [("Benzina", "benzina_eur_l_s03"),
                               ("Diesel",  "diesel_eur_l_s03")]:
    margin_col = f"margine_{fuel_name.lower()}"
    if eur_l_col in merged.columns and "brent_eur" in merged.columns:
        merged[margin_col] = merged[eur_l_col] - merged["brent_eur"] / BRENT_L_FACTOR

FUELS_MARGIN = {
    "Benzina": "margine_benzina",
    "Diesel":  "margine_diesel",
}

ttest_rows = []

for event_name, cfg in EVENTS.items():
    shock = cfg["shock"]
    for fuel_name, margin_col in FUELS_MARGIN.items():
        if margin_col not in merged.columns:
            print(f"  Colonna {margin_col} non trovata — skip")
            continue

        pre  = merged.loc[cfg["pre_start"]:shock, margin_col].dropna()
        post = merged.loc[shock:cfg["post_end"],  margin_col].dropna()

        if len(pre) < 4 or len(post) < 4:
            print(f"  {event_name} | {fuel_name}: campioni troppo piccoli — skip")
            continue

        t_stat, p_val = stats.ttest_ind(pre.values, post.values, equal_var=False)
        cohens_d = (post.mean() - pre.mean()) / np.sqrt(
            (pre.std() ** 2 + post.std() ** 2) / 2
        )
        # IC 95% sulla differenza delle medie (approssimazione Welch-Satterthwaite)
        se_diff = np.sqrt(pre.var() / len(pre) + post.var() / len(post))
        ci_lo = (post.mean() - pre.mean()) - 1.96 * se_diff
        ci_hi = (post.mean() - pre.mean()) + 1.96 * se_diff

        row = {
            "Evento":         event_name,
            "Carburante":     fuel_name,
            "n_pre":          len(pre),
            "n_post":         len(post),
            "mean_pre":       round(pre.mean(),  4),
            "mean_post":      round(post.mean(), 4),
            "delta_mean":     round(post.mean() - pre.mean(), 4),
            "CI_95_low":      round(ci_lo, 4),
            "CI_95_high":     round(ci_hi, 4),
            "t_stat":         round(t_stat, 4),
            "p_value":        round(p_val,  6),
            "cohens_d":       round(cohens_d, 3),
            "H0": "RIFIUTATA" if p_val < ALPHA else "non rifiutata",
        }
        ttest_rows.append(row)
        print(f"  {event_name.split('(')[0].strip():<25} | {fuel_name:<8}: "
              f"Δ={row['delta_mean']:+.4f} EUR/l  "
              f"t={t_stat:.3f}  p={p_val:.4f} {_stars(p_val)}  "
              f"d={cohens_d:.2f}  → H0 {row['H0']}")

if ttest_rows:
    pd.DataFrame(ttest_rows).to_csv("data/ttest_margin.csv", index=False)
    print(f"\n  Salvato: data/ttest_margin.csv ({len(ttest_rows)} test)")


# ─────────────────────────────────────────────────────────────────────────────
# §8. DIFFERENCE-IN-DIFFERENCES (DiD) — Italia vs Germania e Francia
#     Modello OLS:
#       M_{c,t} = α + β1·Italy_c + β2·Post_t + δ·(Italy_c × Post_t) + ε
#     δ = estimatore DiD: variazione del margine IT *relativa* al paese controllo
#         dopo lo shock → isola effetto specifico al mercato italiano
#
#     Due paesi di controllo:
#       • Germania (DE) — principale, grande economia manifatturiera
#       • Francia  (FR) — secondo controllo, economia paragonabile.
#           NOTA: nel 2022 la Francia ha introdotto una "ristourne carburant"
#           (~0.15–0.18 €/l, set–dic 2022) che può comprimere artificialmente
#           il margine osservato FR. I risultati DiD IT vs FR per Ucraina 2022
#           vanno interpretati con cautela (politica fiscale confondente).
#           Rif: Décret n°2022-1153 (23 ago 2022); DGEC rapporto 2022.
#
#     Strategia dati: riutilizza il file EU Oil Bulletin già scaricato in §1
#     (data/eu_oil_bulletin_history.xlsx), estrae colonne DE e FR.
#     Se il file non esiste → skip con nota.
#     Rif: Angrist & Pischke (2009), cap. 5; Card & Krueger (1994).
# ─────────────────────────────────────────────────────────────────────────────

# ── Configurazione paesi di controllo ───────────────────────────────────────
# Aggiungere qui nuovi paesi se disponibili nell'EU Oil Bulletin.
# "prefissi_col": liste di prefissi colonna (uppercase) per identificare le
#   colonne benzina/diesel del paese nel foglio EU Oil Bulletin.
# "nota": stringa di avvertenza metodologica (None = nessuna).
DID_CONTROLS = {
    "Germania": {
        "prefissi_col": ["DE", "GERMANY", "DEUTSCH"],
        "nota": None,
    },
    "Svezia": {
        "prefissi_col": ["SE", "SWEDEN", "SVERIGE"],
        "nota": None,
    },
}

print("\n" + "=" * 65)
print("§8. DIFFERENCE-IN-DIFFERENCES — Italia vs Germania e Francia")
print("    δ = variazione margine IT relativa al paese controllo post-shock")
print("=" * 65)

did_rows = []
CONTROL_PUMPS = {}   # { nome_paese: DataFrame con colonne margine_benzina, margine_diesel }

def _safe_sheet_names_did(path):
    wb = _opxl.load_workbook(path, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names

def _safe_read_sheet_did(path, sheet_name):
    wb = _opxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if not rows:
        return pd.DataFrame()
    header_idx = next((i for i, r in enumerate(rows) if any(v is not None for v in r)), 0)
    headers = [str(h).strip() if h is not None else f"_col{i}"
               for i, h in enumerate(rows[header_idx])]
    df = pd.DataFrame(rows[header_idx + 1:], columns=headers)
    idx_col = headers[0]
    df[idx_col] = pd.to_datetime(df[idx_col], errors="coerce")
    df = df.set_index(idx_col)
    return df[df.index.notna()].sort_index()

def _find_notax(names):
    for s in names:
        su = s.upper()
        if any(k in su for k in ["WO TAX", "WITHOUT", "NO TAX", "NOTAX", "WO TAXES"]):
            return s
    return names[1] if len(names) > 1 else names[0]

EU_HIST_FILE = "data/eu_oil_bulletin_history.xlsx"

try:
    sheet_names_eu = _safe_sheet_names_did(EU_HIST_FILE)
    notax_sheet    = _find_notax(sheet_names_eu)
    df_eu          = _safe_read_sheet_did(EU_HIST_FILE, notax_sheet)
    df_eu          = df_eu.apply(pd.to_numeric, errors="coerce")

    for paese, cfg_paese in DID_CONTROLS.items():
        prefissi = cfg_paese["prefissi_col"]
        # Trova colonne che matchano uno qualsiasi dei prefissi del paese
        all_cols = [c for c in df_eu.columns if any(
            str(c).upper().startswith(p) or p in str(c).upper()
            for p in prefissi
        )]
        b_cols = [c for c in all_cols if any(k in str(c).lower()
                  for k in ["95", "benz", "petrol", "gasol", "euro", "unleaded", "super"])]
        d_cols = [c for c in all_cols if any(k in str(c).lower()
                  for k in ["diesel", "gas_oil", "gasoil"])]
        # Fallback: prendi le prime due colonne disponibili
        if not b_cols and len(all_cols) >= 1:
            b_cols = [all_cols[0]]
        if not d_cols and len(all_cols) >= 2:
            d_cols = [all_cols[1]]

        if b_cols and d_cols:
            pump = pd.concat([
                df_eu[b_cols[0]].rename("benzina_eur_l"),
                df_eu[d_cols[0]].rename("diesel_eur_l"),
            ], axis=1)
            pump = pump[pump.index >= "2019-01-01"].dropna(how="all")
            pump = pump.resample("W-MON").mean()
            for _col in ["benzina_eur_l", "diesel_eur_l"]:
                if _col in pump.columns:
                    _med = pump[_col].dropna().median()
                    pump[_col] = pump[_col] / (1000.0 if _med > 10 else 1.0)
            brent_aligned = merged["brent_eur"].reindex(pump.index).ffill(limit=4)
            pump["margine_benzina"] = pump["benzina_eur_l"] - brent_aligned / BRENT_L_FACTOR
            pump["margine_diesel"]  = pump["diesel_eur_l"]  - brent_aligned / BRENT_L_FACTOR
            CONTROL_PUMPS[paese] = pump
            print(f"  {paese}: {len(pump.dropna())} settimane caricate "
                  f"({b_cols[0]}, {d_cols[0]})")
        else:
            print(f"  {paese}: colonne non trovate nel foglio '{notax_sheet}' — skip")

    if not CONTROL_PUMPS:
        print(f"  Nessun paese di controllo caricato — DiD skip")

except FileNotFoundError:
    print(f"  {EU_HIST_FILE} non trovato — eseguire prima 01_data_pipeline.py")
except Exception as e:
    print(f"  Errore caricamento dati controllo: {e} — DiD skip")

if CONTROL_PUMPS:
    import statsmodels.api as sm

    FUELS_DID = {
        "Benzina": ("margine_benzina", "#d6604d"),
        "Diesel":  ("margine_diesel",  "#31a354"),
    }

    for paese, ctrl_pump in CONTROL_PUMPS.items():
        nota_paese = DID_CONTROLS[paese]["nota"]
        if nota_paese:
            print(f"\n  {nota_paese}")

        for event_name, cfg in EVENTS.items():
            shock      = cfg["shock"]
            pre_start  = cfg["pre_start"]
            post_end   = cfg["post_end"]

            for fuel_name, (margin_col, fc) in FUELS_DID.items():
                if margin_col not in merged.columns or margin_col not in ctrl_pump.columns:
                    continue

                # Costruisci panel IT + paese_controllo nella finestra evento
                it_pre  = merged.loc[pre_start:shock,     margin_col].dropna()
                it_post = merged.loc[shock:post_end,       margin_col].dropna()
                ct_pre  = ctrl_pump.loc[pre_start:shock,  margin_col].dropna()
                ct_post = ctrl_pump.loc[shock:post_end,   margin_col].dropna()

                if any(len(s) < 3 for s in [it_pre, it_post, ct_pre, ct_post]):
                    continue

                # ── Parallel trends test (pre-shock only) ────────────────────
                pt_pass   = None
                pt_pvalue = np.nan
                try:
                    all_pre_dates = it_pre.index.union(ct_pre.index).sort_values()
                    t_vals_it = np.array([(d - all_pre_dates[0]).days for d in it_pre.index], dtype=float)
                    t_vals_ct = np.array([(d - all_pre_dates[0]).days for d in ct_pre.index], dtype=float)
                    rows_pt = (
                        [(1, t, v) for t, v in zip(t_vals_it, it_pre.values)] +
                        [(0, t, v) for t, v in zip(t_vals_ct, ct_pre.values)]
                    )
                    df_pt = pd.DataFrame(rows_pt, columns=["Italy", "t", "Margin"])
                    df_pt["Italy_x_t"] = df_pt["Italy"] * df_pt["t"]
                    X_pt    = sm.add_constant(df_pt[["Italy", "t", "Italy_x_t"]].values)
                    ols_pt  = sm.OLS(df_pt["Margin"].values, X_pt).fit(cov_type="HC3")
                    pt_pvalue = float(ols_pt.pvalues[3])
                    pt_pass   = pt_pvalue >= ALPHA
                    pt_label  = f"PTA p={pt_pvalue:.3f} {'✓ non rigettata' if pt_pass else '✗ VIOLATA'}"
                    print(f"    [{paese}] Parallel Trends: {pt_label}")
                except Exception as e:
                    print(f"    [{paese}] Parallel Trends: errore ({e})")

                rows_panel = (
                    [(1, 0, v) for v in it_pre]  +   # IT pre
                    [(1, 1, v) for v in it_post] +    # IT post
                    [(0, 0, v) for v in ct_pre]  +    # paese_ctrl pre
                    [(0, 1, v) for v in ct_post]       # paese_ctrl post
                )
                df_panel = pd.DataFrame(rows_panel, columns=["Italy", "Post", "Margin"])
                df_panel["Italy_x_Post"] = df_panel["Italy"] * df_panel["Post"]

                X_did   = sm.add_constant(df_panel[["Italy", "Post", "Italy_x_Post"]].values)
                ols_did = sm.OLS(df_panel["Margin"].values, X_did).fit(cov_type="HC3")
                delta     = ols_did.params[3]
                se_delta  = ols_did.bse[3]
                t_did     = ols_did.tvalues[3]
                p_did     = ols_did.pvalues[3]
                ci_lo_did = delta - 1.96 * se_delta
                ci_hi_did = delta + 1.96 * se_delta

                row_did = {
                    "Evento":             event_name,
                    "Paese_controllo":    paese,
                    "Carburante":         fuel_name,
                    "n_IT_pre":           len(it_pre),
                    "n_IT_post":          len(it_post),
                    "n_CT_pre":           len(ct_pre),
                    "n_CT_post":          len(ct_post),
                    "PTA_pvalue":         round(pt_pvalue, 4) if not np.isnan(pt_pvalue) else "N/A",
                    "PTA_non_rigettata":  pt_pass,
                    "delta_DiD":          round(delta,    4),
                    "SE_HC3":             round(se_delta, 4),
                    "CI_95_low":          round(ci_lo_did, 4),
                    "CI_95_high":         round(ci_hi_did, 4),
                    "t_stat":             round(t_did,  3),
                    "p_value":            round(p_did,  6),
                    "R2":                 round(ols_did.rsquared, 3),
                    "H0":                 "RIFIUTATA" if p_did < ALPHA else "non rifiutata",
                }
                did_rows.append(row_did)
                pta_warn = " ⚠ PTA violata" if pt_pass is False else ""
                print(f"  [{paese}] {event_name.split('(')[0].strip():<22} | {fuel_name:<8}: "
                      f"δ={delta:+.4f} EUR/l  SE={se_delta:.4f}  "
                      f"p={p_did:.4f} {_stars(p_did)}  → H0 {row_did['H0']}{pta_warn}")

    if did_rows:
        pd.DataFrame(did_rows).to_csv("data/did_results.csv", index=False)
        print(f"\n  Salvato: data/did_results.csv ({len(did_rows)} stime DiD)")
else:
    print("  DiD non eseguito (nessun dato paese di controllo disponibile).")


# ─────────────────────────────────────────────────────────────────────────────
# FIGURA §7–§8: t-test e DiD (plots/09_ttest_did.png)
# ─────────────────────────────────────────────────────────────────────────────
print("\nGenerazione figura §7-§8 (t-test + DiD)...")

_n_cols = 2
_n_rows = 1 + (1 if did_rows else 0)
fig9, axes9 = plt.subplots(_n_rows, _n_cols,
                            figsize=(14, 5 * _n_rows),
                            squeeze=False)
fig9.suptitle("Test aggiuntivi — Anomalia margine carburanti Italia (evidenza ausiliaria)",
              fontsize=13, fontweight="bold")

# Pannello A: Welch t-test — Δmargine medio per evento × carburante
ax_t = axes9[0, 0]
if ttest_rows:
    df_tt = pd.DataFrame(ttest_rows)
    labels_tt  = [f"{r['Evento'].split('(')[0].strip()}\n{r['Carburante']}"
                  for _, r in df_tt.iterrows()]
    deltas_tt  = df_tt["delta_mean"].values
    ci_lo_arr  = df_tt["delta_mean"].values - df_tt["CI_95_low"].values
    ci_hi_arr  = df_tt["CI_95_high"].values - df_tt["delta_mean"].values
    colors_tt  = ["#e74c3c" if p < ALPHA else "#95a5a6"
                  for p in df_tt["p_value"].values]
    bars_tt = ax_t.barh(range(len(labels_tt)), deltas_tt,
                         color=colors_tt, edgecolor="black", lw=0.5, alpha=0.85)
    for i, (lo, hi) in enumerate(zip(ci_lo_arr, ci_hi_arr)):
        ax_t.errorbar(deltas_tt[i], i,
                      xerr=[[lo], [hi]],
                      fmt="none", color="black", capsize=5, lw=1.5)
    ax_t.axvline(0, color="black", lw=1.0, ls="--")
    ax_t.set_yticks(range(len(labels_tt)))
    ax_t.set_yticklabels(labels_tt, fontsize=8)
    ax_t.set_xlabel("Δ Margine medio post−pre (EUR/litro)", fontsize=10)
    ax_t.set_title("Welch t-test sul margine lordo\n(IC 95%  |  Rosso = p<0.05)",
                   fontsize=11, fontweight="bold")
    for i, (bar, row) in enumerate(zip(bars_tt, df_tt.itertuples())):
        ax_t.text(max(deltas_tt[i], 0) + 0.001, i,
                  _stars(row.p_value), va="center", fontsize=9,
                  color="#8b1a1a" if row.p_value < ALPHA else "#777")

# Pannello B: Welch — Cohen's d (effect size)
ax_d = axes9[0, 1]
if ttest_rows:
    ds     = df_tt["cohens_d"].values
    colors_d = ["#e74c3c" if p < ALPHA else "#95a5a6"
                for p in df_tt["p_value"].values]
    ax_d.barh(range(len(labels_tt)), ds,
              color=colors_d, edgecolor="black", lw=0.5, alpha=0.85)
    ax_d.axvline(0.2, color="gray",     lw=1.2, ls=":", label="piccolo (0.2)")
    ax_d.axvline(0.5, color="orange",   lw=1.2, ls=":", label="medio (0.5)")
    ax_d.axvline(0.8, color="#e74c3c",  lw=1.2, ls=":", label="grande (0.8)")
    ax_d.set_yticks(range(len(labels_tt)))
    ax_d.set_yticklabels(labels_tt, fontsize=8)
    ax_d.set_xlabel("Cohen's d (effect size)", fontsize=10)
    ax_d.set_title("Effect size Welch t-test\n(Cohen's d — soglie standard)",
                   fontsize=11, fontweight="bold")
    ax_d.legend(fontsize=8, loc="lower right")

# Pannello C–D: DiD δ con IC 95% (se disponibile)
if did_rows:
    df_did = pd.DataFrame(did_rows)
    # Label: "Evento\nPaese_ctrl\nCarburante"
    labels_did  = [f"{r['Evento'].split('(')[0].strip()}\nvs {r['Paese_controllo']}\n{r['Carburante']}"
                   for _, r in df_did.iterrows()]
    deltas_did  = df_did["delta_DiD"].values
    ci_lo_did_a = df_did["delta_DiD"].values - df_did["CI_95_low"].values
    ci_hi_did_a = df_did["CI_95_high"].values - df_did["delta_DiD"].values
    colors_did  = ["#e74c3c" if p < ALPHA else "#95a5a6"
                   for p in df_did["p_value"].values]

    ax_did1 = axes9[1, 0]
    bars_did = ax_did1.barh(range(len(labels_did)), deltas_did,
                             color=colors_did, edgecolor="black", lw=0.5, alpha=0.85)
    for i, (lo, hi) in enumerate(zip(ci_lo_did_a, ci_hi_did_a)):
        ax_did1.errorbar(deltas_did[i], i,
                         xerr=[[lo], [hi]],
                         fmt="none", color="black", capsize=5, lw=1.5)
    ax_did1.axvline(0, color="black", lw=1.0, ls="--")
    ax_did1.set_yticks(range(len(labels_did)))
    ax_did1.set_yticklabels(labels_did, fontsize=7)
    ax_did1.set_xlabel("δ DiD (EUR/litro, IC 95% HC3)", fontsize=10)
    ax_did1.set_title(
        "DiD δ: variazione margine IT vs paesi controllo\n(Rosso = p<0.05  |  SE robusti HC3)",
        fontsize=11, fontweight="bold")

    ax_did2 = axes9[1, 1]
    # Time-series: margine IT vs tutti i paesi controllo — primo evento, Benzina
    ev0       = list(EVENTS.keys())[0]
    cfg0      = EVENTS[ev0]
    margin_col_plot = "margine_benzina"

    # Palette colori per i paesi di controllo
    ctrl_colors = ["#2980b9", "#27ae60", "#8e44ad", "#e67e22"]
    if margin_col_plot in merged.columns:
        it_s = merged.loc[cfg0["pre_start"]:cfg0["post_end"], margin_col_plot].dropna()
        ax_did2.plot(it_s.index, it_s.values, color="#d6604d", lw=2.0, label="IT Benzina")

    for (paese, ctrl_pump), col in zip(CONTROL_PUMPS.items(), ctrl_colors):
        if margin_col_plot in ctrl_pump.columns:
            ct_s = ctrl_pump.loc[cfg0["pre_start"]:cfg0["post_end"], margin_col_plot].dropna()
            ax_did2.plot(ct_s.index, ct_s.values, color=col, lw=1.5,
                         ls="--", alpha=0.8, label=f"{paese[:3]} Benzina")

    ax_did2.axvline(cfg0["shock"], color="black", lw=1.5, ls="--")
    ylim_top = ax_did2.get_ylim()[1]
    ax_did2.text(cfg0["shock"] + pd.Timedelta(days=5),
                 ylim_top if ylim_top != 0 else 0.05,
                 ev0.split("(")[0].strip(), rotation=90, fontsize=8, color="black", va="top")
    ax_did2.set_ylabel("Margine lordo (EUR/litro)", fontsize=10)
    ax_did2.set_title(
        f"Margine IT vs paesi controllo — {ev0}\n(continuo = IT  |  tratteggio = controllo)",
        fontsize=11, fontweight="bold")
    ax_did2.legend(fontsize=8, loc="upper left")
    ax_did2.xaxis.set_major_formatter(mdates.DateFormatter("%b %y"))
    ax_did2.xaxis.set_major_locator(mdates.MonthLocator(interval=2))
    plt.setp(ax_did2.xaxis.get_majorticklabels(), rotation=45)

plt.tight_layout(pad=2.0)
plt.savefig("plots/09_ttest_did.png", dpi=DPI, bbox_inches="tight")
plt.close()
print("  Salvato: plots/09_ttest_did.png")


print("\nGenerazione figura riassuntiva §3-§5...")

fig, axes = plt.subplots(2, 3, figsize=(18, 11))
fig.suptitle("Test statistici di supporto — Anomalia margine carburanti Italia (evidenza ausiliaria)",
             fontsize=13, fontweight="bold")

# Pannello 1: CCF
ax = axes[0, 0]
lags_x = list(range(0, MAX_LAG_CCF + 1))
for (fuel_name, ccf_vals), color in zip(ccf_results.items(), ["#e74c3c", "#3498db"]):
    ax.plot(lags_x, ccf_vals, marker="o", color=color, lw=2, label=fuel_name)
ax.axvline(4, color="orange", lw=2, ls="--", label="Soglia 30gg (4 sett.)")
ax.axhline(0, color="black", lw=0.5)
ax.set_xlabel("Lag (settimane)")
ax.set_ylabel("Correlazione")
ax.set_title("Cross-Correlation: Brent → Prezzi Pompa")
ax.legend(fontsize=9)
ax.set_xticks(lags_x)

# Pannello 2: Rolling Correlation
ax = axes[0, 1]
for (fuel_name, rc), color in zip(rolling_corr.items(), ["#e74c3c", "#27ae60"]):
    ax.plot(rc.index, rc.values, color=color, lw=1.5, label=fuel_name)
for ev, ts in WAR_DATES.items():
    if merged.index[0] <= ts <= merged.index[-1]:
        ax.axvline(ts, color="gray", lw=1.2, ls="--")
        ax.text(ts, 0.05, ev.split(" ")[0], rotation=90, fontsize=7,
                color="gray", va="bottom")
ax.set_ylabel(f"Correlazione (rolling {ROLL_WIN} sett.)")
ax.set_title(f"Correlazione mobile Brent–Carburante")
ax.legend(fontsize=9)
ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %y"))
ax.xaxis.set_major_locator(mdates.MonthLocator(interval=6))
plt.setp(ax.xaxis.get_majorticklabels(), rotation=45)

# Pannello 3: KS ECDF (Ucraina)
ax = axes[0, 2]
ev_name_ks = "Ucraina (Feb 2022)"
if ev_name_ks in EVENTS:
    cfg_k = EVENTS[ev_name_ks]
    for (fuel_name, fuel_col), color in zip(FUELS.items(), ["#e74c3c", "#3498db"]):
        if fuel_col not in merged.columns:
            continue
        pre  = merged.loc[cfg_k["pre_start"]:cfg_k["shock"], fuel_col].dropna().values
        post = merged.loc[cfg_k["shock"]:cfg_k["post_end"], fuel_col].dropna().values
        ax.step(np.sort(pre),  np.linspace(0, 1, len(pre)),  color=color, lw=2,
                ls="--", label=f"{fuel_name} pre")
        ax.step(np.sort(post), np.linspace(0, 1, len(post)), color=color, lw=2,
                label=f"{fuel_name} post")
ax.set_xlabel("Prezzo (EUR/litro)")
ax.set_ylabel("ECDF")
ax.set_title(f"KS Test ECDF\n({ev_name_ks})")
ax.legend(fontsize=8)

# Pannello 4: ANOVA F-stat
ax = axes[1, 0]
if anova_rows:
    df_an = pd.DataFrame(anova_rows)
    labels_an = [f"{r['Evento'].split('(')[0].strip()}\n{r['Carburante']}"
                 for _, r in df_an.iterrows()]
    f_vals_an = df_an["F_stat"].values
    colors_an = ["#e74c3c" if p < ALPHA else "#95a5a6"
                 for p in df_an["p_value"].values]
    ax.barh(range(len(labels_an)), f_vals_an, color=colors_an, edgecolor="black", lw=0.5)
    ax.set_yticks(range(len(labels_an)))
    ax.set_yticklabels(labels_an, fontsize=8)
    ax.set_xlabel("F-statistic")
    ax.set_title("ANOVA F-test (3 periodi: pre/shock/post)\nRosso = H0 rifiutata (p<0.05)")

# Pannello 5: Chow test
ax = axes[1, 1]
if chow_rows:
    df_ch = pd.DataFrame(chow_rows)
    labels_ch = [f"{r['Evento'].split('(')[0].strip()}\n{r['Carburante']}"
                 for _, r in df_ch.iterrows()]
    p_vals_ch = [float(p) if p != "N/A" else 1.0 for p in df_ch["p_value"].values]
    colors_ch = ["#e74c3c" if p < ALPHA else "#95a5a6" for p in p_vals_ch]
    ax.barh(range(len(labels_ch)), p_vals_ch, color=colors_ch, edgecolor="black", lw=0.5)
    ax.axvline(ALPHA, color="black", lw=1.5, ls="--", label=f"α={ALPHA}")
    ax.set_yticks(range(len(labels_ch)))
    ax.set_yticklabels(labels_ch, fontsize=8)
    ax.set_xlabel("p-value")
    ax.set_title("Chow Test structural break\nRosso = H0 rifiutata (p<0.05)")
    ax.legend(fontsize=9)

# Pannello 6: Bootstrap CI lag D
ax = axes[1, 2]
if bootstrap_results:
    df_boot = pd.DataFrame(bootstrap_results)
    labels_b = [f"{r['Evento'].split('(')[0].strip()}\n{r['Carburante']}"
                for _, r in df_boot.iterrows()]
    means_b = df_boot["Lag_mean"].values
    ci_lo_b = df_boot["CI_95_low"].values
    ci_hi_b = df_boot["CI_95_high"].values
    colors_b = ["#e74c3c" if h < 30 else "#3498db" for h in ci_hi_b]
    ax.barh(range(len(labels_b)), means_b, color=colors_b, alpha=0.7,
            edgecolor="black", lw=0.5)
    for i, (lo, hi) in enumerate(zip(ci_lo_b, ci_hi_b)):
        ax.errorbar(means_b[i], i, xerr=[[means_b[i] - lo], [hi - means_b[i]]],
                    fmt="none", color="black", capsize=5, lw=2)
    ax.axvline(30, color="orange", lw=2, ls="--", label="Soglia H0 (30gg)")
    ax.set_yticks(range(len(labels_b)))
    ax.set_yticklabels(labels_b, fontsize=8)
    ax.set_xlabel("Lag D (giorni)")
    ax.set_title("Bootstrap 95% CI sul Lag D\nRosso = CI tutto < 30gg")
    ax.legend(fontsize=9)

plt.tight_layout(pad=2.0)
plt.savefig("plots/06_statistical_tests.png", dpi=DPI, bbox_inches="tight")
plt.close()
print("  Salvato: plots/06_statistical_tests.png")

# ─────────────────────────────────────────────────────────────────────────────
# SOMMARIO
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 65)
print("SOMMARIO — Script 03")
print("=" * 65)
print("  §1 Granger:  data/granger_{benzina,diesel}.csv")
print("               plots/03_granger_combined.png")
print("  §2 R&F:      data/rockets_feathers_results.csv")
print("               plots/04_rf_combined.png")
print("  §3 KS/ANOVA/Chow: data/{ks,anova,chow}_results.csv")
print("  §4 CCF:      (nel plot riassuntivo)")
print("  §5 Bootstrap: data/bootstrap_ci.csv")
print("  §6 Sel. reg.: data/regression_selection.csv")
print("  §7 t-test:   data/ttest_margin.csv")
print("  §8 DiD:      data/did_results.csv (se dati DE disponibili)")
print("  Figure:      plots/06_statistical_tests.png")
print("               plots/09_ttest_did.png")
print("\n  NOTE:")
print("  - Granger, R&F, t-test, DiD sono EVIDENZE AUSILIARIE")
print("  - H0 (anomalia margine) è testata in 02_core_analysis.py §D")
print("  - BH correction FDR 5% applicata in script 02 sui test primari")
print("  - DiD richiede data/eu_oil_bulletin_history.xlsx (da script 01)")

plt.rcParams.update(plt.rcParamsDefault)
print("\nScript 03 completato.")
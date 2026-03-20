"""
3_01_data.py  — Pipeline dati completa (v3)
============================================
Scarica e prepara tutte le serie settimanali necessarie per testare H₀.

FONTI DATI
──────────
  Brent (EUR/bbl) + EUR/USD     ← yfinance  BZ=F  e  EURUSD=X
  Prezzi pompa IT, DE, SE       ← EU Weekly Oil Bulletin  (download automatico da URL)
  Futures Eurobob ARA benzina   ← CSV pre-scaricato da Investing.com  ← MANUALE
  Futures London Gas Oil diesel ← CSV pre-scaricato da Investing.com  ← MANUALE
  HICP Italy (2015=100)         ← API ECB SDMX (fallback: tabella annuale)

  NOTA Eurobob / Gas Oil: non esistono API pubbliche gratuite. Scaricare manualmente
  da Investing.com e salvare come:
      data/Eurobob Futures Historical Data.csv
      data/London Gas Oil Futures Historical Data.csv

STRATEGIA CACHE
───────────────
  Se un file scaricato esiste già su disco (es. eu_oil_bulletin_history.xlsx,
  brent_weekly_eur.csv) viene riutilizzato senza re-download. Cancellarlo per
  forzare il refresh.

OUTPUT
──────
  data/3_dataset.csv        — dataset principale IT (crack spread nom. + reali)
  data/3_hicp.csv           — HICP mensile per deflazione
  data/pompa_{xx}.csv       — prezzi pompa paesi controllo (per DiD in 3_03)
                              xx ∈ {de, se, nl, be, dk, fi, at}
  plots/3_01a_brent.png     — Brent EUR/barile nel tempo
  plots/3_01b_pompa_it.png  — prezzi pompa IT benzina e diesel
  plots/3_01c_crack.png     — crack spread nominale vs reale (HICP-deflato)
  plots/3_01d_confronto.png — prezzi pompa IT vs DE vs SE a confronto
"""

import os
import warnings
from datetime import date, timedelta

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import requests
import openpyxl as _opxl

warnings.filterwarnings("ignore")
os.makedirs("data",  exist_ok=True)
os.makedirs("plots", exist_ok=True)

# ── Costanti ──────────────────────────────────────────────────────────────────
DATA_START = "2019-01-01"
TODAY_END  = (date.today() + timedelta(days=7)).strftime("%Y-%m-%d")
DPI        = 160

DENSITY_BENZ = 0.74;  L_PER_T_BENZ = 1000.0 / DENSITY_BENZ   # ≈1351 L/t
DENSITY_DIES = 0.84;  L_PER_T_DIES = 1000.0 / DENSITY_DIES   # ≈1190 L/t

# Linee verticali negli eventi sui grafici
EVENTS_LINES = {
    "Russia-Ucraina\n(24 feb 2022)":  ("2022-02-24", "#e74c3c"),
    "Iran-Israele\n(13 giu 2025)":    ("2025-06-13", "#e67e22"),
}

# HICP Italy annuale (2015=100) — fallback se API ECB non raggiungibile
# Fonte: Eurostat prc_hicp_aind
HICP_FALLBACK = {
    2019: 103.3, 2020: 102.0, 2021: 104.7, 2022: 113.2,
    2023: 121.5, 2024: 124.6, 2025: 126.8, 2026: 128.5,
}

# Paesi da estrarre per il DiD.
# IT = paese trattamento; tutti gli altri = controlli.
#
# INCLUSI (nessun intervento significativo sui prezzi ex-tasse):
#   DE - Germania     : no price cap; Tankrabatt giu–ago 2022 era sussidio fiscale
#                       che NON altera il margine ex-tasse → incluso
#   NL - Paesi Bassi  : mercato libero, nessun intervento ex-tasse
#   BE - Belgio       : meccanismo max-price solo su retail inc. tasse
#   DK - Danimarca    : mercato libero
#   FI - Finlandia    : mercato libero
#   AT - Austria      : nessun price cap ex-tasse
#
# ESCLUSI (interventi distorsivi sui margini ex-tasse):
#   FR - Francia      : ristorno fiscale ~0.15–0.18 €/L (set–dic 2022) →
#                       comprime artificialmente il margine FR osservato
#   HU - Ungheria     : price cap rigido su prezzi ex-tasse (nov 2021 – lug 2022)
#   ES - Spagna       : sussidio 0.20 €/L (apr–dic 2022) → distorce margini
#   PL - Polonia      : riduzione IVA 2022 → altera struttura costi
#   SE - Svezia: ESCLUSA per diesel (reduktionsplikt 30.5% → 6% gen 2024
#                        gonfia artificialmente il margine SE 2022-2023 e crea
#                        una rottura strutturale a gen 2024 non geopolitica).
#                        Usabile per benzina (mandato 7.8%, effetto < 2 ct/L).
COUNTRY_PREFIXES = {
    "IT": ["IT_"],
    "DE": ["DE_"],
    "NL": ["NL_"],
    "BE": ["BE_"],
    "DK": ["DK_"],
    "FI": ["FI_"],
    "AT": ["AT_"],
}
FUEL_KW_BENZ = ["95", "benz", "petrol", "unleaded", "super", "euro_", "euro95"]
FUEL_KW_DIES = ["diesel", "gasoil", "gas_oil"]


# ════════════════════════════════════════════════════════════════════════════
# HELPERS
# ════════════════════════════════════════════════════════════════════════════

def _event_lines(ax, y_max, idx):
    """Disegna linee verticali per gli eventi geopolitici."""
    for label, (dt, color) in EVENTS_LINES.items():
        ts = pd.Timestamp(dt)
        if len(idx) and idx[0] <= ts <= idx[-1]:
            ax.axvline(ts, color=color, lw=1.6, ls="--", alpha=0.85)
            ax.text(ts + pd.Timedelta(days=6), y_max * 0.96,
                    label, rotation=90, fontsize=7, color=color, va="top")


def _wb_sheets(path):
    wb = _opxl.load_workbook(path, read_only=True, data_only=True)
    ns = wb.sheetnames; wb.close(); return ns


def _wb_read(path, sheet):
    wb = _opxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if not rows:
        return pd.DataFrame()
    hdr_i = next((i for i, r in enumerate(rows) if sum(v is not None for v in r) >= 3), 0)
    hdrs  = [str(h).strip() if h is not None else f"_c{j}"
             for j, h in enumerate(rows[hdr_i])]
    df    = pd.DataFrame(rows[hdr_i + 1:], columns=hdrs)
    ic    = hdrs[0]
    df[ic] = pd.to_datetime(df[ic], errors="coerce")
    return df.set_index(ic).loc[lambda x: x.index.notna()].sort_index()


def _notax_sheet(names):
    for s in names:
        if any(k in s.upper() for k in ["WO TAX", "WITHOUT", "NO TAX", "NOTAX", "WO TAXES"]):
            return s
    return names[1] if len(names) > 1 else names[0]


def _cols_for_country(df, prefixes, fuel_benz, fuel_dies):
    """Trova colonne benzina/diesel per un paese usando solo startswith (no substring)."""
    all_c  = [c for c in df.columns
               if any(str(c).upper().startswith(p) for p in prefixes)]
    b_cols = [c for c in all_c if any(k in str(c).lower() for k in fuel_benz)]
    d_cols = [c for c in all_c if any(k in str(c).lower() for k in fuel_dies)]
    if not b_cols and all_c:          b_cols = [all_c[0]]
    if not d_cols and len(all_c) > 1: d_cols = [all_c[1]]
    return b_cols, d_cols


def _load_investing_csv(path, raw_col="Price"):
    """Carica CSV da Investing.com (formato standard con colonna Price)."""
    df = pd.read_csv(path, thousands=",")
    df.columns = [c.lstrip("﻿") for c in df.columns]
    df["Date"] = pd.to_datetime(df["Date"], format="%m/%d/%Y", errors="coerce")
    df = df.dropna(subset=["Date"]).set_index("Date").sort_index()
    if raw_col in df.columns:
        df[raw_col] = pd.to_numeric(
            df[raw_col].astype(str).str.replace(",", ""), errors="coerce")
    return df


# ════════════════════════════════════════════════════════════════════════════
# 1. EUR/USD e Brent
# ════════════════════════════════════════════════════════════════════════════
print("1. Scarico EUR/USD e Brent (yfinance)...")
brent_csv = "data/brent_weekly_eur.csv"

if os.path.exists(brent_csv):
    # Riusa file esistente (evita re-download in ambienti senza internet)
    brent_w = pd.read_csv(brent_csv, index_col=0, parse_dates=True)
    print(f"   Caricato da cache: {brent_csv}  ({len(brent_w)} settimane)")
else:
    # Download fresco
    try:
        import yfinance as yf
        raw_fx  = yf.download("EURUSD=X", start=DATA_START, end=TODAY_END, progress=False)
        raw_brt = yf.download("BZ=F",     start=DATA_START, end=TODAY_END, progress=False)
        if raw_fx.empty or raw_brt.empty:
            raise ValueError("risposta yfinance vuota")
        eurusd_w = raw_fx["Close"].squeeze().rename("eurusd").resample("W-MON").mean()
        brent_w  = raw_brt["Close"].squeeze().rename("brent_usd").resample("W-MON").mean().to_frame()
        brent_w  = brent_w.join(eurusd_w, how="left")
        brent_w["eurusd"]   = brent_w["eurusd"].ffill()
        brent_w["brent_eur"] = brent_w["brent_usd"] / brent_w["eurusd"]
        brent_w["brent_eur"] = brent_w["brent_eur"].interpolate(method="time", limit=2)
        brent_w.to_csv(brent_csv)
        print(f"   Scaricato: {len(brent_w)} settimane")
    except Exception as exc:
        raise SystemExit(f"Brent non disponibile e cache mancante: {exc}")


# ════════════════════════════════════════════════════════════════════════════
# 2. EU Weekly Oil Bulletin  (download da URL ufficiale DG Energy)
# ════════════════════════════════════════════════════════════════════════════
print("2. EU Weekly Oil Bulletin (download da URL o cache)...")

EU_SOURCES = [
    # URL principale — storico 2005-presente, aggiornato settimanalmente
    ("storico 2005-presente",
     "https://energy.ec.europa.eu/document/download/"
     "906e60ca-8b6a-44e7-8589-652854d2fd3f_en?"
     "filename=Weekly_Oil_Bulletin_Prices_History_maticni_4web.xlsx",
     "data/eu_oil_bulletin_history.xlsx"),
    # Fallback URL alternativo
    ("settimanale senza tasse (fallback)",
     "https://energy.ec.europa.eu/document/download/"
     "78311f92-68f8-4b82-b5cf-1293beeaae77_en?"
     "filename=Weekly+Oil+Bulletin+Weekly+prices+without+taxes.xlsx",
     "data/eu_oil_bulletin_notax.xlsx"),
]

df_eu_raw  = None
eu_xl_path = None

for label, url, fname in EU_SOURCES:
    try:
        if os.path.exists(fname) and os.path.getsize(fname) > 10_000:
            print(f"   Uso cache: {fname}")
        else:
            print(f"   Download: {label}...")
            resp = requests.get(url, timeout=90, headers={"User-Agent": "Mozilla/5.0"})
            resp.raise_for_status()
            cnt = resp.content
            if not (cnt[:2] == b"PK" or cnt[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"):
                raise ValueError("risposta non-Excel")
            with open(fname, "wb") as fh:
                fh.write(cnt)
            print(f"   Salvato: {fname}")

        sheets    = _wb_sheets(fname)
        sheet     = _notax_sheet(sheets)
        df_eu_raw = _wb_read(fname, sheet).apply(pd.to_numeric, errors="coerce")
        eu_xl_path = fname
        print(f"   OK — foglio '{sheet}', {len(df_eu_raw.columns)} col, {len(df_eu_raw)} righe")
        break
    except Exception as exc:
        print(f"   Fallito ({label}): {exc}")

if df_eu_raw is None:
    raise SystemExit("EU Oil Bulletin non disponibile. Verifica connessione o file cache.")


# ════════════════════════════════════════════════════════════════════════════
# 3. Estrai prezzi pompa per IT, DE, SE
# ════════════════════════════════════════════════════════════════════════════
print("3. Estraggo prezzi pompa per IT, DE, SE...")
pompa_dict: dict[str, pd.DataFrame] = {}

for country, prefixes in COUNTRY_PREFIXES.items():
    b_cols, d_cols = _cols_for_country(df_eu_raw, prefixes, FUEL_KW_BENZ, FUEL_KW_DIES)
    if not b_cols or not d_cols:
        print(f"   {country}: colonne non trovate — skip")
        continue
    p = pd.concat([df_eu_raw[b_cols[0]].rename("benzina_eur_l"),
                   df_eu_raw[d_cols[0]].rename("diesel_eur_l")], axis=1)
    p = p[p.index >= DATA_START].dropna(how="all").resample("W-MON").mean()
    # Normalizza EUR/1000L → EUR/L (EU Bulletin pubblica in EUR/1000L)
    for col in ["benzina_eur_l", "diesel_eur_l"]:
        if p[col].dropna().median() > 10:
            p[col] /= 1000.0
    # Interpola gap brevi (max 2 settimane)
    full_idx = pd.date_range(p.index[0], p.index[-1], freq="W-MON")
    p = p.reindex(full_idx).interpolate(method="time", limit=2)
    pompa_dict[country] = p
    print(f"   {country}: {p.dropna(how='all').shape[0]} settimane  "
          f"[{b_cols[0]}, {d_cols[0]}]")

if "IT" not in pompa_dict:
    raise SystemExit("Prezzi pompa Italia non trovati nel Bulletin.")

pompa_it = pompa_dict["IT"]
pompa_it.to_csv("data/prezzi_pompa_italia.csv")

# Salva tutti i paesi controllo (non-IT) per uso in 3_03_did.py
for country in COUNTRY_PREFIXES:
    if country == "IT":
        continue
    if country in pompa_dict:
        pompa_dict[country].to_csv(f"data/pompa_{country.lower()}.csv")
        print(f"   Salvato data/pompa_{country.lower()}.csv")


# ════════════════════════════════════════════════════════════════════════════
# 4. Merge Brent + pompa IT
# ════════════════════════════════════════════════════════════════════════════
print("4. Merge Brent + pompa IT...")
merged = brent_w.reindex(pompa_it.index).join(pompa_it)
for col in ["brent_eur", "eurusd"]:
    if col in merged.columns:
        merged[col] = merged[col].ffill(limit=4)
merged = merged.dropna(subset=["benzina_eur_l", "diesel_eur_l"])
merged.to_csv("data/dataset_merged.csv")
print(f"   {len(merged)} settimane | {merged.index[0].date()} – {merged.index[-1].date()}")


# ════════════════════════════════════════════════════════════════════════════
# 5. Futures wholesale (CSV pre-scaricati da Investing.com — manuale)
# ════════════════════════════════════════════════════════════════════════════
print("5. Carico futures wholesale da CSV (Investing.com)...")
merged_f  = merged.copy()
eurusd_al = merged_f["eurusd"].ffill().bfill() if "eurusd" in merged_f.columns \
            else pd.Series(1.1, index=merged_f.index)

for csv_path, raw_col, eur_col, l_per_t in [
    ("data/Eurobob Futures Historical Data.csv",
     "eurobob_usd_tonne", "eurobob_eur_l", L_PER_T_BENZ),
    ("data/London Gas Oil Futures Historical Data.csv",
     "gasoil_usd_tonne",  "gasoil_eur_l",  L_PER_T_DIES),
]:
    if not os.path.exists(csv_path):
        print(f"   SKIP — non trovato: {csv_path}")
        continue
    try:
        df_f = _load_investing_csv(csv_path)
        df_f.rename(columns={"Price": raw_col}, inplace=True)
        wk = df_f[raw_col].resample("W-MON").mean().ffill()
        merged_f = merged_f.join(wk.rename(raw_col), how="left")
        merged_f[eur_col] = (merged_f[raw_col] / eurusd_al) / l_per_t
        print(f"   {raw_col}: mediana {merged_f[eur_col].median():.4f} EUR/L")
    except Exception as exc:
        print(f"   {raw_col}: ERRORE — {exc}")

# Crack spread = prezzo pompa − prezzo wholesale
if "eurobob_eur_l" in merged_f.columns:
    merged_f["margine_benz_crack"] = merged_f["benzina_eur_l"] - merged_f["eurobob_eur_l"]
if "gasoil_eur_l" in merged_f.columns:
    merged_f["margine_dies_crack"] = merged_f["diesel_eur_l"]  - merged_f["gasoil_eur_l"]

merged_f.to_csv("data/dataset_merged_with_futures.csv")
print(f"   Dataset con futures: {len(merged_f)} settimane")


# ════════════════════════════════════════════════════════════════════════════
# 6. HICP Italy — deflatore per margini reali
# ════════════════════════════════════════════════════════════════════════════
print("6. HICP Italy (API ECB o fallback tabella annuale)...")


def _download_hicp_ecb() -> pd.Series | None:
    """HICP mensile IT dall'API ECB SDMX 2.1 (indice 2015=100)."""
    url = ("https://data-api.ecb.europa.eu/service/data/ICP/"
           "M.IT.N.000000.4.INX?format=csvdata&detail=dataonly")
    try:
        r = requests.get(url, timeout=20)
        r.raise_for_status()
        from io import StringIO
        df = pd.read_csv(StringIO(r.text))
        val_col  = next((c for c in df.columns if c.upper() in ("OBS_VALUE", "VALUE")), None)
        date_col = next((c for c in df.columns if c.upper() in ("TIME_PERIOD", "DATE")), None)
        if not val_col or not date_col:
            return None
        s = pd.to_numeric(df.set_index(date_col)[val_col], errors="coerce").dropna()
        s.index = pd.to_datetime(s.index, format="%Y-%m", errors="coerce")
        s = s.dropna().sort_index()
        s.name = "hicp"
        print(f"   ECB: {len(s)} mesi  ({s.index[0].date()} – {s.index[-1].date()})")
        return s
    except Exception as exc:
        print(f"   ECB fallback: {exc}")
        return None


def _hicp_fallback() -> pd.Series:
    """Costruisce HICP mensile da tabella annuale per interpolazione lineare."""
    pts = [(pd.Timestamp(f"{y}-07-01"), v) for y, v in HICP_FALLBACK.items()]
    s   = pd.Series(dict(pts), name="hicp").sort_index()
    idx = pd.date_range(f"{min(HICP_FALLBACK)}-01-01",
                        f"{max(HICP_FALLBACK)}-12-01", freq="MS")
    s   = s.reindex(s.index.union(idx)).interpolate("time").reindex(idx)
    s.name = "hicp"
    print(f"   Fallback tabella annuale: {len(s)} mesi interpolati")
    return s


hicp_m = _download_hicp_ecb()
if hicp_m is None:
    hicp_m = _hicp_fallback()
hicp_m.to_frame().to_csv("data/3_hicp.csv")

hicp_2019_avg = float(hicp_m[hicp_m.index.year == 2019].mean())
print(f"   HICP media 2019 (base deflazione): {hicp_2019_avg:.2f}")

# Allinea HICP a frequenza settimanale
hicp_w = hicp_m.resample("W-MON").first().ffill()

# Aggiungi deflatore e margini reali al dataset
merged_f = merged_f.join(hicp_w.rename("hicp"), how="left")
merged_f["hicp"]       = merged_f["hicp"].ffill().bfill()
merged_f["deflatore"]  = merged_f["hicp"] / hicp_2019_avg
for nom, real in [("margine_benz_crack", "margine_benz_real"),
                  ("margine_dies_crack", "margine_dies_real")]:
    if nom in merged_f.columns:
        merged_f[real] = merged_f[nom] / merged_f["deflatore"]

merged_f.to_csv("data/3_dataset.csv")
print(f"   ✓ data/3_dataset.csv  ({len(merged_f)} settimane)")


# ════════════════════════════════════════════════════════════════════════════
# 7. GRAFICI
# ════════════════════════════════════════════════════════════════════════════
print("7. Genero grafici...")
df = merged_f[merged_f.index >= DATA_START].copy()


def _fmt_xaxis(ax, interval=4):
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %Y"))
    ax.xaxis.set_major_locator(mdates.MonthLocator(interval=interval))
    plt.setp(ax.get_xticklabels(), rotation=35, fontsize=8)


# ── 7a. Brent EUR/barile ─────────────────────────────────────────────────
fig, ax = plt.subplots(figsize=(14, 4))
ax.plot(df.index, df["brent_eur"], color="#2166ac", lw=1.8)
_event_lines(ax, float(df["brent_eur"].max()), df.index)
ax.set_title("Brent Crude Oil 2019– (EUR/barile)", fontweight="bold")
ax.set_ylabel("EUR/barile");  ax.grid(alpha=0.25)
_fmt_xaxis(ax)
plt.tight_layout()
fig.savefig("plots/3_01a_brent.png", dpi=DPI, bbox_inches="tight")
plt.close(fig)

# ── 7b. Prezzi pompa IT benzina e diesel ─────────────────────────────────
fig, axes = plt.subplots(2, 1, figsize=(14, 7), sharex=True)
for ax, col, color, title in [
    (axes[0], "benzina_eur_l", "#d6604d", "Benzina Italia senza tasse (EUR/L)"),
    (axes[1], "diesel_eur_l",  "#4393c3", "Diesel Italia senza tasse (EUR/L)"),
]:
    ax.plot(df.index, df[col], color=color, lw=1.8)
    _event_lines(ax, float(df[col].max()), df.index)
    ax.set_ylabel("EUR/L");  ax.set_title(title, fontweight="bold");  ax.grid(alpha=0.25)
_fmt_xaxis(axes[1])
fig.suptitle("Prezzi pompa Italia senza tasse — EU Oil Bulletin", fontsize=12, fontweight="bold")
plt.tight_layout()
fig.savefig("plots/3_01b_pompa_it.png", dpi=DPI, bbox_inches="tight")
plt.close(fig)

# ── 7c. Crack spread nominale e reale (HICP-deflato) ─────────────────────
if "margine_benz_crack" in df.columns:
    fig, axes = plt.subplots(2, 1, figsize=(14, 8), sharex=True)
    for ax, col_nom, col_real, color, fuel in [
        (axes[0], "margine_benz_crack", "margine_benz_real", "#d6604d", "Benzina"),
        (axes[1], "margine_dies_crack", "margine_dies_real", "#4393c3", "Diesel"),
    ]:
        if col_nom not in df.columns:
            continue
        bl = df.loc["2019", col_nom].dropna()
        ax.plot(df.index, df[col_nom], color=color, lw=1.8, label="Nominale")
        if col_real in df.columns:
            ax.plot(df.index, df[col_real], color=color, lw=1.2,
                    ls="--", alpha=0.55, label="Reale (HICP-deflato)")
        if len(bl) >= 4:
            mu, sd = bl.mean(), bl.std()
            ax.axhline(mu, color="#555", lw=1.0, ls="-.", alpha=0.7)
            ax.axhspan(mu - 2*sd, mu + 2*sd, alpha=0.10, color="#888888",
                       label="Baseline 2019 ±2σ")
        _event_lines(ax, float(df[col_nom].max()), df.index)
        ax.set_ylabel("EUR/L");  ax.legend(fontsize=8, loc="upper left")
        ax.set_title(f"Crack spread {fuel}  (margine lordo = pompa − wholesale)",
                     fontweight="bold");  ax.grid(alpha=0.25)
    _fmt_xaxis(axes[1])
    fig.suptitle("Margine lordo crack spread — nominale vs HICP-deflato\n"
                 "La banda grigia mostra la normale variabilità 2019 (±2σ)",
                 fontsize=11, fontweight="bold")
    plt.tight_layout()
    fig.savefig("plots/3_01c_crack.png", dpi=DPI, bbox_inches="tight")
    plt.close(fig)
    print("   ✓ plots/3_01c_crack.png")

# ── 7d. Confronto prezzi pompa IT vs DE vs SE ─────────────────────────────
country_colors = {"IT": "#d62728", "DE": "#1f77b4", "SE": "#2ca02c"}
fig, axes = plt.subplots(2, 1, figsize=(14, 8), sharex=True)
for ax, col, title in [
    (axes[0], "benzina_eur_l", "Benzina senza tasse (EUR/L)"),
    (axes[1], "diesel_eur_l",  "Diesel senza tasse (EUR/L)"),
]:
    for country, df_p in pompa_dict.items():
        if col in df_p.columns:
            ax.plot(df_p.index, df_p[col], color=country_colors.get(country, "gray"),
                    lw=1.6, label=country, alpha=0.9)
    ymax = max((df_p[col].max() for df_p in pompa_dict.values() if col in df_p.columns),
               default=1.0)
    _event_lines(ax, float(ymax), pompa_it.index)
    ax.set_ylabel("EUR/L");  ax.set_title(title, fontweight="bold")
    ax.legend(fontsize=9);  ax.grid(alpha=0.25)
_fmt_xaxis(axes[1])
fig.suptitle("Confronto prezzi pompa senza tasse: Italia vs Germania vs Svezia\n"
             "(stesso fonte: EU Weekly Oil Bulletin)",
             fontsize=11, fontweight="bold")
plt.tight_layout()
fig.savefig("plots/3_01d_confronto.png", dpi=DPI, bbox_inches="tight")
plt.close(fig)

print("   ✓ plots/3_01a_brent.png")
print("   ✓ plots/3_01b_pompa_it.png")
print("   ✓ plots/3_01d_confronto.png")
print("\nScript 3_01 completato.")
